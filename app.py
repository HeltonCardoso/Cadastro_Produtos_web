from collections import defaultdict
from io import BytesIO
import sys
from pathlib import Path
from flask import Flask, abort, current_app, json, make_response, render_template, request, send_file, send_from_directory, redirect, url_for, flash, jsonify
from gspread import service_account
import gspread
import requests
from models import Processo, db, Usuario, Perfil, ItemProcessado
from config import Config
import os
from datetime import datetime, timedelta
import pandas as pd
from werkzeug.utils import secure_filename
from processamento.cadastro_produto_web import executar_processamento
from processamento.extrair_atributos import extrair_atributos_processamento
from processamento.api_anymarket import obter_token_anymarket_seguro
from processamento.comparar_prazos import processar_comparacao
from processamento.google_sheets import ler_planilha_google
from token_manager_secure import ml_token_manager
from mercadolivre_api_secure import ml_api_secure
from utils.stats_utils import get_processing_stats, obter_dados_grafico_7dias
import logging
from logging.handlers import RotatingFileHandler
from log_utils import (registrar_processo,registrar_itens_processados,obter_historico_processos,contar_processos_hoje)
from processamento.api_anymarket import (consultar_api_anymarket, excluir_foto_anymarket, excluir_fotos_planilha_anymarket)
from google_sheets_utils import (carregar_configuracao_google_sheets, salvar_configuracao_google_sheets,listar_abas_google_sheets,testar_conexao_google_sheets)
from routes_intelipost import intelipost_bp
from flask_caching import Cache, logger
#from metrics_api import metrics_bp
from processamento.google_sheets import ler_planilha_google
from google_auth import GoogleSheetsOAuth, GoogleTokenManager
from functools import wraps
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from ml_oauth import ml_oauth_bp






sys.path.append(str(Path(__file__).parent))

# ============================================
# CONFIGURAÇÃO DO APP
# ============================================

app = Flask(__name__, static_folder='static', static_url_path='/static')
app.config.from_object(Config)
app.secret_key = Config.SECRET_KEY

# 🔹 PRIMEIRO: Inicializa o banco de dados
#db.init_app(app)
db.init_app(app)
cache = Cache(app, config={
    'CACHE_TYPE': 'SimpleCache',
    'CACHE_DEFAULT_TIMEOUT': 60
})
# 🔹 SEGUNDO: Configura o Login Manager (DEPOIS do db.init_app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '⚠️ Faça login para acessar esta página'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    """Carrega o usuário da sessão"""
    from models import Usuario
    return Usuario.query.get(int(user_id))

# 🔹 TERCEIRO: Registra os blueprints
#  app.register_blueprint(metrics_bp)
#app.register_blueprint(intelipost_bp)

# ============================================
# DECORATORS DE PERMISSÃO (CORRIGIDOS)
# ============================================

from functools import wraps
from flask_login import current_user
from flask import redirect, url_for, flash  # ← IMPORTANTE: adicione estes imports!

def master_required(f):
    """Decorator para rotas que só Master pode acessar"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Faça login para continuar', 'warning')
            return redirect(url_for('login'))
        if not current_user.is_master():
            flash('Acesso negado. Apenas Master pode acessar.', 'danger')
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated_function

def permissao_modulo(modulo):
    """Decorator para verificar permissão de módulo"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('Faça login para continuar', 'warning')
                return redirect(url_for('login'))
            if not current_user.has_permission(modulo):
                flash(f'Acesso negado. Módulo {modulo} não disponível para seu perfil.', 'danger')
                return redirect(url_for('home'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# 🔹 SEGUNDO: Registra os blueprints (DEPOIS de init_app)
#app.register_blueprint(metrics_bp)
app.register_blueprint(intelipost_bp)
app.register_blueprint(ml_oauth_bp)
from routes_ml_dashboard import ml_dashboard_bp
app.register_blueprint(ml_dashboard_bp)

# Configuração de logs
handler = RotatingFileHandler('app.log', maxBytes=10000, backupCount=1)
handler.setLevel(logging.INFO)
app.logger.addHandler(handler)

# 🔹 TERCEIRO: Cria as tabelas dentro do contexto
with app.app_context():
    # Cria as pastas necessárias
    os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
    
    # Cria as tabelas do banco
    try:
        db.create_all()
        print("✅ Banco de dados inicializado com sucesso!")
        
        # ============================================
        # CRIA PERFIS PADRÃO
        # ============================================
        from models import Perfil, Usuario
        
        # Lista de perfis padrão
        perfis_padrao = [
            {'nome': 'Master', 'descricao': 'Acesso total ao sistema'},
            {'nome': 'SAC', 'descricao': 'Acesso a pedidos e clientes'},
            {'nome': 'Cadastro', 'descricao': 'Acesso a produtos e atributos'},
            {'nome': 'Financeiro', 'descricao': 'Acesso a relatórios financeiros'},
        ]
        
        for p in perfis_padrao:
            perfil = Perfil.query.filter_by(nome=p['nome']).first()
            if not perfil:
                perfil = Perfil(nome=p['nome'], descricao=p['descricao'])
                db.session.add(perfil)
                print(f"✅ Perfil criado: {p['nome']}")
        
        db.session.commit()
        
        # ============================================
        # CRIA USUÁRIOS PADRÃO
        # ============================================
        
        # Usuário Master
        perfil_master = Perfil.query.filter_by(nome='Master').first()
        if perfil_master:
            master = Usuario.query.filter_by(username='master').first()
            if not master:
                master = Usuario(
                    username='master',
                    email='master@sistema.com',
                    perfil_id=perfil_master.id,
                    is_active=True
                )
                master.set_password('master123')
                db.session.add(master)
                print("✅ Usuário Master criado: master / master123")
        
        # Usuário SAC
        perfil_sac = Perfil.query.filter_by(nome='SAC').first()
        if perfil_sac:
            sac = Usuario.query.filter_by(username='sac').first()
            if not sac:
                sac = Usuario(
                    username='sac',
                    email='sac@sistema.com',
                    perfil_id=perfil_sac.id,
                    is_active=True
                )
                sac.set_password('sac123')
                db.session.add(sac)
                print("✅ Usuário SAC criado: sac / sac123")
        
        # Usuário Cadastro
        perfil_cadastro = Perfil.query.filter_by(nome='Cadastro').first()
        if perfil_cadastro:
            cadastro = Usuario.query.filter_by(username='cadastro').first()
            if not cadastro:
                cadastro = Usuario(
                    username='cadastro',
                    email='cadastro@sistema.com',
                    perfil_id=perfil_cadastro.id,
                    is_active=True
                )
                cadastro.set_password('cadastro123')
                db.session.add(cadastro)
                print("✅ Usuário Cadastro criado: cadastro / cadastro123")
        
        db.session.commit()
        
        print("\n" + "="*50)
        print("🎉 SISTEMA INICIALIZADO COM SUCESSO!")
        print("="*50)
        print("📝 PERFIS DISPONÍVEIS:")
        print("   - Master (Acesso total)")
        print("   - SAC (Pedidos e clientes)")
        print("   - Cadastro (Produtos e atributos)")
        print("   - Financeiro (Relatórios)")
        print("\n🔑 USUÁRIOS PADRÃO:")
        print("   master / master123 (Master)")
        print("   sac / sac123 (SAC)")
        print("   cadastro / cadastro123 (Cadastro)")
        print("="*50)
        print("⚠️  ALTERE AS SENHAS PELO SISTEMA APÓS O PRIMEIRO LOGIN!")
        print("="*50)
        
    except Exception as e:
        print(f"⚠️ Erro ao criar tabelas: {e}")
        import traceback
        traceback.print_exc()
        app.logger.error(f"Erro ao criar tabelas: {e}")

def obter_ultima_planilha():
    try:
        upload_folder = app.config["UPLOAD_FOLDER"]
        
        if not os.path.exists(upload_folder):
            app.logger.warning(f"Pasta uploads não encontrada: {upload_folder}")
            return None, None

        planilhas = []
        for f in os.listdir(upload_folder):
            file_path = os.path.join(upload_folder, f)
            if os.path.isfile(file_path) and f.lower().endswith(('.xlsx', '.xls', '.csv')):
                planilhas.append((f, os.path.getmtime(file_path)))

        if not planilhas:
            return None, None

        # Ordena por data (mais recente primeiro)
        planilhas.sort(key=lambda x: x[1], reverse=True)
        ultima_planilha = planilhas[0][0]
        
        return ultima_planilha, datetime.fromtimestamp(planilhas[0][1]).strftime('%Y-%m-%d %H:%M:%S')
    
    except Exception as e:
        app.logger.error(f"Erro em obter_ultima_planilha: {str(e)}")
        return None, None


@app.route("/")
@login_required
def index():
    """Redireciona para o dashboard do perfil"""
    if current_user.is_master():
        return redirect(url_for('dashboard_master'))
    elif current_user.perfil == 'SAC':
        return redirect(url_for('dashboard_sac'))
    elif current_user.perfil == 'Cadastro':
        return redirect(url_for('dashboard_cadastro'))
    else:
        return redirect(url_for('dashboard_master'))

@app.route('/home')
@login_required
def home():
    """Página home - redireciona para o dashboard apropriado"""
    if current_user.is_master():
        return redirect(url_for('dashboard_master'))
    elif current_user.perfil == 'SAC':
        return redirect(url_for('dashboard_sac'))
    elif current_user.perfil == 'Cadastro':
        return redirect(url_for('dashboard_cadastro'))
    else:
        return redirect(url_for('dashboard_master'))

@app.route('/sequencia-cadastros')
@login_required
#@permissao_modulo('produtos')
def sequencia_cadastros():
    """Página de sequência de cadastros"""
    return render_template('sequencia_cadastros.html', 
                         active_page='sequencia_cadastros',
                         page_title='Sequência de Cadastros')

@app.route('/admin/usuarios')
@login_required
@master_required
def admin_usuarios():
    usuarios = Usuario.query.all()
    perfis = Perfil.query.all()
    return render_template('admin/usuarios.html', usuarios=usuarios, perfis=perfis, page_title='Gerenciar Usuários')


# ============================================
# ROTAS DE GERENCIAMENTO DE USUÁRIOS
# ============================================

@app.route('/auth/admin/usuario/novo', methods=['POST'])
@login_required
@master_required
def admin_usuario_novo():
    """Cria um novo usuário"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        email    = data.get('email', '').strip()
        password = data.get('password', '')
        role     = data.get('role', 'SAC')

        if not username or not email or not password:
            return jsonify({'success': False, 'error': 'Usuário, e-mail e senha são obrigatórios'}), 400

        if Usuario.query.filter_by(username=username).first():
            return jsonify({'success': False, 'error': 'Nome de usuário já existe'}), 400

        if Usuario.query.filter_by(email=email).first():
            return jsonify({'success': False, 'error': 'E-mail já cadastrado'}), 400

        perfil = Perfil.query.filter_by(nome=role).first()
        if not perfil:
            return jsonify({'success': False, 'error': f'Perfil "{role}" não encontrado'}), 400

        novo = Usuario(
            username=username,
            email=email,
            perfil_id=perfil.id,
            is_active=True
        )
        novo.set_password(password)
        db.session.add(novo)
        db.session.commit()

        return jsonify({'success': True, 'id': novo.id})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/auth/admin/usuario/<int:usuario_id>/editar', methods=['PUT'])
@login_required
@master_required
def admin_usuario_editar(usuario_id):
    """Edita um usuário existente"""
    try:
        usuario = Usuario.query.get_or_404(usuario_id)
        data    = request.get_json()

        username  = data.get('username', '').strip()
        email     = data.get('email', '').strip()
        password  = data.get('password', '')
        role      = data.get('role')
        is_active = data.get('is_active')

        if not username or not email:
            return jsonify({'success': False, 'error': 'Usuário e e-mail são obrigatórios'}), 400

        # Verifica duplicidade apenas se mudou
        outro = Usuario.query.filter_by(username=username).first()
        if outro and outro.id != usuario_id:
            return jsonify({'success': False, 'error': 'Nome de usuário já existe'}), 400

        outro_email = Usuario.query.filter_by(email=email).first()
        if outro_email and outro_email.id != usuario_id:
            return jsonify({'success': False, 'error': 'E-mail já cadastrado'}), 400

        usuario.username = username
        usuario.email    = email

        if role:
            perfil = Perfil.query.filter_by(nome=role).first()
            if not perfil:
                return jsonify({'success': False, 'error': f'Perfil "{role}" não encontrado'}), 400
            usuario.perfil_id = perfil.id

        if is_active is not None:
            usuario.is_active = bool(is_active)

        if password:
            usuario.set_password(password)

        db.session.commit()
        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/auth/admin/usuario/<int:usuario_id>/excluir', methods=['DELETE'])
@login_required
@master_required
def admin_usuario_excluir(usuario_id):
    """Exclui um usuário"""
    try:
        # Impede que o próprio usuário logado se exclua
        if usuario_id == current_user.id:
            return jsonify({'success': False, 'error': 'Você não pode excluir seu próprio usuário'}), 400

        usuario = Usuario.query.get_or_404(usuario_id)
        db.session.delete(usuario)
        db.session.commit()
        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/google/sheets/list', methods=['GET'])
def api_listar_google_sheets():
    """Lista planilhas do usuário"""
    try:
        token_manager = GoogleTokenManager()
        credentials = token_manager.load_tokens()
        
        if not credentials or not credentials.valid:
            return jsonify({'success': False, 'error': 'Não conectado ao Google'}), 401
        
        from googleapiclient.discovery import build
        service = build('drive', 'v3', credentials=credentials)
        
        # Busca arquivos do tipo planilha
        results = service.files().list(
            q="mimeType='application/vnd.google-apps.spreadsheet'",
            pageSize=20,
            fields="files(id, name, modifiedTime)"
        ).execute()
        
        files = results.get('files', [])
        
        return jsonify({
            'success': True,
            'sheets': [
                {'id': f['id'], 'name': f['name'], 'modified': f.get('modifiedTime')}
                for f in files
            ]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/keepalive", methods=["GET", "POST"])
def keepalive():
    return {"status": "alive"}, 200

    
@app.route('/api/google/sheets/test', methods=['POST'])
def api_testar_google_sheet():
    """Testa acesso a uma planilha específica"""
    try:
        data = request.get_json()
        sheet_id = data.get('sheet_id')
        
        if not sheet_id:
            return jsonify({'success': False, 'error': 'ID da planilha obrigatório'}), 400
        
        token_manager = GoogleTokenManager()
        credentials = token_manager.load_tokens()
        
        if not credentials or not credentials.valid:
            return jsonify({'success': False, 'error': 'Não conectado ao Google'}), 401
        
        from googleapiclient.discovery import build
        service = build('sheets', 'v4', credentials=credentials)
        
        # Tenta ler informações da planilha
        spreadsheet = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
        
        # Pega nomes das abas
        sheets = spreadsheet.get('sheets', [])
        sheet_names = [s['properties']['title'] for s in sheets]
        
        # Tenta ler primeiras linhas
        result = service.spreadsheets().values().get(
            spreadsheetId=sheet_id,
            range='A1:E5'
        ).execute()
        
        values = result.get('values', [])
        
        return jsonify({
            'success': True,
            'message': f'Conexão bem sucedida! Planilha: {spreadsheet["properties"]["title"]}',
            'preview': {
                'total_linhas': len(values),
                'total_colunas': len(values[0]) if values else 0,
                'abas': sheet_names,
                'primeiras_linhas': values[:3]
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/google/sheets/config', methods=['POST'])
def api_salvar_config_google_sheets():
    """Salva configuração da planilha padrão"""
    try:
        data = request.get_json()
        sheet_id = data.get('sheet_id')
        sheet_aba = data.get('sheet_aba', '')
        
        config_file = Path('config/google_sheets_config.json')
        config_file.parent.mkdir(exist_ok=True)
        
        config = {
            'sheet_id': sheet_id,
            'sheet_aba': sheet_aba,
            'ultima_atualizacao': datetime.now().isoformat()
        }
        
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2)
        
        return jsonify({'success': True, 'message': 'Configuração salva!'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/google/sheets/config', methods=['GET'])
def api_carregar_config_google_sheets():
    """Carrega configuração da planilha padrão"""
    try:
        config_file = Path('config/google_sheets_config.json')
        
        if config_file.exists():
            with open(config_file, 'r') as f:
                config = json.load(f)
        else:
            config = {'sheet_id': '', 'sheet_aba': ''}
        
        return jsonify({'success': True, 'config': config})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
@app.route('/api/dashboard/atualizar-metricas')
@cache.cached(timeout=30)
def api_atualizar_metricas():
    """API para atualizar métricas do dashboard"""
    try:
        stats = get_processing_stats()
        
        # Métricas Mercado Livre
        ml_autenticado = ml_token_manager.is_authenticated()
        ml_metricas = {}
        if ml_autenticado:
            try:
                dados_ml = obter_dados_completos_perfil()
                ml_metricas = dados_ml.get('metricas', {})
            except:
                ml_metricas = {}
        
        # Métricas AnyMarket
        anymarket_stats = obter_estatisticas_anymarket_7_dias()
        
        return jsonify({
            'sucesso': True,
            'sistema': {
                'processamentos_hoje': stats['hoje'],
                'sucessos_hoje': stats['sucessos_hoje'],
                'erros_hoje': stats['erros_hoje'],
                'ultima_execucao': stats['ultima']
            },
            'mercadolivre': {
                'autenticado': ml_autenticado,
                'metricas': ml_metricas
            },
            'anymarket': anymarket_stats,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500


@app.route('/testar-endpoints-campanhas')
def testar_endpoints_campanhas():
    """Página para testar endpoints de campanhas"""
    from token_manager_secure import ml_token_manager
    
    esta_autenticado = ml_token_manager.is_authenticated()
    
    if not esta_autenticado:
        flash('Autentique-se primeiro no Mercado Livre', 'warning')
        return redirect(url_for('consultar_mercado_livre'))
    
    return render_template(
        'testar_endpoints.html',
        active_page='campanhas_mercadolivre',
        page_title='Testar Endpoints de Campanhas',
        esta_autenticado=esta_autenticado
    )

@app.route('/api/testar-endpoints', methods=['POST'])
def api_testar_endpoints():
    """API para testar endpoints"""
    try:
        import requests
        from token_manager_secure import ml_token_manager
        
        token = ml_token_manager.get_valid_token()
        
        if not token:
            return jsonify({'sucesso': False, 'erro': 'Token não disponível'})
        
        headers = {'Authorization': f'Bearer {token}'}
        base_url = "https://api.mercadolibre.com"
        
        resultados = []
        
        # Endpoints para testar
        endpoints = [
            {
                'nome': 'seller-promotions (principal)',
                'url': f'{base_url}/seller-promotions',
                'params': {'promotion_type': 'SELLER_CAMPAIGN', 'app_version': 'v2', 'limit': 5}
            },
            {
                'nome': 'seller_discounts',
                'url': f'{base_url}/seller_discounts',
                'params': {}
            },
            {
                'nome': 'users/me',
                'url': f'{base_url}/users/me',
                'params': {}
            }
        ]
        
        # Testar cada endpoint
        for endpoint in endpoints:
            try:
                response = requests.get(
                    endpoint['url'],
                    headers=headers,
                    params=endpoint['params'],
                    timeout=15
                )
                
                resultado = {
                    'nome': endpoint['nome'],
                    'url': endpoint['url'],
                    'status': response.status_code,
                    'sucesso': response.status_code == 200
                }
                
                if response.status_code == 200:
                    try:
                        data = response.json()
                        resultado['dados'] = data
                        
                        # Verificar se tem campanhas
                        if isinstance(data, list):
                            resultado['possui_campanhas'] = any('promotion_id' in str(item) for item in data)
                        elif isinstance(data, dict):
                            resultado['possui_campanhas'] = any('promotion' in key.lower() for key in data.keys())
                        else:
                            resultado['possui_campanhas'] = False
                            
                    except:
                        resultado['dados'] = response.text[:500]
                        resultado['possui_campanhas'] = False
                else:
                    resultado['erro'] = response.text[:200]
                
                resultados.append(resultado)
                
            except Exception as e:
                resultados.append({
                    'nome': endpoint['nome'],
                    'url': endpoint['url'],
                    'status': 'erro',
                    'sucesso': False,
                    'erro': str(e)
                })
        
        return jsonify({
            'sucesso': True,
            'resultados': resultados
        })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)})
    
@app.route('/campanhas-ativas')
def campanhas_ativas():
    """Página para visualizar campanhas ativas"""
    from token_manager_secure import ml_token_manager
    
    esta_autenticado = ml_token_manager.is_authenticated()
    
    if not esta_autenticado:
        flash('Autentique-se primeiro no Mercado Livre', 'warning')
        return redirect(url_for('consultar_mercado_livre'))
    
    return render_template(
        'campanhas_ativas.html',
        active_page='campanhas_mercadolivre',
        page_title='Campanhas Ativas no Mercado Livre',
        esta_autenticado=esta_autenticado
    )
@app.route('/api/campanhas-ativas', methods=['GET'])
@cache.cached(timeout=60)
def api_campanhas_ativas():
    """API para buscar TODAS as promoções do vendedor usando o endpoint oficial"""
    try:
        import requests
        from token_manager_secure import ml_token_manager
        
        # Obter token válido
        token = ml_token_manager.get_valid_token()
        
        if not token:
            return jsonify({
                'sucesso': False, 
                'erro': 'Token não disponível. Autentique-se primeiro.'
            })
        
        headers = {
            'Authorization': f'Bearer {token}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }
        
        base_url = "https://api.mercadolibre.com"
        
        # Primeiro, obter informações do usuário
        user_response = requests.get(
            f'{base_url}/users/me',
            headers=headers,
            timeout=15
        )
        
        if user_response.status_code != 200:
            return jsonify({
                'sucesso': False,
                'erro': 'Não foi possível obter informações do usuário'
            })
        
        user_data = user_response.json()
        user_id = user_data.get('id')
        
        # ENDPOINT CORRETO: Consultar TODAS as promoções do usuário
        # Documentação oficial: https://developers.mercadolivre.com.br/pt_br/gerenciar-ofertas [citation:6]
        promotions_response = requests.get(
            f'{base_url}/seller-promotions/users/{user_id}',
            headers=headers,
            params={'app_version': 'v2'},  # Versão obrigatória da API
            timeout=15
        )
        
        if promotions_response.status_code == 200:
            promotions_data = promotions_response.json()
            
            # A estrutura da resposta é { "results": [...] } [citation:6]
            promotions_list = promotions_data.get('results', [])
            
            if not promotions_list:
                return jsonify({
                    'sucesso': True,
                    'quantidade_total': 0,
                    'quantidade_ativas': 0,
                    'campanhas': [],
                    'mensagem': 'Nenhuma promoção encontrada para este usuário.',
                    'user_id': user_id,
                    'user_nickname': user_data.get('nickname', 'N/A'),
                    'endpoint_utilizado': f'/seller-promotions/users/{user_id}'
                })
            
            # Processar todas as promoções encontradas
            campanhas_processadas = []
            tipos_encontrados = set()
            
            for promotion in promotions_list:
                campanha_info = processar_promocao_unificada(promotion)
                if campanha_info:
                    campanhas_processadas.append(campanha_info)
                    tipos_encontrados.add(campanha_info['tipo'])
            
            # Separar campanhas ativas (status = 'started') [citation:1][citation:2][citation:6]
            campanhas_ativas = [
                c for c in campanhas_processadas 
                if c.get('status_original') == 'started'  # 'started' é o status para ativo
            ]
            
            return jsonify({
                'sucesso': True,
                'quantidade_total': len(campanhas_processadas),
                'quantidade_ativas': len(campanhas_ativas),
                'campanhas': campanhas_ativas,
                'todas_campanhas': campanhas_processadas,
                'tipos_encontrados': list(tipos_encontrados),
                'user_id': user_id,
                'user_nickname': user_data.get('nickname', 'N/A'),
                'endpoint_utilizado': f'/seller-promotions/users/{user_id}'
            })
            
        else:
            return jsonify({
                'sucesso': False,
                'erro': f'Erro ao buscar promoções. Status: {promotions_response.status_code}',
                'detalhes': promotions_response.text[:500],
                'user_id': user_id,
                'user_nickname': user_data.get('nickname', 'N/A')
            })
            
    except requests.exceptions.Timeout:
        return jsonify({'sucesso': False, 'erro': 'Timeout na requisição'})
    except requests.exceptions.ConnectionError:
        return jsonify({'sucesso': False, 'erro': 'Erro de conexão'})
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)})

def processar_promocao_unificada(promocao):
    """Processa dados de promoção baseado na documentação oficial unificada [citation:6]"""
    if not promocao:
        return None
    
    try:
        # Status possíveis segundo documentação [citation:1][citation:2][citation:6]
        # - pending: Promoção aprovada, mas ainda não iniciou
        # - started: Promoção ativa
        # - finished: Promoção finalizada
        
        status_map = {
            'pending': 'Pendente',
            'started': 'Ativa',
            'finished': 'Finalizada',
            'candidate': 'Candidata'
        }
        
        status = promocao.get('status', 'N/A')
        status_formatado = status_map.get(status, status)
        
        # Extrair benefícios se existirem [citation:6]
        benefits = promocao.get('benefits', {})
        
        return {
            'id': promocao.get('id', 'N/A'),
            'nome': promocao.get('name', 'Sem nome'),
            'tipo': promocao.get('type', 'N/A'),  # DEAL, SELLER_CAMPAIGN, LIGHTNING, etc.
            'status': status_formatado,
            'status_original': status,
            'is_active': status == 'started',
            'data_inicio': promocao.get('start_date', 'N/A'),
            'data_fim': promocao.get('finish_date', 'N/A'),
            'deadline': promocao.get('deadline_date', 'N/A'),  # Prazo para aceitar convite
            'beneficios': {
                'tipo': benefits.get('type'),
                'meli_percent': benefits.get('meli_percent'),  # % que o MELI paga
                'seller_percent': benefits.get('seller_percent'),  # % que o vendedor paga
                'buy_quantity': benefits.get('buy_quantity'),  # Para VOLUME (ex: compre 3)
                'pay_quantity': benefits.get('pay_quantity')   # Para VOLUME (ex: pague 2)
            } if benefits else None,
            'detalhes': promocao
        }
    except Exception as e:
        return {
            'id': 'erro',
            'nome': f'Erro ao processar: {str(e)}',
            'tipo': promocao.get('type', 'desconhecido'),
            'dados_brutos': promocao
        }

@app.route('/api/campanha/<promotion_id>/itens', methods=['GET'])
def api_campanha_itens(promotion_id):
    """Busca os itens de uma campanha com status de elegibilidade"""
    try:
        from token_manager_secure import ml_token_manager
        import requests
        
        token = ml_token_manager.get_valid_token()
        if not token:
            return jsonify({'sucesso': False, 'erro': 'Token não disponível'})
        
        promotion_type = request.args.get('type', '')
        
        headers = {'Authorization': f'Bearer {token}'}
        base_url = "https://api.mercadolibre.com"
        
        # Endpoint para buscar itens da campanha [citation:1][citation:2][citation:4]
        url = f'{base_url}/seller-promotions/promotions/{promotion_id}/items'
        params = {
            'promotion_type': promotion_type,
            'app_version': 'v2',
            'limit': 100
        }
        
        response = requests.get(url, headers=headers, params=params, timeout=15)
        
        if response.status_code == 200:
            data = response.json()
            items = data.get('results', [])
            
            # Mapeamento de status para português
            status_map = {
                'candidate': 'Elegível',
                'pending': 'Programado',
                'started': 'Participando',
                'finished': 'Finalizado'
            }
            
            itens_processados = []
            for item in items:
                # Calcular horas restantes para início/prazo
                horas_para_inicio = calcular_horas_restantes(item.get('start_date', ''))
                horas_para_fim = calcular_horas_restantes(item.get('end_date', ''))
                
                item_info = {
                    'id': item.get('id'),
                    'status': item.get('status'),
                    'status_pt': status_map.get(item.get('status'), item.get('status')),
                    'preco_promocional': item.get('price', 0),
                    'preco_original': item.get('original_price', 0),
                    'desconto': calcular_desconto(item.get('price', 0), item.get('original_price', 0)),
                    'percentual_meli': item.get('meli_percentage', 0),
                    'percentual_vendedor': item.get('seller_percentage', 0),
                    'data_inicio': item.get('start_date', ''),
                    'data_fim': item.get('end_date', ''),
                    'horas_para_inicio': horas_para_inicio,
                    'horas_para_fim': horas_para_fim,
                    'is_urgente': horas_para_inicio is not None and horas_para_inicio <= 24,  # Urgente se <= 24h
                    'preco_minimo': item.get('min_discounted_price'),  # Preço mínimo permitido [citation:1]
                    'preco_maximo': item.get('max_discounted_price'),  # Preço máximo permitido [citation:1]
                    'preco_sugerido': item.get('suggested_discounted_price'),  # Preço sugerido [citation:1]
                    'stock_min': item.get('stock', {}).get('min') if item.get('stock') else None,  # Estoque mínimo [citation:4]
                    'stock_max': item.get('stock', {}).get('max') if item.get('stock') else None,  # Estoque máximo [citation:4]
                    'top_deal_price': item.get('top_deal_price')  # Preço para compradores Top [citation:1]
                }
                
                # Buscar detalhes do anúncio (título, imagem)
                try:
                    item_response = requests.get(
                        f'{base_url}/items/{item.get("id")}',
                        headers=headers,
                        timeout=10
                    )
                    if item_response.status_code == 200:
                        item_data = item_response.json()
                        item_info['titulo'] = item_data.get('title', 'Sem título')
                        item_info['thumbnail'] = item_data.get('thumbnail', '')
                        item_info['permalink'] = item_data.get('permalink', '')
                except:
                    item_info['titulo'] = 'Erro ao carregar título'
                
                itens_processados.append(item_info)
            
            return jsonify({
                'sucesso': True,
                'itens': itens_processados,
                'total': len(itens_processados)
            })
        else:
            return jsonify({
                'sucesso': False,
                'erro': f'Erro {response.status_code}',
                'detalhes': response.text[:200]
            })
            
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)})

def calcular_horas_restantes(data_str):
    """Calcula horas restantes para uma data"""
    if not data_str or data_str == 'N/A':
        return None
    try:
        from datetime import datetime
        data = datetime.fromisoformat(data_str.replace('Z', '+00:00'))
        agora = datetime.now(data.tzinfo)
        diff = data - agora
        horas = diff.total_seconds() / 3600
        return round(horas, 1)
    except:
        return None

def calcular_desconto(preco_promocional, preco_original):
    """Calcula porcentagem de desconto"""
    if preco_original and preco_original > 0 and preco_promocional < preco_original:
        return round(((preco_original - preco_promocional) / preco_original) * 100, 1)
    return 0

   
def processar_campanha_ml(campanha, tipo_promocao):
    """Processa dados de campanha baseado na documentação oficial [citation:2][citation:5]"""
    if not campanha:
        return None
    
    try:
        # Mapeamento de status conforme documentação [citation:2][citation:5]
        status_map = {
            'pending': 'Pendente',    # Promoção aprovada, mas ainda não iniciou
            'started': 'Ativa',        # Promoção ativa [citation:5]
            'finished': 'Finalizada',  # Promoção finalizada
            'candidate': 'Candidata'   # Item candidato [citation:2]
        }
        
        status = campanha.get('status', 'N/A')
        status_formatado = status_map.get(status, status)
        
        return {
            'id': campanha.get('id', 'N/A'),
            'nome': campanha.get('name', 'Sem nome'),
            'tipo': tipo_promocao,
            'subtipo': campanha.get('sub_type', 'N/A'),  # FIXED_PERCENTAGE ou FLEXIBLE_PERCENTAGE [citation:5]
            'status': status_formatado,
            'status_original': status,
            'is_active': status == 'started',  # Apenas 'started' é considerado ativo [citation:5]
            'data_inicio': campanha.get('start_date', 'N/A'),
            'data_fim': campanha.get('finish_date', 'N/A'),
            'detalhes': campanha
        }
    except Exception as e:
        return None

def processar_promocao_ml(promocao):
    """Processa dados de promoção do Mercado Livre baseado na documentação oficial [citation:2]"""
    if not promocao:
        return None
    
    try:
        # Mapeamento de status baseado na documentação [citation:1][citation:3]
        status_map = {
            'pending': 'Pendente',
            'started': 'Ativa',  # Status para promoção ativa [citation:1]
            'finished': 'Finalizada',
            'candidate': 'Candidato',
            'active': 'Ativa'
        }
        
        status = promocao.get('status', 'N/A')
        status_formatado = status_map.get(status, status)
        
        # Determinar se é uma campanha ativa baseado no status [citation:1]
        is_active = status in ['started', 'active']
        
        # Extrair informações de benefícios se existirem [citation:2]
        beneficios = promocao.get('benefits', {})
        
        return {
            'id': promocao.get('id', 'N/A'),
            'nome': promocao.get('name', 'Sem nome'),
            'tipo': promocao.get('type', 'N/A'),  # DEAL, SELLER_CAMPAIGN, etc. [citation:2]
            'status': status_formatado,
            'status_original': status,
            'is_active': is_active,
            'data_inicio': promocao.get('start_date', 'N/A'),
            'data_fim': promocao.get('finish_date', 'N/A'),
            'deadline': promocao.get('deadline_date', 'N/A'),  # Prazo para aceitar convite [citation:2]
            'beneficios': {
                'tipo': beneficios.get('type'),
                'meli_percent': beneficios.get('meli_percent'),  # % que o MELI paga [citation:2]
                'seller_percent': beneficios.get('seller_percent')  # % que o vendedor paga [citation:2]
            } if beneficios else None,
            'detalhes': promocao  # Dados brutos para debug
        }
    except Exception as e:
        return {
            'id': 'erro',
            'nome': f'Erro ao processar: {str(e)}',
            'tipo': promocao.get('type', 'desconhecido'),
            'dados_brutos': promocao
        }

def processar_campanha(campanha):
    """Função auxiliar para processar dados da campanha"""
    if not campanha:
        return None
    
    try:
        # Extrair informações relevantes
        return {
            'id': campanha.get('id') or campanha.get('promotion_id') or campanha.get('campaign_id', 'N/A'),
            'nome': campanha.get('name') or campanha.get('title') or campanha.get('campaign_name', 'Sem nome'),
            'tipo': campanha.get('type') or campanha.get('promotion_type', 'N/A'),
            'status': campanha.get('status') or campanha.get('campaign_status', 'N/A'),
            'data_inicio': campanha.get('start_date') or campanha.get('date_created', 'N/A'),
            'data_fim': campanha.get('end_date') or campanha.get('date_modified', 'N/A'),
            'desconto': campanha.get('discount') or campanha.get('discount_percentage', 0),
            'itens_envolvidos': campanha.get('items_quantity') or campanha.get('total_items', 0),
            'detalhes': campanha  # Incluir dados brutos para debug
        }
    except Exception as e:
        return {
            'id': 'erro',
            'nome': f'Erro ao processar: {str(e)}',
            'dados_brutos': campanha
        }


            
@app.route('/api/mercadolivre/atualizar-manufacturing', methods=['POST'])
def atualizar_manufacturing():
    """Rota para atualizar manufacturing time"""
    try:
        data = request.get_json()
        mlb = data.get('mlb')
        dias = data.get('dias')
        atualizacoes = data.get('atualizacoes')  # Para múltiplos
        
        if atualizacoes:
            # Atualização em massa
            resultado = ml_api_secure.atualizar_multiplos_manufacturing(atualizacoes)
        elif mlb and dias:
            # Atualização única
            resultado = ml_api_secure.atualizar_manufacturing_time(mlb, dias)
        else:
            return jsonify({
                'sucesso': False,
                'erro': 'Parâmetros insuficientes. Forneça mlb/dias ou atualizacoes'
            })
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        })

@app.route('/api/mercadolivre/debug-mlb/<mlb>')
def debug_mlb(mlb):
    """Rota simples para debug de um MLB específico"""
    try:
        from mercadolivre_api_secure import ml_api_secure
        ml_api_secure.debug_json_completo(mlb)
        return jsonify({
            'sucesso': True,
            'mensagem': f'Debug do MLB {mlb} realizado - verifique o console do servidor'
        })
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        })

@app.route('/api/dashboard/metricas-gerais')
def api_metricas_gerais():
    """API para obter métricas gerais para dashboard"""
    try:
        stats = get_processing_stats()
        
        # Métricas Mercado Livre
        ml_autenticado = ml_token_manager.is_authenticated()
        ml_metricas = {}
        if ml_autenticado:
            try:
                dados_ml = obter_dados_completos_perfil()
                ml_metricas = dados_ml.get('metricas', {})
            except:
                ml_metricas = {}
        
        # Métricas AnyMarket
        anymarket_token_configurado = verificar_token_anymarket_configurado()
        anymarket_metricas = {}
        
        return jsonify({
            'sucesso': True,
            'sistema': {
                'processamentos_hoje': stats['hoje'],
                'sucessos_hoje': stats['sucessos_hoje'],
                'erros_hoje': stats['erros_hoje'],
                'ultima_execucao': stats['ultima']
            },
            'mercadolivre': {
                'autenticado': ml_autenticado,
                'metricas': ml_metricas
            },
            'anymarket': {
                'token_configurado': anymarket_token_configurado
            },
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/dashboard/atualizar-planilhas')
def api_atualizar_planilhas():
    """API para atualizar lista de planilhas"""
    try:
        ultima_planilha, ultima_planilha_data = obter_ultima_planilha()
        
        return jsonify({
            'sucesso': True,
            'ultima_planilha': ultima_planilha,
            'ultima_planilha_data': ultima_planilha_data,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
     
@app.route('/pedidos-anymarket')
@login_required
@permissao_modulo('pedidos')  ####Esta rota apenas chama a função abaixo api_pedidos_anymarket
def pedidos_anymarket():
    """Página principal de pedidos do AnyMarket"""
    return render_template('pedidos_anymarket.html',
                            active_page='pedidos',
                            active_module='anymarket',
                            page_title='Pedidos - Anymarket'
                            )


@app.route('/canais-transmissao')
def canais_transmissao():
    """Página para consultar canais de transmissão do AnyMarket"""
    try:
        # Verificar se token está configurado
        token_configurado = verificar_token_anymarket_configurado()
        partner_id = request.args.get('partner_id', '')
        dados = []
        stats = {
            'total_canais': 0,
            'total_ativos': 0,
            'total_inativos': 0,
            'total_transmitindo': 0
        }
        token_preview = ''
        
        if token_configurado and partner_id:
            try:
                from processamento.api_anymarket import consultar_canais_transmissao
                token = obter_token_anymarket_seguro()
                
                # Mostrar preview do token (primeiros e últimos caracteres)
                if token:
                    token_preview = token[:10] + '...' + token[-10:]
                
                resultado = consultar_canais_transmissao(partner_id)
                
                if resultado.get('sucesso') and resultado.get('dados'):
                    dados = resultado['dados']
                    
                    # Calcular estatísticas
                    stats['total_canais'] = len(dados)
                    stats['total_ativos'] = sum(1 for c in dados if c.get('active'))
                    stats['total_inativos'] = stats['total_canais'] - stats['total_ativos']
                    stats['total_transmitindo'] = stats['total_ativos']  # Simplificado
                    
                    flash(f'{len(dados)} canais encontrados!', 'success')
                elif resultado.get('erro'):
                    flash(f'Erro: {resultado["erro"]}', 'danger')
                    
            except Exception as e:
                flash(f'Erro ao consultar API: {str(e)}', 'danger')
                print(f"❌ Erro: {str(e)}")
                import traceback
                traceback.print_exc()
        
        elif partner_id and not token_configurado:
            flash('Token do AnyMarket não configurado. Configure em Configurações.', 'warning')
        
        return render_template(
            'canais_transmissao.html',
            active_page='canais_transmissao',
            active_module='anymarket',
            page_title='Canais de Transmissão - AnyMarket',
            dados=dados,
            stats=stats,
            token_configurado=token_configurado,
            token_preview=token_preview,
            partner_id=partner_id
        )
        
    except Exception as e:
        flash(f'Erro interno: {str(e)}', 'danger')
        return render_template(
            'canais_transmissao.html',
            active_page='canais_transmissao',
            active_module='anymarket',
            page_title='Canais de Transmissão - AnyMarket',
            dados=[],
            stats={},
            token_configurado=False,
            token_preview='',
            partner_id=''
        )

@app.route('/alterar-modo-envio')
def alterar_modo_envio():
    """Página para alterar modo de envio (ME1 ↔ ME2)"""
    # Obtém informações da conta atual
    accounts = ml_token_manager.get_all_accounts()
    current_account = None
    conta_atual_id = ml_token_manager.current_account_id
    
    if conta_atual_id and conta_atual_id in ml_token_manager.accounts:
        current_account = ml_token_manager.accounts[conta_atual_id]
    
    # Verifica se tem alguma conta autenticada
    esta_autenticado = False
    for account in accounts:
        if account.get('has_token'):
            esta_autenticado = True
            break
    
    return render_template(
        'alterar_modo_envio.html',
        active_page='alterar_modo_envio',
        active_module='mercadolivre',
        page_title='Alterar Modo de Envio',
        esta_autenticado=esta_autenticado,
        conta_atual=current_account,
        conta_atual_id=conta_atual_id,
        todas_contas=accounts
    )

@app.route('/api/mercadolivre/verificar-me2/<mlb>')
def api_verificar_me2(mlb):
    """API para verificar requisitos ME2 de um MLB"""
    try:
        resultado = ml_api_secure.verificar_requisitos_me2(mlb)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/alterar-me2', methods=['POST'])
def api_alterar_me2():
    """API para alterar modo de envio para ME2"""
    try:
        data = request.get_json()
        mlb = data.get('mlb')
        mlbs = data.get('mlbs')
        
        if mlbs:
            # Migração em massa
            resultado = ml_api_secure.alterar_multiplos_para_me2(mlbs)
        elif mlb:
            # Migração única
            resultado = ml_api_secure.alterar_para_me2(mlb)
        else:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum MLB fornecido'
            }), 400
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/debug-envio/<mlb>')
def api_debug_envio(mlb):
    """Rota para debug da mudança de envio"""
    try:
        resultado = ml_api_secure.debug_mudanca_envio(mlb)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
    
@app.route('/api/anymarket/canais-transmissao')
def api_canais_transmissao():
    """API para consultar canais de transmissão (retorna JSON)"""
    try:
        auth_header = request.headers.get('Authorization', '')
        partner_id = request.args.get('partner_id', '')
        
        # Se não veio no header, tenta obter do token seguro
        token = None
        if auth_header.startswith('Bearer '):
            token = auth_header.replace('Bearer ', '')
        else:
            from processamento.api_anymarket import obter_token_anymarket_seguro
            token = obter_token_anymarket_seguro()
        
        if not token:
            return jsonify({
                'success': False,
                'error': 'Token não fornecido e não configurado'
            }), 401
        
        from processamento.api_anymarket import consultar_canais_transmissao
        resultado = consultar_canais_transmissao(partner_id, token)
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    
@app.route('/api/anymarket/pedidos')
@login_required
def api_pedidos_anymarket():
    """API para buscar pedidos do AnyMarket - CORREÇÃO DO FILTRO DE DATA"""
    try:
        # Obter token do header Authorization
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'success': False, 'error': 'Token de autenticação não fornecido'}), 401
        
        token = auth_header.replace('Bearer ', '')
        
        # Parâmetros da requisição
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        status = request.args.get('status')
        marketplace = request.args.get('marketplace')
        data_inicio = request.args.get('createdAfter')  # ✅ Note: createAfter (sem 'd')
        data_fim = request.args.get('createdBefore')
        
        
        # ✅ VALIDAÇÃO DA PÁGINA
        if page < 1:
            page = 1
        
        sort_field = request.args.get('sort', 'createdAt')
        sort_direction = request.args.get('sortDirection', 'DESC')

        # Construir URL da API AnyMarket
        url = "https://api.anymarket.com.br/v2/orders"
        
        # ✅ CORREÇÃO: Usar offset em vez de page
        offset = (page - 1) * limit
        
        params = {
            'offset': offset,
            'limit': limit,
        }
        
        if sort_field and sort_direction:
            params['sort'] = sort_field
            params['sortDirection'] = sort_direction

        # Adicionar filtros APENAS se fornecidos
        if status and status.strip():
            params['status'] = status.strip()
            
        if marketplace and marketplace.strip():
            params['marketplace'] = marketplace.strip()
        
        # ✅ CORREÇÃO CRÍTICA: Processar datas INDEPENDENTEMENTE
        if data_inicio and data_inicio.strip():
            try:
                # Validar e formatar data início
                datetime.strptime(data_inicio, '%Y-%m-%d')
                params['createdAfter'] = f"{data_inicio}T00:00:00-03:00"
                print(f"✅ Filtro data início: {data_inicio} -> {params['createdAfter']}")
            except ValueError as e:
                print(f"⚠️ Data início em formato inválido: {data_inicio}, erro: {e}")
        else:
            print("ℹ️ Data início não fornecida")
        
        if data_fim and data_fim.strip():
            try:
                # Validar e formatar data fim
                datetime.strptime(data_fim, '%Y-%m-%d')
                params['createdBefore'] = f"{data_fim}T23:59:59-03:00"
                print(f"✅ Filtro data fim: {data_fim} -> {params['createdBefore']}")
            except ValueError as e:
                print(f"⚠️ Data fim em formato inválido: {data_fim}, erro: {e}")
        else:
            print("ℹ️ Data fim não fornecida")
        
        # Fazer requisição para a API AnyMarket
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json',
            'gumgaToken': token
        }
        
        print(f"🔍 SOLICITANDO PÁGINA {page} (offset: {offset}) para AnyMarket")
        print(f"📋 Parâmetros FINAIS: {params}")
        
        response = requests.get(url, params=params, headers=headers, timeout=60)
        
        print(f"📡 Resposta da API: {response.status_code}")
        
        if response.status_code != 200:
            error_detail = response.text
            print(f"❌ Erro {response.status_code}: {error_detail}")
            return jsonify({
                'success': False, 
                'error': f'Erro na API AnyMarket: {response.status_code}',
                'details': error_detail[:500] if error_detail else 'Sem detalhes'
            }), response.status_code
        
        data = response.json()
        
        # Processar resposta
        orders = data.get('content', [])
        pagination_data = data.get('page', {})
        
        # ✅ CORREÇÃO CRÍTICA: Calcular paginação corretamente
        total_elements = pagination_data.get('totalElements', 0)
        page_size = pagination_data.get('size', limit)
        
        # ✅ CÁLCULO CORRETO DA PAGINAÇÃO
        total_pages = max(1, (total_elements + page_size - 1) // page_size)
        current_page = page  # Usar a página solicitada
        
        print(f"📊 PAGINAÇÃO CALCULADA:")
        print(f"   - Total elementos: {total_elements}")
        print(f"   - Tamanho da página: {page_size}")
        print(f"   - Total de páginas: {total_pages}")
        print(f"   - Página atual: {current_page}")
        
        # ✅ CORREÇÃO: Calcular navegação baseado nos cálculos
        has_next = current_page < total_pages
        has_prev = current_page > 1
        
        pagination = {
            'currentPage': current_page,
            'totalPages': total_pages,
            'totalElements': total_elements,
            'hasNext': has_next,
            'hasPrev': has_prev,
            'pageSize': page_size
        }
        
        # Calcular estatísticas
        stats = {
            'total': len(orders),
            'pendentes': len([o for o in orders if o.get('status') == 'PENDING']),
            'valorTotal': sum(float(o.get('total', 0)) for o in orders),
            'totalGeral': total_elements
        }
        
        return jsonify({
            'success': True,
            'orders': orders,
            'stats': stats,
            'pagination': pagination,
            'filters': {
                'createdAfter': data_inicio or '',
                'createdBefore': data_fim or '',
                'status': status or '',
                'marketplace': marketplace or '',
                'sort': sort_field,
                'sortDirection': sort_direction
            },
            'debug': {
                'pagina_solicitada': page,
                'offset_calculado': offset,
                'total_pages_calculado': total_pages,
                'api_url': response.url,
                'ordenacao': f'{sort_field} {sort_direction}'
            }
        })
        
    except Exception as e:
        print(f"❌ Erro na API pedidos: {str(e)}")
        return jsonify({'success': False, 'error': f'Erro interno: {str(e)}'}), 500
 
# ========================================
# ROTAS MERCADO LIVRE (MLB)
# ========================================

@app.route('/consultar-mercado-livre')
def consultar_mercado_livre():
    """Página principal para consulta de MLBs"""
    # Obtém informações da conta atual
    accounts = ml_token_manager.get_all_accounts()
    current_account = None
    conta_atual_id = ml_token_manager.current_account_id
    
    if conta_atual_id and conta_atual_id in ml_token_manager.accounts:
        current_account = ml_token_manager.accounts[conta_atual_id]
    
    # Verifica se tem alguma conta autenticada
    esta_autenticado = False
    for account in accounts:
        if account.get('has_token'):
            esta_autenticado = True
            break
    
    return render_template(
        'consultar_mercado_livre.html',
        active_page='consultar_mercado_livre',
        active_module='mercadolivre',
        page_title='Consulta Mercado Livre',
        esta_autenticado=esta_autenticado,
        conta_atual=current_account,
        conta_atual_id=conta_atual_id,
        todas_contas=accounts
    )

@app.route('/api/mercadolivre/configurar', methods=['POST'])
def api_configurar_mercadolivre():
    """API para configurar Client ID e Secret"""
    try:
        data = request.get_json()
        client_id = data.get('client_id')
        client_secret = data.get('client_secret')
        
        if not client_id or not client_secret:
            return jsonify({'sucesso': False, 'erro': 'Client ID e Client Secret são obrigatórios'}), 400
        
        ml_token_manager.set_config(client_id, client_secret)
        
        return jsonify({
            'sucesso': True, 
            'mensagem': 'Configuração salva com sucesso!'
        })
            
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500


@app.route('/api/mercadolivre/forcar-renovacao', methods=['POST'])
def api_forcar_renovacao():
    """Força a renovação do token"""
    try:
        token_data = ml_token_manager.load_tokens()
        if not token_data or not token_data.get('refresh_token'):
            return jsonify({'sucesso': False, 'erro': 'Nenhum refresh token disponível'}), 400
        
        print("🔄 Forçando renovação do token...")
        novo_token = ml_token_manager.refresh_token(token_data['refresh_token'])
        
        if novo_token:
            # Testa o novo token
            if ml_token_manager.testar_token_api(novo_token):
                return jsonify({
                    'sucesso': True, 
                    'mensagem': 'Token renovado com sucesso!'
                })
            else:
                return jsonify({
                    'sucesso': False, 
                    'erro': 'Token renovado mas não funciona na API'
                })
        else:
            return jsonify({
                'sucesso': False, 
                'erro': 'Falha na renovação do token. É necessário nova autenticação.'
            })
            
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

# ============================================
# ROTAS PARA ALTERAR MODELO DO PRODUTO
# ============================================

@app.route('/mercadolivre/alterar-modelo')
@login_required
@permissao_modulo('produtos')
def alterar_modelo_ml():
    """Página para alterar modelo de produtos ML"""
    return render_template(
        'mercadolivre/alterar_modelo.html',
        active_page='alterar_modelo_ml',
        active_module='mercadolivre',
        page_title='Alterar Modelo - Mercado Livre'
    )

@app.route('/api/mercadolivre/buscar-modelo', methods=['POST'])
@login_required
def api_buscar_modelo():
    """Busca o modelo atual de um produto"""
    try:
        if not ml_token_manager.is_authenticated():
            return jsonify({'sucesso': False, 'erro': 'Não autenticado no Mercado Livre'}), 401
        
        data = request.get_json()
        mlb = data.get('mlb')
        
        if not mlb:
            return jsonify({'sucesso': False, 'erro': 'MLB é obrigatório'}), 400
        
        resultado = ml_api_secure.buscar_atributo_modelo(mlb)
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/buscar-modelos-disponiveis', methods=['POST'])
@login_required
def api_buscar_modelos_disponiveis():
    """Busca os modelos disponíveis para um produto"""
    try:
        if not ml_token_manager.is_authenticated():
            return jsonify({'sucesso': False, 'erro': 'Não autenticado no Mercado Livre'}), 401
        
        data = request.get_json()
        mlb = data.get('mlb')
        
        resultado = ml_api_secure.buscar_modelos_disponiveis(mlb)
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

        
@app.route('/api/mercadolivre/alterar-modelo', methods=['POST'])
@login_required
def api_alterar_modelo():
    """Altera o modelo de um produto"""
    try:
        if not ml_token_manager.is_authenticated():
            return jsonify({'sucesso': False, 'erro': 'Não autenticado no Mercado Livre'}), 401
        
        data = request.get_json()
        mlb = data.get('mlb')
        mlbs = data.get('mlbs')
        novo_modelo_nome = data.get('modelo_nome')
        
        if not novo_modelo_nome:
            return jsonify({'sucesso': False, 'erro': 'Nome do modelo é obrigatório'}), 400
        
        if mlbs and len(mlbs) > 0:
            # Alteração em lote
            resultado = ml_api_secure.alterar_modelo_multiplos(mlbs, novo_modelo_nome)
        elif mlb:
            # Alteração única
            resultado = ml_api_secure.alterar_modelo_produto(mlb, novo_modelo_nome)
        else:
            return jsonify({'sucesso': False, 'erro': 'MLB ou lista de MLBs é obrigatória'}), 400
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
      
@app.route('/api/mercadolivre/excluir-definitivo', methods=['POST'])
@login_required
def api_excluir_mlb_definitivo():
    """
    Rota para exclusão definitiva de MLB (2 etapas)
    """
    try:
        if not ml_token_manager.is_authenticated():
            return jsonify({
                'sucesso': False, 
                'erro': 'Não autenticado no Mercado Livre',
                'codigo': 'nao_autenticado'
            }), 401
        
        data = request.get_json()
        mlb = data.get('mlb')
        mlbs = data.get('mlbs')  # Para exclusão em lote
        
        if not mlb and not mlbs:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum MLB fornecido',
                'codigo': 'sem_mlb'
            }), 400
        
        # Função para validar e limpar MLB
        def processar_mlb(mlb_str):
            if not mlb_str:
                return None
            mlb_str = str(mlb_str).strip().upper()
            # Remove caracteres não alfanuméricos
            import re
            mlb_str = re.sub(r'[^A-Z0-9]', '', mlb_str)
            
            if mlb_str.startswith('MLB'):
                return mlb_str
            elif mlb_str.isdigit():
                return 'MLB' + mlb_str
            else:
                numeros = re.sub(r'\D', '', mlb_str)
                return 'MLB' + numeros if numeros else None
        
        if mlbs:
            # Exclusão em lote
            mlbs_validos = []
            for m in mlbs:
                mlb_processado = processar_mlb(m)
                if mlb_processado:
                    mlbs_validos.append(mlb_processado)
            
            if not mlbs_validos:
                return jsonify({
                    'sucesso': False,
                    'erro': 'Nenhum MLB válido encontrado',
                    'codigo': 'mlbs_invalidos'
                }), 400
            
            print(f"🔍 Iniciando exclusão em lote de {len(mlbs_validos)} MLBs")
            
            resultados = []
            for mlb_valido in mlbs_validos:
                resultado = ml_api_secure.excluir_anuncio_definitivo(mlb_valido)
                resultado['mlb'] = mlb_valido
                resultados.append(resultado)
                
                # Delay entre exclusões para evitar rate limit
                import time
                time.sleep(1)
            
            sucessos = sum(1 for r in resultados if r.get('sucesso'))
            
            return jsonify({
                'sucesso': sucessos > 0,
                'resultados': resultados,
                'total': len(mlbs_validos),
                'sucessos': sucessos,
                'erros': len(mlbs_validos) - sucessos
            })
            
        else:
            # Exclusão única
            mlb_processado = processar_mlb(mlb)
            
            if not mlb_processado:
                return jsonify({
                    'sucesso': False,
                    'erro': f'MLB inválido: {mlb}',
                    'codigo': 'mlb_invalido'
                }), 400
            
            print(f"🔍 Iniciando exclusão única do MLB: {mlb_processado}")
            resultado = ml_api_secure.excluir_anuncio_definitivo(mlb_processado)
            
            return jsonify(resultado)
            
    except Exception as e:
        print(f"❌ ERRO NA ROTA DE EXCLUSÃO: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'erro': f'Erro interno: {str(e)}',
            'codigo': 'erro_interno'
        }), 500
################################################################ ATRIBUTOS ##############################################################################

@app.route('/api/mercadolivre/atributos-disponiveis/<mlb>')
@login_required
@permissao_modulo('produtos')
def api_atributos_disponiveis(mlb):
    """Retorna todos os atributos que podem ser alterados no produto"""
    try:
        if not ml_token_manager.is_authenticated():
            return jsonify({'sucesso': False, 'erro': 'Não autenticado'}), 401
        
        token = ml_token_manager.get_valid_token()
        headers = {'Authorization': f'Bearer {token}'}
        
        # Busca o produto
        response = requests.get(
            f'https://api.mercadolibre.com/items/{mlb}',
            headers=headers,
            timeout=15
        )
        
        if response.status_code != 200:
            return jsonify({'sucesso': False, 'erro': 'Produto não encontrado'}), 404
        
        dados = response.json()
        
        # Lista de atributos que queremos permitir alterar
        atributos_permitidos = [
            'MODEL', 'BRAND', 'COLOR', 'MATERIAL', 'WARRANTY_TIME',
            'SELLER_SKU', 'GTIN', 'HEIGHT', 'WIDTH', 'DEPTH', 'WEIGHT'
        ]
        
        atributos_encontrados = []
        for attr in dados.get('attributes', []):
            if attr.get('id') in atributos_permitidos:
                atributos_encontrados.append({
                    'id': attr.get('id'),
                    'name': attr.get('name'),
                    'current_value': attr.get('value_name'),
                    'current_id': attr.get('value_id'),
                    'value_type': attr.get('value_type')
                })
        
        # Adiciona atributos que não existem no produto (para criar)
        ids_encontrados = [a['id'] for a in atributos_encontrados]
        for attr_id in atributos_permitidos:
            if attr_id not in ids_encontrados:
                nome_map = {
                    'MODEL': 'Modelo', 'BRAND': 'Marca', 'COLOR': 'Cor',
                    'MATERIAL': 'Material', 'WARRANTY_TIME': 'Garantia',
                    'SELLER_SKU': 'SKU', 'GTIN': 'GTIN/EAN',
                    'HEIGHT': 'Altura', 'WIDTH': 'Largura',
                    'DEPTH': 'Profundidade', 'WEIGHT': 'Peso'
                }
                atributos_encontrados.append({
                    'id': attr_id,
                    'name': nome_map.get(attr_id, attr_id),
                    'current_value': None,
                    'current_id': None,
                    'value_type': 'string',
                    'not_exists': True
                })
        
        return jsonify({
            'sucesso': True,
            'mlb': mlb,
            'titulo': dados.get('title'),
            'atributos': atributos_encontrados,
            'is_catalog': dados.get('catalog_listing', False)
        })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
    
@app.route('/api/mercadolivre/alterar-atributos', methods=['POST'])
@login_required
@permissao_modulo('produtos')
def api_alterar_atributos():
    """Altera múltiplos atributos de um produto de uma vez"""
    try:
        if not ml_token_manager.is_authenticated():
            return jsonify({'sucesso': False, 'erro': 'Não autenticado'}), 401
        
        data = request.get_json()
        mlb = data.get('mlb')
        atributos = data.get('atributos', {})  # Dict com {id: value}
        
        if not mlb:
            return jsonify({'sucesso': False, 'erro': 'MLB é obrigatório'}), 400
        
        if not atributos:
            return jsonify({'sucesso': False, 'erro': 'Nenhum atributo para alterar'}), 400
        
        token = ml_token_manager.get_valid_token()
        headers = {'Authorization': f'Bearer {token}'}
        
        # Busca o produto atual
        response = requests.get(
            f'https://api.mercadolibre.com/items/{mlb}',
            headers=headers,
            timeout=15
        )
        
        if response.status_code != 200:
            return jsonify({'sucesso': False, 'erro': 'Produto não encontrado'}), 404
        
        dados = response.json()
        
        # Verifica se é produto de catálogo
        if dados.get('catalog_listing', False):
            return jsonify({
                'sucesso': False,
                'erro': '⚠️ Produto de CATÁLOGO! As alterações podem levar até 48h para serem validadas pelo Mercado Livre.'
            }), 400
        
        # Atualiza os atributos
        atributos_atuais = dados.get('attributes', [])
        atributos_alterados = []
        
        for attr in atributos_atuais:
            attr_id = attr.get('id')
            if attr_id in atributos:
                novo_valor = atributos[attr_id]
                if novo_valor:  # Só altera se tiver valor
                    attr['value_name'] = novo_valor
                    attr['value_id'] = None  # Limpa o ID quando usa nome personalizado
                    atributos_alterados.append(attr_id)
        
        # Adiciona atributos que não existiam
        for attr_id, novo_valor in atributos.items():
            if novo_valor:
                existe = False
                for attr in atributos_atuais:
                    if attr.get('id') == attr_id:
                        existe = True
                        break
                
                if not existe:
                    nome_map = {
                        'MODEL': 'Modelo', 'BRAND': 'Marca', 'COLOR': 'Cor',
                        'MATERIAL': 'Material', 'WARRANTY_TIME': 'Garantia',
                        'SELLER_SKU': 'SKU', 'GTIN': 'GTIN/EAN',
                        'HEIGHT': 'Altura', 'WIDTH': 'Largura',
                        'DEPTH': 'Profundidade', 'WEIGHT': 'Peso'
                    }
                    atributos_atuais.append({
                        'id': attr_id,
                        'name': nome_map.get(attr_id, attr_id),
                        'value_name': novo_valor,
                        'value_id': None,
                        'value_struct': None
                    })
                    atributos_alterados.append(attr_id)
        
        # Envia a atualização
        payload = {'attributes': atributos_atuais}
        
        response_put = requests.put(
            f'https://api.mercadolibre.com/items/{mlb}',
            headers=headers,
            json=payload,
            timeout=30
        )
        
        if response_put.status_code == 200:
            return jsonify({
                'sucesso': True,
                'mlb': mlb,
                'atributos_alterados': atributos_alterados,
                'mensagem': f'{len(atributos_alterados)} atributo(s) alterado(s) com sucesso!'
            })
        else:
            return jsonify({
                'sucesso': False,
                'erro': f'Erro ao atualizar: {response_put.status_code}',
                'detalhes': response_put.text[:500]
            }), response_put.status_code
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
    

@app.route('/mercadolivre/alterar-atributos')
@login_required
@permissao_modulo('produtos')
def alterar_atributos_ml():
    """Página para alterar múltiplos atributos do produto"""
    return render_template(
        'mercadolivre/alterar_atributos.html',
        active_page='alterar_atributos_ml',
        active_module='mercadolivre',
        page_title='Alterar Atributos - Mercado Livre'
    )
    
    #####################################################################################################################
@app.route('/api/mercadolivre/autenticar', methods=['POST'])
def api_autenticar_mercadolivre():
    """API para autenticar no Mercado Livre"""
    try:
        data = request.get_json()
        access_token = data.get('access_token')
        refresh_token = data.get('refresh_token')
        
        if not access_token or not refresh_token:
            return jsonify({'sucesso': False, 'erro': 'Access token e refresh token são obrigatórios'}), 400
        
        # Salva os tokens
        token_data = {
            'access_token': access_token,
            'refresh_token': refresh_token,
            'expires_in': 21600  # 6 horas
        }
        
        if ml_token_manager.save_tokens(token_data):
            # Testa a conexão
            if ml_api_secure.testar_conexao():
                return jsonify({
                    'sucesso': True, 
                    'mensagem': 'Autenticação realizada com sucesso!'
                })
            else:
                return jsonify({
                    'sucesso': False, 
                    'erro': 'Tokens inválidos ou sem permissão'
                })
        else:
            return jsonify({
                'sucesso': False, 
                'erro': 'Erro ao salvar tokens'
            })
            
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/status')
def api_status_mercadolivre():
    """API para verificar status da autenticação"""
    try:
        esta_autenticado = ml_token_manager.is_authenticated()
        conexao_ativa = False
        
        if esta_autenticado:
            conexao_ativa = ml_api_secure.testar_conexao()
        
        return jsonify({
            'sucesso': True,
            'autenticado': esta_autenticado,
            'conexao_ativa': conexao_ativa
        })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/buscar-mlb', methods=['POST'])
@login_required
def api_buscar_mlb():
    """API para buscar anúncios por MLB"""
    try:
        if not ml_token_manager.is_authenticated():
            return jsonify({'sucesso': False, 'erro': 'Não autenticado no Mercado Livre'}), 401
        
        data = request.get_json()
        mlbs = data.get('mlbs', [])
        tipo_busca = data.get('tipo_busca', 'mlbs')
        
        if tipo_busca == 'mlbs' and not mlbs:
            return jsonify({'sucesso': False, 'erro': 'Nenhum MLB fornecido'}), 400
        
        # Limpar e validar MLBs
        mlbs_validos = []
        for mlb in mlbs:
            mlb_limpo = mlb.strip().upper()
            if mlb_limpo.startswith('MLB'):
                mlbs_validos.append(mlb_limpo)
            else:
                # Tentar adicionar MLB se não tiver
                if mlb_limpo.replace('MLB', '').isalnum():
                    mlb_formatado = f"MLB{mlb_limpo.replace('MLB', '')}"
                    mlbs_validos.append(mlb_formatado)
        
        if tipo_busca == 'mlbs' and not mlbs_validos:
            return jsonify({'sucesso': False, 'erro': 'Nenhum MLB válido encontrado'}), 400
        
        # Buscar dados
        if tipo_busca == 'mlbs':
            resultado = ml_api_secure.buscar_anuncios_mlbs(mlbs_validos)
        elif tipo_busca == 'meus_anuncios':
            resultado = ml_api_secure.buscar_meus_anuncios()
        else:
            return jsonify({'sucesso': False, 'erro': 'Tipo de busca não suportado'}), 400
        
        return jsonify(resultado)
        
    except Exception as e:
        print(f"❌ Erro ao buscar MLBs: {str(e)}")
        return jsonify({'sucesso': False, 'erro': str(e)}), 500


@app.route('/api/mercadolivre/buscar-todos', methods=['POST'])
@login_required
def api_buscar_todos_anuncios():
    """
    Busca TODOS os anúncios do usuário com filtros opcionais.
 
    Body JSON esperado:
    {
        "status": "active",          // "active" | "paused" | "closed" | "all"
        "data_de": "2024-01-01",     // opcional, YYYY-MM-DD
        "data_ate": "2024-12-31",    // opcional, YYYY-MM-DD
        "limite": 500                // opcional, máximo de anúncios
    }
    """
    try:
        if not ml_token_manager.is_authenticated():
            return jsonify({'sucesso': False, 'erro': 'Não autenticado no Mercado Livre'}), 401
 
        data = request.get_json() or {}
 
        status    = data.get('status', 'active')
        data_de   = data.get('data_de')    # 'YYYY-MM-DD' ou None
        data_ate  = data.get('data_ate')   # 'YYYY-MM-DD' ou None
        limite    = data.get('limite')     # int ou None
 
        resultado = ml_api_secure.buscar_todos_anuncios(
            status=status,
            data_criacao_de=data_de,
            data_criacao_ate=data_ate,
            limite_total=int(limite) if limite else None
        )
 
        return jsonify(resultado)
 
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
@app.route('/api/mercadolivre/analisar-envio-manufacturing', methods=['POST'])
def api_analisar_envio_manufacturing():
    """API para análise específica de envio e manufacturing time"""
    try:
        if not ml_token_manager.is_authenticated():
            return jsonify({'sucesso': False, 'erro': 'Não autenticado no Mercado Livre'}), 401
        
        data = request.get_json()
        mlbs = data.get('mlbs', [])
        tipo_busca = data.get('tipo_busca', 'mlbs')
        
        # Buscar dados
        if tipo_busca == 'mlbs':
            resultado_busca = ml_api_secure.buscar_anuncios_mlbs(mlbs)
        elif tipo_busca == 'meus_anuncios':
            resultado_busca = ml_api_secure.buscar_meus_anuncios()
        else:
            return jsonify({'sucesso': False, 'erro': 'Tipo de busca não suportado'}), 400
        
        if not resultado_busca['sucesso']:
            return jsonify(resultado_busca)
        
        resultados = resultado_busca['resultados']
        resultados_validos = [r for r in resultados if 'error' not in r]
        
        # Estatísticas
        total_me2 = sum(1 for r in resultados_validos if r.get('shipping_mode') == 'me2')
        total_me1 = sum(1 for r in resultados_validos if r.get('shipping_mode') == 'me1')
        com_manufacturing = sum(1 for r in resultados_validos if r.get('manufacturing_time') not in [0, '0', 'N/A', None, ''])
        
        return jsonify({
            'sucesso': True,
            'estatisticas': {
                'total_analisado': len(resultados_validos),
                'me2': total_me2,
                'me1': total_me1,
                'com_manufacturing': com_manufacturing,
                'sem_manufacturing': len(resultados_validos) - com_manufacturing
            },
            'resultados': resultados,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        print(f"❌ Erro na análise: {str(e)}")
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/desautenticar', methods=['POST'])
def api_desautenticar_mercadolivre():
    """API para remover autenticação"""
    try:
        if ml_token_manager.remove_tokens():
            return jsonify({
                'sucesso': True,
                'mensagem': 'Desautenticado com sucesso'
            })
        else:
            return jsonify({
                'sucesso': False,
                'erro': 'Erro ao remover tokens'
            })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/configuracao')
def api_configuracao_mercadolivre():
    """API para obter configuração atual (sem dados sensíveis)"""
    try:
        configurado = ml_token_manager.client_id is not None and ml_token_manager.client_secret is not None
        
        return jsonify({
            'sucesso': True,
            'configurado': configurado
        })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
    

@app.route('/api/anymarket/pedidos/<int:order_id>')
def api_detalhes_pedido_anymarket(order_id):
    """API para buscar detalhes de um pedido específico"""
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'success': False, 'error': 'Token de autenticação não fornecido'}), 401
        
        token = auth_header.replace('Bearer ', '')
        
        url = f"https://api.anymarket.com.br/v2/orders/{order_id}"
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json',
            'gumgaToken': token
        }
        
        response = requests.get(url, headers=headers, timeout=30)
        
        if response.status_code != 200:
            return jsonify({
                'success': False, 
                'error': f'Erro na API AnyMarket: {response.status_code}'
            }), response.status_code
        
        order_data = response.json()
        
        return jsonify({
            'success': True,
            'order': order_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Erro interno: {str(e)}'}), 500

@app.route('/api/tokens/anymarket/obter')
def api_obter_token():
    """API para obter token do AnyMarket - VERSÃO TOLERANTE"""
    try:
        tokens_file = 'tokens_secure.json'
        
        # 🔹 CORREÇÃO: Se o arquivo não existe, retorna que não há token
        if not os.path.exists(tokens_file):
            return jsonify({
                'success': False, 
                'error': 'Token não configurado',
                'arquivo_existe': False
            }), 404
        
        with open(tokens_file, 'r', encoding='utf-8') as f:
            tokens = json.load(f)
        
        # Tenta estrutura nova primeiro
        token_data = tokens.get('anymarket')
        if token_data and token_data.get('token'):
            return jsonify({
                'success': True,
                'token': token_data['token'],
                'criado_em': token_data.get('criado_em'),
                'arquivo_existe': True
            })
        
        # Tenta estrutura antiga com IDs aleatórios
        for key, value in tokens.items():
            if isinstance(value, dict) and value.get('tipo') == 'anymarket' and value.get('token'):
                return jsonify({
                    'success': True,
                    'token': value['token'],
                    'criado_em': value.get('criado_em'),
                    'arquivo_existe': True,
                    'estrutura_antiga': True
                })
        
        return jsonify({
            'success': False, 
            'error': 'Token não encontrado',
            'arquivo_existe': True
        }), 404
        
    except json.JSONDecodeError:
        return jsonify({
            'success': False, 
            'error': 'Arquivo de tokens corrompido',
            'arquivo_existe': True
        }), 500
    except Exception as e:
        return jsonify({
            'success': False, 
            'error': str(e),
            'arquivo_existe': os.path.exists('tokens_secure.json')
        }), 500

    
@app.route('/uploads/<filename>')
def baixar_arquivo(filename):
    try:
        # Decodifica caracteres especiais na URL
        from urllib.parse import unquote
        filename = unquote(filename)
        
        upload_folder = app.config['UPLOAD_FOLDER']
        safe_filename = secure_filename(os.path.basename(filename))
        file_path = os.path.join(upload_folder, safe_filename)
        
        if not os.path.exists(file_path):
            # Tenta encontrar o arquivo sem sanitização (para compatibilidade)
            for f in os.listdir(upload_folder):
                if f.startswith(os.path.splitext(safe_filename)[0]):
                    file_path = os.path.join(upload_folder, f)
                    break
            
            if not os.path.exists(file_path):
                abort(404)
        
        return send_from_directory(
            upload_folder,
            os.path.basename(file_path),
            as_attachment=True,
            download_name=safe_filename  # Força o nome no download
        )
        
    except Exception as e:
        app.logger.error(f"Erro ao baixar {filename}: {str(e)}")
        abort(500)

@app.route('/configuracoes/tokens')
@login_required
def configurar_tokens():
    config = carregar_configuracao_google_sheets()
    
    # Verifica se tem token do AnyMarket configurado
    token_anymarket_configurado = verificar_token_anymarket_configurado()
    
    # 🔹 VERIFICA SE TEM TOKEN DO INTELIPOST CONFIGURADO
    token_intelipost_configurado = verificar_token_intelipost_configurado()
    
    return render_template(
        "config_tokens.html",
        active_page='configuracao',
        config=config,
        token_anymarket_configurado=token_anymarket_configurado,
        token_intelipost_configurado=token_intelipost_configurado,  # 🔹 NOVO
        page_title='Configuração de Tokens'
    )

def verificar_token_intelipost_configurado():
    """Verifica se o token do Intelipost está configurado"""
    try:
        tokens_file = 'tokens_secure.json'
        if not os.path.exists(tokens_file):
            return False
        
        with open(tokens_file, 'r', encoding='utf-8') as f:
            tokens = json.load(f)
        
        # Procura por token intelipost
        if 'intelipost' in tokens and tokens['intelipost'].get('api_key'):
            return True
        
        # Tenta estrutura antiga
        for key, value in tokens.items():
            if isinstance(value, dict) and value.get('tipo') == 'intelipost' and value.get('api_key'):
                return True
        
        return False
        
    except Exception:
        return False

@app.route('/api/tokens/intelipost/salvar', methods=['POST'])
def salvar_token_intelipost():
    """Salva token do Intelipost no arquivo seguro"""
    try:
        data = request.get_json()
        api_key = data.get('api_key')
        
        if not api_key:
            return jsonify({'sucesso': False, 'erro': 'API Key é obrigatória'}), 400
        
        tokens_file = 'tokens_secure.json'
        tokens = {}
        
        # Carrega tokens existentes
        if os.path.exists(tokens_file):
            try:
                with open(tokens_file, 'r', encoding='utf-8') as f:
                    tokens = json.load(f)
            except json.JSONDecodeError:
                tokens = {}
        
        # Salva token Intelipost
        tokens['intelipost'] = {
            'api_key': api_key,
            'criado_em': datetime.now().isoformat(),
            'ultima_atualizacao': datetime.now().isoformat(),
            'tipo': 'intelipost',
            'descricao': 'API Key para rastreamento Intelipost'
        }
        
        # Salva arquivo
        with open(tokens_file, 'w', encoding='utf-8') as f:
            json.dump(tokens, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Token Intelipost salvo em {tokens_file}")
        
        return jsonify({
            'sucesso': True, 
            'mensagem': 'API Key do Intelipost salva com sucesso!',
            'salvo_em': tokens_file
        })
        
    except Exception as e:
        print(f"❌ Erro ao salvar token Intelipost: {str(e)}")
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/tokens/intelipost/obter')
def obter_token_intelipost():
    """Obtém status do token Intelipost (não retorna a chave)"""
    try:
        tokens_file = 'tokens_secure.json'
        
        if not os.path.exists(tokens_file):
            return jsonify({
                'sucesso': False,
                'configurado': False,
                'mensagem': 'Token não configurado'
            })
        
        with open(tokens_file, 'r', encoding='utf-8') as f:
            tokens = json.load(f)
        
        # Procura token Intelipost
        token_configurado = False
        token_data = None
        
        if 'intelipost' in tokens:
            token_data = tokens['intelipost']
            token_configurado = bool(token_data.get('api_key'))
        else:
            # Tenta estrutura antiga
            for key, value in tokens.items():
                if isinstance(value, dict) and value.get('tipo') == 'intelipost' and value.get('api_key'):
                    token_data = value
                    token_configurado = True
                    break
        
        return jsonify({
            'sucesso': True,
            'configurado': token_configurado,
            'criado_em': token_data.get('criado_em') if token_data else None,
            'mensagem': 'Token configurado' if token_configurado else 'Token não configurado',
            'caracteres': len(token_data.get('api_key', '')) if token_data and token_data.get('api_key') else 0
        })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/tokens/intelipost/testar', methods=['POST'])
def testar_token_intelipost():
    """Testa o token do Intelipost - VERSÃO DEFINITIVA"""
    try:
        data = request.get_json()
        api_key = data.get('api_key', '').strip()
        
        if not api_key:
            return jsonify({'sucesso': False, 'erro': 'API Key é obrigatória'}), 400
        
        print(f"🧪 Testando API Key Intelipost...")
        print(f"🔑 Tamanho: {len(api_key)} caracteres")
        
        try:
            from processamento.intelipost_api import IntelipostAPI
            
            # Cria cliente
            api_client = IntelipostAPI(api_key=api_key)
            
            # 🔹 TESTE 1: Conexão básica
            print("🔄 Teste 1: Conexão básica...")
            resultado = api_client.testar_conexao()
            
            if resultado.get('sucesso') and resultado.get('conectado'):
                # 🔹 TESTE 2: Com pedido de teste (se quiser)
                print("🔄 Teste 2: Com pedido de teste...")
                resultado_teste = api_client.testar_com_pedido_real("PEDIDO0001")
                
                if resultado_teste.get('sucesso'):
                    resultado.update(resultado_teste)
                    resultado['teste_pedido'] = 'bem_sucedido'
                else:
                    resultado['teste_pedido'] = 'falhou_mas_api_funciona'
                    resultado['aviso'] = 'API funciona mas pedido de teste não encontrado'
            
            print(f"📊 Resultado final: {resultado}")
            return jsonify(resultado)
            
        except ImportError:
            return jsonify({
                'sucesso': False,
                'conectado': False,
                'erro': 'Módulo Intelipost não encontrado',
                'sugestao': 'Verifique se processamento/intelipost_api.py existe'
            })
        except Exception as api_error:
            print(f"❌ Erro na API: {str(api_error)}")
            return jsonify({
                'sucesso': False,
                'conectado': False,
                'erro': f'Erro: {str(api_error)}'
            })
        
    except Exception as e:
        print(f"❌ Erro geral: {str(e)}")
        return jsonify({
            'sucesso': False,
            'conectado': False,
            'erro': f'Erro interno: {str(e)}'
        })


    
@app.route('/api/tokens/intelipost/remover', methods=['POST'])
def remover_token_intelipost():
    """Remove token do Intelipost"""
    try:
        tokens_file = 'tokens_secure.json'
        
        if os.path.exists(tokens_file):
            with open(tokens_file, 'r', encoding='utf-8') as f:
                tokens = json.load(f)
            
            if 'intelipost' in tokens:
                del tokens['intelipost']
                
                with open(tokens_file, 'w', encoding='utf-8') as f:
                    json.dump(tokens, f, indent=2, ensure_ascii=False)
        
        return jsonify({'sucesso': True, 'mensagem': 'Token Intelipost removido'})
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500


def verificar_token_anymarket_configurado():
    """Verifica se o token do AnyMarket está configurado (sem retornar o token)"""
    try:
        tokens_file = 'tokens_secure.json'
        if not os.path.exists(tokens_file):
            return False
        
        with open(tokens_file, 'r') as f:
            tokens = json.load(f)
        
        return 'anymarket' in tokens and 'token' in tokens['anymarket']
        
    except Exception:
        return False
    
@app.route('/api/anymarket/testar-token', methods=['POST'])
def testar_token_anymarket():
    """API para testar o token do AnyMarket - VERSÃO ROBUSTA"""
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'success': False, 'error': 'Token de autenticação não fornecido'}), 401
        
        token = auth_header.replace('Bearer ', '')
        
        # 🔹 MÚLTIPLAS TENTATIVAS COM DIFERENTES PARÂMETROS
        test_cases = [
            # Caso 1: Com datas específicas e limit=5
            {
                'params': {
                    'page': 1,
                    'limit': 5,
                    'createdAt.start': '2024-01-01T00:00:00-03:00',
                    'createdAt.end': '2024-12-31T23:59:59-03:00'
                },
                'description': 'Com datas fixas'
            },
            # Caso 2: Apenas paginação básica
            {
                'params': {
                    'page': 1,
                    'limit': 5
                },
                'description': 'Paginação básica'
            },
            # Caso 3: Sem parâmetros (API usa defaults)
            {
                'params': {},
                'description': 'Sem parâmetros'
            }
        ]
        
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json',
            'gumgaToken': token
        }
        
        url = "https://api.anymarket.com.br/v2/orders"
        
        for i, test_case in enumerate(test_cases):
            try:
                print(f"🧪 Teste {i+1}: {test_case['description']}")
                response = requests.get(url, params=test_case['params'], headers=headers, timeout=10)
                
                if 200 <= response.status_code < 300:
                    return jsonify({
                        'success': True, 
                        'message': f'Token válido! Conexão estabelecida com a API AnyMarket',
                        'status_code': response.status_code,
                        'test_used': test_case['description']
                    })
                
            except requests.exceptions.RequestException:
                continue  # Tenta o próximo caso se houver erro de conexão
        
        # Se nenhum caso funcionou, retorna o último erro
        return jsonify({
            'success': False, 
            'error': f'Não foi possível conectar à API AnyMarket. Status: {response.status_code}',
            'status_code': response.status_code
        }), response.status_code
            
    except Exception as e:
        return jsonify({
            'success': False, 
            'error': f'Erro ao testar token: {str(e)}'
        }), 500
        

@app.route("/consultar-anymarket", methods=["GET", "POST"])
def consultar_anymarket():
    resultado = None
    acao = None
    
    if request.method == "POST":
        acao = request.form.get('action', 'consultar')
        
        if acao == 'consultar':
            try:
                product_id = request.form.get('product_id', '').strip()
                api_token = request.form.get('api_token', '').strip()
                
                if not product_id:
                    flash("ID do produto é obrigatório", "danger")
                    return redirect(url_for('consultar_anymarket'))
                
                # ✅ Se não forneceu token, usa o seguro
                if not api_token:
                    from processamento.api_anymarket import obter_token_anymarket_seguro
                    api_token = obter_token_anymarket_seguro()
                
                inicio = datetime.now()
                resultado = consultar_api_anymarket(product_id, api_token)
                tempo_segundos = (datetime.now() - inicio).total_seconds()
                
                registrar_processo(
                    modulo="anymarket",
                    qtd_itens=resultado.get('quantidade_fotos', 0),
                    tempo_execucao=tempo_segundos,
                    status="sucesso" if resultado.get('sucesso') else "erro",
                    erro_mensagem=resultado.get('erro') if not resultado.get('sucesso') else None
                )
                
                if resultado.get('sucesso'):
                    flash(f"Consulta realizada com sucesso! {resultado.get('quantidade_fotos', 0)} fotos encontradas.", "success")
                else:
                    flash(f"Erro na consulta: {resultado.get('erro', 'Erro desconhecido')}", "danger")
                    
            except Exception as e:
                flash(f"Erro ao consultar API: {str(e)}", "danger")
        
        elif acao == 'excluir_lote':
            try:
                if 'planilha' not in request.files:
                    flash("Nenhum arquivo enviado", "danger")
                    return redirect(url_for('consultar_anymarket'))
                
                planilha = request.files['planilha']
                if planilha.filename == '':
                    flash("Nenhum arquivo selecionado", "danger")
                    return redirect(url_for('consultar_anymarket'))
                
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                nome_arquivo = secure_filename(planilha.filename)
                caminho_planilha = os.path.join(app.config['UPLOAD_FOLDER'], nome_arquivo)
                planilha.save(caminho_planilha)
                
                inicio = datetime.now()
                resultado = excluir_fotos_planilha_anymarket(caminho_planilha)
                tempo_segundos = (datetime.now() - inicio).total_seconds()
                
                registrar_processo(
                    modulo="anymarket_exclusao",
                    qtd_itens=resultado.get('total_processado', 0),
                    tempo_execucao=tempo_segundos,
                    status="sucesso" if resultado.get('sucesso') else "erro",
                    erro_mensagem=resultado.get('erro') if not resultado.get('sucesso') else None
                )
                
                if resultado.get('sucesso'):
                    flash(f"Exclusão em lote concluída! {resultado.get('total_sucesso', 0)} de {resultado.get('total_processado', 0)} fotos excluídas.", "success")
                else:
                    flash(f"Erro na exclusão em lote: {resultado.get('erro', 'Erro desconhecido')}", "danger")
                    
            except Exception as e:
                flash(f"Erro ao processar planilha: {str(e)}", "danger")
    
    return render_template(
        "consultar_anymarket.html",
        active_page='consultar_anymarket',
        active_module='anymarket',
        resultado=resultado,
        acao=acao,
        historico_processos=obter_historico_processos("anymarket"),
        processos_hoje=contar_processos_hoje("anymarket"),
        stats=get_processing_stats("anymarket"),
        page_title='Consulta Fotos'
    )

@app.route('/api/tokens/anymarket/obter', methods=['GET'])
def obter_token_anymarket():
    """Obtém token do AnyMarket do arquivo seguro"""
    try:
        tokens_file = 'tokens_secure.json'
        if not os.path.exists(tokens_file):
            return jsonify({'success': False, 'error': 'Token não configurado'}), 404
        
        with open(tokens_file, 'r', encoding='utf-8') as f:
            tokens = json.load(f)
        
        token_data = tokens.get('anymarket')
        if not token_data or not token_data.get('token'):
            return jsonify({'success': False, 'error': 'Token não encontrado'}), 404
        
        return jsonify({
            'success': True,
            'token': token_data['token'],
            'criado_em': token_data.get('criado_em')
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/metrics/teste')
def api_metrics_teste():
    """Rota de teste que retorna dados fixos para verificar o frontend"""
    return jsonify({
        'sucesso': True,
        'dados': {
            'mercadolivre': {
                'status': 'online',
                'vendas_7d': 15430.00,
                'pedidos_7d': 47,
                'ticket_medio': 328.30,
                'anuncios_ativos': 124
            },
            'anymarket': {
                'status': 'online',
                'vendas_7d': 8230.00,
                'pedidos_7d': 23,
                'ticket_medio': 357.83,
                'marketplaces_ativos': 3,
                'total_itens': 45,
                'fotos_erro': 3
            },
            'intelipost': {
                'status': 'online',
                'em_transito': 15,
                'entregues_7d': 42,
                'atrasados': 2,
                'prazo_medio': 3.2
            },
            'sistema': {
                'processamentos_hoje': 12,
                'ultima_planilha': 'exemplo.xlsx',
                'ultima_planilha_data': '2024-01-15 14:30:00',
                'stats_modulos': {
                    'cadastro': 5,
                    'atributos': 3,
                    'prazos': 2,
                    'anymarket': 2
                }
            }
        },
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/teste-planilha')
def api_teste_planilha():
    """Rota para testar e ver os dados da planilha"""
    try:
        from google_sheets import ler_planilha_google
        
        # CONFIGURE AQUI - Substitua pelos seus dados
        SHEET_ID = '1JMcU1mhbW0Q2IyLo15wHAoFLEY3IdFLEG0bMssP6y6o'
        ABA_NOME = 'Sequência Cad.'  # Nome da sua aba
        
        # Lê a planilha
        df = ler_planilha_google(SHEET_ID, ABA_NOME)
        
        # Converte para dicionário
        dados = df.to_dict(orient='records')
        
        # Estatísticas básicas
        info = {
            'total_linhas': len(df),
            'total_colunas': len(df.columns),
            'colunas': df.columns.tolist(),
            'primeiras_5_linhas': dados[:5],
            'amostra': dados[0] if dados else None
        }
        
        return jsonify({
            'sucesso': True,
            'info': info
        })
        
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500
    
@app.route('/api/tokens/anymarket/salvar', methods=['POST'])
def salvar_token_anymarket():
    """Salva token do AnyMarket no arquivo seguro - VERSÃO QUE CRIA ARQUIVO"""
    try:
        data = request.get_json()
        token = data.get('token')
        
        if not token:
            return jsonify({'success': False, 'error': 'Token não fornecido'}), 400
        
        tokens_file = 'tokens_secure.json'
        tokens = {}
        
        # 🔹 CORREÇÃO: Se o arquivo existe, carrega. Se não, cria estrutura vazia.
        if os.path.exists(tokens_file):
            try:
                with open(tokens_file, 'r', encoding='utf-8') as f:
                    tokens = json.load(f)
            except json.JSONDecodeError:
                # Se o arquivo estiver corrompido, recria
                tokens = {}
        else:
            # Arquivo não existe - cria estrutura vazia
            tokens = {}
            print("📁 Arquivo tokens_secure.json não encontrado - criando novo...")
        
        # Garante que a estrutura tenha o objeto anymarket
        tokens['anymarket'] = {
            'token': token,
            'criado_em': datetime.now().isoformat(),
            'ultimo_uso': datetime.now().isoformat()
        }
        
        # 🔹 CORREÇÃO: Garante que o diretório existe
        os.makedirs(os.path.dirname(tokens_file) or '.', exist_ok=True)
        
        # Salva o arquivo
        with open(tokens_file, 'w', encoding='utf-8') as f:
            json.dump(tokens, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Token salvo com segurança em {tokens_file}")
        
        return jsonify({
            'success': True, 
            'message': 'Token salvo com segurança',
            'arquivo_criado': not os.path.exists(tokens_file)  # Indica se foi criado agora
        })
        
    except Exception as e:
        print(f"❌ Erro ao salvar token: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/tokens/anymarket/remover', methods=['POST'])
def remover_token_anymarket():
    """Remove token do AnyMarket"""
    try:
        tokens_file = 'tokens_secure.json'
        
        if os.path.exists(tokens_file):
            with open(tokens_file, 'r', encoding='utf-8') as f:
                tokens = json.load(f)
            
            if 'anymarket' in tokens:
                del tokens['anymarket']
                
                with open(tokens_file, 'w', encoding='utf-8') as f:
                    json.dump(tokens, f, indent=2, ensure_ascii=False)
        
        return jsonify({'success': True, 'message': 'Token removido'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


    
@app.route('/verificar-status-mlb/<mlb>')
def verificar_status_mlb(mlb):
    """
    Verifica o status real de um MLB no Mercado Livre
    Útil para diagnosticar exclusões
    """
    try:
        if not ml_token_manager.is_authenticated():
            return jsonify({
                'sucesso': False,
                'erro': 'Não autenticado'
            }), 401
        
        headers = ml_api_secure._get_headers()
        
        # Tenta acessar o item diretamente
        response = requests.get(
            f"https://api.mercadolibre.com/items/{mlb}",
            headers=headers,
            timeout=10
        )
        
        resultado = {
            'mlb': mlb,
            'status_code': response.status_code,
            'existe': response.status_code == 200,
            'mensagem': ''
        }
        
        if response.status_code == 200:
            data = response.json()
            resultado['status'] = data.get('status')
            resultado['sub_status'] = data.get('sub_status', [])
            resultado['titulo'] = data.get('title')
            
            # Verifica se está marcado como deletado
            if 'deleted' in data.get('sub_status', []):
                resultado['mensagem'] = '✅ Anúncio está marcado como DELETADO no sistema'
            else:
                resultado['mensagem'] = '⚠️ Anúncio ainda existe (não foi deletado)'
        elif response.status_code == 404:
            resultado['mensagem'] = '✅ Anúncio NÃO EXISTE mais (excluído com sucesso)'
        else:
            resultado['mensagem'] = f'❌ Erro na consulta: {response.status_code}'
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500
      
@app.route("/excluir-foto-anymarket", methods=["POST"])
def excluir_foto_anymarket_route():
    """API para exclusão individual de foto"""
    try:
        data = request.get_json()
        product_id = data.get('product_id')
        photo_id = data.get('photo_id')
        
        if not product_id or not photo_id:
            return jsonify({'sucesso': False, 'erro': 'ID do produto e da foto são obrigatórios'}), 400
        
        resultado = excluir_foto_anymarket(product_id, photo_id)
        
        registrar_processo(
            modulo="anymarket_exclusao",
            qtd_itens=1,
            tempo_execucao=0,
            status="sucesso" if resultado.get('sucesso') else "erro",
            erro_mensagem=resultado.get('erro') if not resultado.get('sucesso') else None
        )
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route("/excluir-fotos-lote", methods=["POST"])
def excluir_fotos_lote_route():
    """API para exclusão em lote de fotos"""
    try:
        data = request.get_json()
        fotos = data.get('fotos', [])
        
        if not fotos:
            return jsonify({'sucesso': False, 'erro': 'Nenhuma foto selecionada'}), 400
        
        total_sucesso = 0
        total_erro = 0
        resultados = []
        
        for foto in fotos:
            product_id = foto.get('product_id')
            photo_id = foto.get('photo_id')
            
            if product_id and photo_id:
                resultado = excluir_foto_anymarket(product_id, photo_id)
                resultados.append(resultado)
                
                if resultado.get('sucesso'):
                    total_sucesso += 1
                else:
                    total_erro += 1
        
        registrar_processo(
            modulo="anymarket_exclusao",
            qtd_itens=len(fotos),
            tempo_execucao=0,
            status="sucesso" if total_erro == 0 else "parcial",
            erro_mensagem=f"{total_erro} erro(s)" if total_erro > 0 else None
        )
        
        return jsonify({
            'sucesso': True,
            'total_processado': len(fotos),
            'total_sucesso': total_sucesso,
            'total_erro': total_erro,
            'resultados': resultados
        })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
   

def limpar_moeda(valor):
    if pd.isna(valor):
        return None

    valor = str(valor)

    return (
        valor
        .replace("R$", "")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", ".")
    )
    
@app.route("/preencher-planilha", methods=["GET", "POST"])
@login_required
@permissao_modulo('produtos')
def preencher_planilha():
    nome_arquivo_saida = None
    config = carregar_configuracao_google_sheets()
    abas = []
    sheet_id_input = request.form.get('sheet_id', config.get('sheet_id', ''))
    aba_selecionada = request.form.get('aba_nome', '')
    preview_data = None

    # Aba ativa (upload ou google)
    aba_ativa = request.args.get('aba', 'upload')

    if request.method == "POST":
        action_type = request.form.get('action_type', '')

        try:
            # ====================================
            # 🔹 LISTAR ABAS DO GOOGLE SHEETS
            # ====================================
            if action_type == "listar_abas":
                sheet_id = request.form.get('sheet_id', '').strip()
                if not sheet_id:
                    flash("Informe o ID da planilha Google", "danger")
                    return redirect(url_for("preencher_planilha", aba="google"))

                abas = listar_abas_google_sheets(sheet_id)
                config = carregar_configuracao_google_sheets()

                return render_template(
                    "preencher_planilha.html",
                    historico_processos=obter_historico_processos("cadastro"),
                    processos_hoje=contar_processos_hoje("cadastro"),
                    stats=get_processing_stats("cadastro"),
                    nome_arquivo_saida=None,
                    config=config,
                    abas=abas,
                    sheet_id_input=sheet_id,
                    aba_selecionada=None,
                    aba_ativa="google",
                    page_title='Cadastro Produto'
                )

            # ====================================
            # 🔹 PREVIEW DA ABA
            # ====================================
            elif action_type == "preview_aba":
                sheet_id = request.form.get('sheet_id', '').strip()
                aba_nome = request.form.get('aba_nome', '').strip()

                if not sheet_id or not aba_nome:
                    flash("ID da planilha e aba são obrigatórios para preview", "danger")
                    return redirect(url_for("preencher_planilha", aba="google"))

                df_preview = ler_planilha_google(sheet_id, aba_nome)
                preview_data = {
                    "total_linhas": len(df_preview),
                    "total_colunas": len(df_preview.columns),
                    "colunas": df_preview.columns.tolist(),
                    "linhas": df_preview.head(10).to_dict(orient="records")
                }

                return render_template(
                    "preencher_planilha.html",
                    historico_processos=obter_historico_processos("cadastro"),
                    processos_hoje=contar_processos_hoje("cadastro"),
                    stats=get_processing_stats("cadastro"),
                    nome_arquivo_saida=None,
                    config=config,
                    abas=listar_abas_google_sheets(sheet_id),
                    sheet_id_input=sheet_id,
                    aba_selecionada=aba_nome,
                    aba_ativa="google",
                    preview_data=preview_data,
                    page_title='Cadastro Produto'
                )

            # ====================================
            # 🔹 PROCESSAR (GOOGLE SHEETS) - AGORA SEM ARQUIVO DESTINO
            # ====================================
            elif action_type == "conectar_google":
                sheet_id = request.form.get('sheet_id', '').strip()
                aba_nome = request.form.get('aba_nome', '').strip()

                if not sheet_id or not aba_nome:
                    flash("ID da planilha e aba são obrigatórios", "danger")
                    return redirect(url_for("preencher_planilha", aba="google"))

                salvar_configuracao_google_sheets(sheet_id, aba_nome)
                config = carregar_configuracao_google_sheets()

                # 🔹 AGORA USA O MODELO FIXO - NÃO PRECISA DE UPLOAD
                # Processa usando apenas o Google Sheets como origem
                arquivo_saida, qtd_produtos, tempo_segundos, produtos_processados = executar_processamento(
                    {"sheet_id": sheet_id, "aba": aba_nome}
                    # 🔹 Não passa planilha_destino - usa o modelo fixo
                )

                nome_arquivo_saida = os.path.basename(arquivo_saida)

                registrar_processo(
                    modulo="cadastro",
                    qtd_itens=qtd_produtos,
                    tempo_execucao=tempo_segundos,
                    status="sucesso"
                )
                registrar_itens_processados("cadastro", produtos_processados)

                flash("Cadastro concluído com sucesso a partir do Google Sheets!", "success")
                aba_ativa = "google"

            # ====================================
            # 🔹 PROCESSAR (UPLOAD LOCAL) - AGORA APENAS ARQUIVO ORIGEM
            # ====================================
            elif 'arquivo_origem' in request.files:
                origem = request.files["arquivo_origem"]

                if origem.filename == '':
                    flash("Nenhum arquivo de origem selecionado", "danger")
                    registrar_processo(
                        modulo="cadastro",
                        qtd_itens=0,
                        tempo_execucao=0,
                        status="erro",
                        erro_mensagem="Nenhum arquivo de origem selecionado"
                    )
                    return redirect(url_for("preencher_planilha", aba="upload"))

                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

                nome_origem = secure_filename(origem.filename)
                caminho_origem = os.path.join(app.config['UPLOAD_FOLDER'], nome_origem)
                origem.save(caminho_origem)

                # 🔹 AGORA USA O MODELO FIXO - NÃO PRECISA DE ARQUIVO DESTINO
                arquivo_saida, qtd_produtos, tempo_segundos, produtos_processados = executar_processamento(
                    caminho_origem
                    # 🔹 Não passa planilha_destino - usa o modelo fixo
                )

                nome_arquivo_saida = os.path.basename(arquivo_saida)

                registrar_processo(
                    modulo="cadastro",
                    qtd_itens=qtd_produtos,
                    tempo_execucao=tempo_segundos,
                    status="sucesso"
                )
                registrar_itens_processados("cadastro", produtos_processados)

                flash("Planilha preenchida com sucesso usando modelo fixo!", "success")
                aba_ativa = "upload"

        except Exception as e:
            erro_msg = str(e)
            if "faltando as seguintes colunas" in erro_msg:
                colunas_faltando = erro_msg.split(":")[1].strip()
                erro_msg = f"Planilha fora do padrão. Colunas faltantes: {colunas_faltando}"

            registrar_processo(
                modulo="cadastro",
                qtd_itens=0,
                tempo_execucao=0,
                status="erro",
                erro_mensagem=erro_msg
            )
            flash(f"Erro: {erro_msg}", "danger")
            return redirect(url_for("preencher_planilha", aba=aba_ativa))

    return render_template(
        "preencher_planilha.html",
        active_page='preencher_planilha',
        active_module='cadastro',
        historico_processos=obter_historico_processos("cadastro"),
        processos_hoje=contar_processos_hoje("cadastro"),
        stats=get_processing_stats("cadastro"),
        nome_arquivo_saida=nome_arquivo_saida,
        config=config,
        abas=abas,
        sheet_id_input=sheet_id_input,
        aba_selecionada=aba_selecionada,
        aba_ativa=aba_ativa,
        preview_data=preview_data,
        page_title='Cadastro Produto'
    )


@app.route("/extrair-atributos", methods=["GET", "POST"])
def extrair_atributos():
    nome_arquivo_saida = None
    config = carregar_configuracao_google_sheets()
    abas = []
    preview_data = None
    sheet_id_input = request.form.get('sheet_id', config.get('sheet_id', ''))
    aba_selecionada = request.form.get('aba_nome', '')
    
    # Verifica se deve manter a aba Google ativa
    aba_ativa = request.args.get('aba', 'upload')  # 'upload' ou 'google'
    
    try:
        if request.method == "POST":
            action_type = request.form.get('action_type', '')
            
            # Se for para listar abas - USA FUNÇÃO COMPATÍVEL
            if action_type == 'listar_abas':
                sheet_id = request.form.get('sheet_id', '').strip()
                if sheet_id:
                    try:
                        # CORREÇÃO: Usa a função original para compatibilidade
                        from google_sheets_utils import listar_abas_google_sheets
                        abas = listar_abas_google_sheets(sheet_id)
                        flash(f"{len(abas)} abas encontradas", "success")
                        sheet_id_input = sheet_id
                        aba_ativa = 'google'  # Mantém na aba Google
                    except Exception as e:
                        flash(f"Erro ao listar abas: {str(e)}", "danger")
                else:
                    flash("Informe o ID da planilha primeiro", "warning")
            
            # Se for para fazer preview de uma aba
            elif action_type == 'preview_aba':
                sheet_id = request.form.get('sheet_id', '').strip()
                aba_nome = request.form.get('aba_nome', '').strip()
                if sheet_id and aba_nome:
                    try:
                        from google_sheets_utils import obter_dados_aba
                        preview_data = obter_dados_aba(sheet_id, aba_nome)
                        flash(f"Preview da aba '{aba_nome}' carregado", "success")
                        sheet_id_input = sheet_id
                        aba_selecionada = aba_nome
                        aba_ativa = 'google'  # Mantém na aba Google
                    except Exception as e:
                        flash(f"Erro ao carregar preview: {str(e)}", "danger")
                else:
                    flash("Selecione uma aba para visualizar", "warning")
            
            # Se for para fazer preview de uma aba
            elif action_type == 'preview_aba':
                sheet_id = request.form.get('sheet_id', '').strip()
                aba_nome = request.form.get('aba_nome', '').strip()
                if sheet_id and aba_nome:
                    try:
                        from google_sheets_utils import obter_dados_aba
                        preview_data = obter_dados_aba(sheet_id, aba_nome)
                        flash(f"Preview da aba '{aba_nome}' carregado", "success")
                        sheet_id_input = sheet_id
                        aba_selecionada = aba_nome
                        aba_ativa = 'google'  # Mantém na aba Google
                    except Exception as e:
                        flash(f"Erro ao carregar preview: {str(e)}", "danger")
                else:
                    flash("Selecione uma aba para visualizar", "warning")
            
            # Se for para processar com Google Sheets
            elif action_type == 'conectar_google':
                sheet_id = request.form.get('sheet_id', '').strip()
                aba_nome = request.form.get('aba_nome', '').strip()
                
                if not sheet_id or not aba_nome:
                    flash("ID da planilha e aba são obrigatórios", "danger")
                    return redirect(url_for("extrair_atributos", aba='google'))
                
                # Salva a configuração completa
                salvar_configuracao_google_sheets(sheet_id, aba_nome)
                config = carregar_configuracao_google_sheets()
                
                inicio = datetime.now()
                caminho_saida, qtd_itens, tempo_segundos, _ = extrair_atributos_processamento({
                    'sheet_id': sheet_id,
                    'aba': aba_nome
                })
                
                nome_arquivo_saida = os.path.basename(caminho_saida)
                
                registrar_processo(
                    modulo="atributos",
                    qtd_itens=qtd_itens,
                    tempo_execucao=tempo_segundos,
                    status="sucesso"
                )
                
                flash("Extração do Google Sheets concluída com sucesso!", "success")
                aba_ativa = 'google'  # Mantém na aba Google
            
            # Modo upload de arquivo (apenas se for submit do formulário de upload)
            elif 'arquivo' in request.files:
                arquivo = request.files["arquivo"]
                if arquivo.filename == '':
                    flash("Nenhum arquivo selecionado", "danger")
                    registrar_processo(
                        modulo="atributos",
                        qtd_itens=0,
                        tempo_execucao=0,
                        status="erro",
                        erro_mensagem="Nenhum arquivo selecionado"
                    )
                    return redirect(url_for("extrair_atributos", aba='upload'))
                
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                nome_arquivo = secure_filename(arquivo.filename)
                caminho_arquivo = os.path.join(app.config['UPLOAD_FOLDER'], nome_arquivo)
                arquivo.save(caminho_arquivo)

                caminho_saida, qtd_itens, tempo_segundos, _ = extrair_atributos_processamento(
                    caminho_arquivo
                )
                
                nome_arquivo_saida = os.path.basename(caminho_saida)
                
                registrar_processo(
                    modulo="atributos",
                    qtd_itens=qtd_itens,
                    tempo_execucao=tempo_segundos,
                    status="sucesso"
                )
                
                flash("Extração concluída com sucesso!", "success")
                aba_ativa = 'upload'  # Mantém na aba Upload
    
    except Exception as e:
        erro_msg = str(e)
        if "faltando as seguintes colunas" in erro_msg:
            colunas_faltando = erro_msg.split(":")[1].strip()
            erro_msg = f"Planilha fora do padrão. Colunas faltantes: {colunas_faltando}"
        
        registrar_processo(
            modulo="atributos",
            qtd_itens=0,
            tempo_execucao=0,
            status="erro",
            erro_mensagem=erro_msg
        )
        flash(f"Erro: {erro_msg}", "danger")
    
    return render_template(
        "extrair_atributos.html",
        active_page='extrair_atributos',
        active_module='cadastro',
        historico_processos=obter_historico_processos("atributos"),
        processos_hoje=contar_processos_hoje("atributos"),
        stats=get_processing_stats("atributos"),
        nome_arquivo_saida=nome_arquivo_saida,
        config=config,
        abas=abas,
        preview_data=preview_data,
        sheet_id_input=sheet_id_input,
        aba_selecionada=aba_selecionada,
        aba_ativa=aba_ativa,
        page_title='Extração de Atributos'
    )

@app.route('/api/dashboard/sequencia-cadastro')
def api_dashboard_sequencia_cadastro():
    """
    API que retorna métricas da planilha Sequência Cad. para o dashboard
    """
    try:
        from google_sheets import ler_planilha_google
        from datetime import datetime, timedelta
        import pandas as pd
        
        # CONFIGURAÇÃO - Use o mesmo ID da sua rota de teste
        SHEET_ID = '1JMcU1mhbW0Q2IyLo15wHAoFLEY3IdFLEG0bMssP6y6o'
        ABA_NOME = 'Sequência Cad.'
        
        logger.info(f"📊 Buscando dados da planilha: {ABA_NOME}")
        
        # Lê a planilha
        df = ler_planilha_google(SHEET_ID, ABA_NOME)
        
        if df.empty:
            return jsonify({
                'sucesso': False,
                'erro': 'Planilha vazia',
                'dados': {}
            })
        
        logger.info(f"✅ {len(df)} linhas encontradas")
        logger.info(f"📋 Colunas: {df.columns.tolist()}")
        
        # ============================================
        # MÉTRICAS BÁSICAS
        # ============================================
        metricas = {
            'total_registros': len(df),
            'colunas': df.columns.tolist(),
            'ultima_atualizacao': datetime.now().isoformat()
        }
        
        # ============================================
        # ANALISAR COLUNAS DA PLANILHA
        # ============================================
        colunas_lower = [str(c).lower().strip() for c in df.columns]
        
        # 1. PROcurar coluna de DATA
        nomes_data = ['data', 'data cadastro', 'data_cadastro', 'dt_cadastro', 'criado em', 'criação']
        coluna_data = None
        
        for i, col_lower in enumerate(colunas_lower):
            if any(nome in col_lower for nome in nomes_data):
                coluna_data = df.columns[i]
                break
        
        if coluna_data:
            logger.info(f"📅 Coluna de data encontrada: {coluna_data}")
            
            # Converte para datetime
            df['_data_convertida'] = pd.to_datetime(df[coluna_data], errors='coerce', dayfirst=True)
            df_validas = df[df['_data_convertida'].notna()]
            
            hoje = datetime.now().date()
            ontem = hoje - timedelta(days=1)
            sete_dias_atras = hoje - timedelta(days=7)
            trinta_dias_atras = hoje - timedelta(days=30)
            
            datas_validas = df_validas['_data_convertida'].dt.date
            
            metricas.update({
                'cadastros_hoje': int(len(datas_validas[datas_validas == hoje])),
                'cadastros_ontem': int(len(datas_validas[datas_validas == ontem])),
                'cadastros_7dias': int(len(datas_validas[datas_validas >= sete_dias_atras])),
                'cadastros_30dias': int(len(datas_validas[datas_validas >= trinta_dias_atras])),
                'tem_coluna_data': True,
                'nome_coluna_data': coluna_data
            })
        else:
            metricas.update({
                'cadastros_hoje': 0,
                'cadastros_ontem': 0,
                'cadastros_7dias': 0,
                'cadastros_30dias': 0,
                'tem_coluna_data': False
            })
        
        # 2. Procurar coluna de STATUS
        nomes_status = ['status', 'situação', 'situacao', 'etapa', 'fase', 'andamento']
        coluna_status = None
        
        for i, col_lower in enumerate(colunas_lower):
            if any(nome in col_lower for nome in nomes_status):
                coluna_status = df.columns[i]
                break
        
        if coluna_status:
            logger.info(f"📊 Coluna de status encontrada: {coluna_status}")
            
            # Conta valores únicos
            status_counts = df[coluna_status].value_counts().to_dict()
            
            # Tenta identificar pendentes vs concluídos
            pendentes = 0
            concluidos = 0
            outros = 0
            
            for status, qtd in status_counts.items():
                status_str = str(status).upper()
                if any(p in status_str for p in ['PENDENTE', 'AGUARDANDO', 'EM ANDAMENTO', 'PROCESSANDO', 'ANDAMENTO']):
                    pendentes += qtd
                elif any(c in status_str for c in ['CONCLUÍDO', 'CONCLUIDO', 'FINALIZADO', 'APROVADO', 'OK', 'PRONTO']):
                    concluidos += qtd
                else:
                    outros += qtd
            
            metricas.update({
                'status_distribuicao': status_counts,
                'pendentes': pendentes,
                'concluidos': concluidos,
                'outros_status': outros,
                'tem_coluna_status': True,
                'nome_coluna_status': coluna_status
            })
        else:
            metricas.update({
                'status_distribuicao': {},
                'pendentes': 0,
                'concluidos': 0,
                'outros_status': 0,
                'tem_coluna_status': False
            })
        
        # 3. Procurar coluna de RESPONSÁVEL
        nomes_resp = ['responsavel', 'responsável', 'usuario', 'usuário', 'criado por', 'user']
        coluna_resp = None
        
        for i, col_lower in enumerate(colunas_lower):
            if any(nome in col_lower for nome in nomes_resp):
                coluna_resp = df.columns[i]
                break
        
        if coluna_resp:
            resp_counts = df[coluna_resp].value_counts().head(5).to_dict()
            metricas.update({
                'top_responsaveis': resp_counts,
                'tem_coluna_responsavel': True
            })
        
        # 4. ÚLTIMOS 10 REGISTROS para tabela
        ultimos_registros = []
        for _, row in df.head(10).iterrows():
            registro = {}
            for col in df.columns[:6]:  # Primeiras 6 colunas
                registro[col] = str(row[col])[:100]
            ultimos_registros.append(registro)
        
        metricas['ultimos_registros'] = ultimos_registros
        
        # 5. Estatísticas por período (se tiver data)
        if coluna_data and len(df_validas) > 0:
            # Cadastros por dia (últimos 7 dias)
            cadastros_por_dia = []
            for i in range(6, -1, -1):
                data = (datetime.now() - timedelta(days=i)).date()
                qtd = len(datas_validas[datas_validas == data])
                cadastros_por_dia.append({
                    'data': data.strftime('%d/%m'),
                    'quantidade': qtd
                })
            metricas['cadastros_por_dia'] = cadastros_por_dia
        
        logger.info(f"✅ Métricas calculadas: Total={metricas['total_registros']}, Hoje={metricas.get('cadastros_hoje', 0)}")
        
        return jsonify({
            'sucesso': True,
            'dados': metricas
        })
        
    except Exception as e:
        logger.error(f"❌ Erro: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@app.route('/api/dashboard/sequencia-cadastros-corrigido')
def api_dashboard_sequencia_cadastros_corrigido():
    """
    Versão corrigida que lê a planilha com cabeçalho na linha 2
    """
    try:
        import gspread
        from google.oauth2 import service_account
        from pathlib import Path
        import pandas as pd
        from datetime import datetime
        
        # Configuração
        SHEET_ID = '1JMcU1mhbW0Q2IyLo15wHAoFLEY3IdFLEG0bMssP6y6o'
        ABA_NOME = 'Sequência Cad.'
        
        # Caminho do credentials.json
        current_dir = Path(__file__).parent
        credentials_path = current_dir / "credentials.json"
        
        if not credentials_path.exists():
            return jsonify({
                'sucesso': False,
                'erro': 'Arquivo credentials.json não encontrado'
            })
        
        # Autentica
        credentials = service_account.Credentials.from_service_account_file(
            str(credentials_path),
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        
        gc = gspread.authorize(credentials)
        planilha = gc.open_by_key(SHEET_ID)
        worksheet = planilha.worksheet(ABA_NOME)
        
        # Pega TODOS os valores
        todos_valores = worksheet.get_all_values()
        
        if not todos_valores or len(todos_valores) < 3:  # Precisa ter título + cabeçalho + dados
            return jsonify({
                'sucesso': False,
                'erro': 'Planilha sem dados suficientes'
            })
        
        # 🔥 LINHA 1: Título (ignorar)
        titulo = todos_valores[0]
        print(f"📌 Título: {titulo}")
        
        # 🔥 LINHA 2: Cabeçalhos (nomes das colunas)
        cabecalhos = todos_valores[1]
        print(f"📋 Cabeçalhos encontrados: {cabecalhos}")
        
        # LINHAS 3 em diante: Dados
        dados = todos_valores[2:]  # Pula as duas primeiras linhas
        
        # Filtra linhas vazias
        dados_filtrados = []
        for linha in dados:
            # Verifica se a linha tem algum conteúdo relevante
            if any(celula.strip() for celula in linha):
                dados_filtrados.append(linha)
        
        print(f"📊 Total de linhas com dados: {len(dados_filtrados)}")
        
        # Converte para lista de dicionários
        registros = []
        for linha in dados_filtrados:
            registro = {}
            for i, coluna in enumerate(cabecalhos):
                if coluna and coluna.strip():  # Só se a coluna tiver nome
                    if i < len(linha):
                        valor = linha[i].strip()
                        if valor:  # Só adiciona se tiver valor
                            registro[coluna] = valor
            if registro:  # Só adiciona se tiver pelo menos um campo
                registros.append(registro)
        
        # Converte para DataFrame
        df = pd.DataFrame(registros)
        
        print(f"✅ DataFrame criado com {len(df)} linhas")
        print(f"📋 Colunas no DataFrame: {df.columns.tolist()}")
        
        # Se não tiver dados, retorna erro
        if df.empty:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum dado encontrado após o cabeçalho'
            })
        
        # ============================================
        # ANÁLISE DOS DADOS
        # ============================================
        
        # 1. Distribuição por SITUAÇÃO
        situacoes = {}
        if 'SITUAÇÃO' in df.columns:
            situacoes_raw = df['SITUAÇÃO'].dropna()
            situacoes_raw = situacoes_raw[situacoes_raw != '']
            situacoes = situacoes_raw.value_counts().to_dict()
            print(f"📊 Situações: {situacoes}")
        
        # 2. Distribuição por RESPONSÁVEL
        responsaveis = {}
        if 'RESPONSÁVEL' in df.columns:
            responsaveis_raw = df['RESPONSÁVEL'].dropna()
            responsaveis_raw = responsaveis_raw[responsaveis_raw != '']
            responsaveis = responsaveis_raw.value_counts().to_dict()
            print(f"👥 Responsáveis: {responsaveis}")
        
        # 3. Prazos
        prazos = {}
        if 'PRAZO FORNECEDOR' in df.columns:
            com_prazo = df['PRAZO FORNECEDOR'].notna().sum()
            prazos = {
                'com_prazo': int(com_prazo),
                'sem_prazo': int(len(df) - com_prazo)
            }
        
        # 4. Marcas por situação
        marcas_por_situacao = {}
        if 'SITUAÇÃO' in df.columns and 'MARCA' in df.columns:
            for situacao in df['SITUAÇÃO'].dropna().unique():
                if situacao and situacao != '':
                    marcas = df[df['SITUAÇÃO'] == situacao]['MARCA'].dropna().tolist()
                    marcas_limpas = [str(m) for m in marcas if m and m != '']
                    if marcas_limpas:
                        marcas_por_situacao[str(situacao)] = marcas_limpas
        
        # 5. Últimos registros
        ultimos_registros = []
        for _, row in df.head(10).iterrows():
            registro = {
                'marca': str(row.get('MARCA', 'N/A')) if pd.notna(row.get('MARCA')) else 'N/A',
                'situacao': str(row.get('SITUAÇÃO', 'N/A')) if pd.notna(row.get('SITUAÇÃO')) else 'N/A',
                'responsavel': str(row.get('RESPONSÁVEL', 'N/A')) if pd.notna(row.get('RESPONSÁVEL')) else 'N/A',
                'observacao': str(row.get('OBSERVAÇÃO', 'N/A')) if pd.notna(row.get('OBSERVAÇÃO')) else 'N/A',
                'prazo': str(row.get('PRAZO FORNECEDOR', 'N/A')) if pd.notna(row.get('PRAZO FORNECEDOR')) else 'N-A'
            }
            ultimos_registros.append(registro)
        
        resultado = {
            'total_cadastros': len(df),
            'situacoes': situacoes,
            'responsaveis': responsaveis,
            'prazos': prazos,
            'marcas_por_situacao': marcas_por_situacao,
            'ultimos_cadastros': ultimos_registros,
            'colunas': df.columns.tolist(),
            'cabecalhos': [c for c in cabecalhos if c],  # Remove colunas vazias
            'ultima_atualizacao': datetime.now().isoformat()
        }
        
        print(f"✅ Dashboard gerado com sucesso!")
        print(f"   Total: {resultado['total_cadastros']}")
        print(f"   Situações: {len(situacoes)}")
        print(f"   Responsáveis: {len(responsaveis)}")
        
        return jsonify({
            'sucesso': True,
            'dados': resultado
        })
        
    except Exception as e:
        print(f"❌ Erro: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500
    

@app.route('/api/dashboard/sequencia-cadastros')
def api_dashboard_sequencia_cadastros():
    """
    API que retorna dados da sequência de cadastros
    Usando google_sheets_utils (que já existe e funciona)
    """
    try:
        # Import do utils (que já funciona)
        from google_sheets_utils import obter_dados_aba
        import pandas as pd
        from datetime import datetime
        
        print("✅ Import google_sheets_utils OK")
        
        # CONFIGURAÇÃO
        SHEET_ID = '1JMcU1mhbW0Q2IyLo15wHAoFLEY3IdFLEG0bMssP6y6o'
        ABA_NOME = 'Sequência Cad.'
        
        print(f"📊 Lendo planilha: {ABA_NOME}")
        
        # Busca dados usando o utils (sem limite)
        dados_aba = obter_dados_aba(SHEET_ID, ABA_NOME, limite_linhas=1000)
        
        if not dados_aba or not dados_aba.get('dados'):
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum dado encontrado na planilha'
            })
        
        # Converte para DataFrame
        df = pd.DataFrame(dados_aba['dados'])
        
        print(f"✅ {len(df)} linhas encontradas")
        print(f"📋 Colunas: {df.columns.tolist()}")
        
        # Remove linhas totalmente vazias
        df = df.dropna(how='all')
        
        # ============================================
        # MAPEAMENTO DAS COLUNAS
        # ============================================
        print("📋 Colunas disponíveis:")
        for col in df.columns:
            print(f"   - '{col}'")
        
        # ============================================
        # 1. DISTRIBUIÇÃO POR SITUAÇÃO
        # ============================================
        situacoes = {}
        if 'SITUAÇÃO' in df.columns:
            situacoes_raw = df['SITUAÇÃO'].dropna()
            situacoes_raw = situacoes_raw[situacoes_raw != '']
            situacoes = situacoes_raw.value_counts().to_dict()
            print(f"📊 Situações encontradas: {situacoes}")
        
        # ============================================
        # 2. DISTRIBUIÇÃO POR RESPONSÁVEL
        # ============================================
        responsaveis = {}
        if 'RESPONSÁVEL' in df.columns:
            responsaveis_raw = df['RESPONSÁVEL'].dropna()
            responsaveis_raw = responsaveis_raw[responsaveis_raw != '']
            responsaveis = responsaveis_raw.value_counts().to_dict()
            print(f"👥 Responsáveis: {responsaveis}")
        
        # ============================================
        # 3. PRAZOS
        # ============================================
        prazos = {}
        if 'PRAZO FORNECEDOR' in df.columns:
            com_prazo = df['PRAZO FORNECEDOR'].notna().sum()
            prazos = {
                'com_prazo': int(com_prazo),
                'sem_prazo': int(len(df) - com_prazo)
            }
        
        # ============================================
        # 4. MARCAS POR SITUAÇÃO
        # ============================================
        marcas_por_situacao = {}
        if 'SITUAÇÃO' in df.columns and 'MARCA' in df.columns:
            for situacao in df['SITUAÇÃO'].dropna().unique():
                if situacao and situacao != '':
                    marcas = df[df['SITUAÇÃO'] == situacao]['MARCA'].dropna().tolist()
                    marcas_limpas = [str(m) for m in marcas if m and m != '']
                    if marcas_limpas:
                        marcas_por_situacao[str(situacao)] = marcas_limpas
        
        # ============================================
        # 5. ÚLTIMOS CADASTROS (10 primeiros)
        # ============================================
        ultimos_cadastros = []
        for idx, row in df.head(10).iterrows():
            item = {
                'marca': str(row.get('MARCA', 'N/A')) if pd.notna(row.get('MARCA')) else 'N/A',
                'situacao': str(row.get('SITUAÇÃO', 'N/A')) if pd.notna(row.get('SITUAÇÃO')) else 'N/A',
                'responsavel': str(row.get('RESPONSÁVEL', 'N/A')) if pd.notna(row.get('RESPONSÁVEL')) else 'N/A',
                'observacao': str(row.get('OBSERVAÇÃO', 'N/A')) if pd.notna(row.get('OBSERVAÇÃO')) else 'N/A',
                'prazo': str(row.get('PRAZO FORNECEDOR', 'N/A')) if pd.notna(row.get('PRAZO FORNECEDOR')) else 'N/A'
            }
            ultimos_cadastros.append(item)
        
        # ============================================
        # RESULTADO
        # ============================================
        resultado = {
            'total_cadastros': len(df),
            'situacoes': situacoes,
            'responsaveis': responsaveis,
            'prazos': prazos,
            'marcas_por_situacao': marcas_por_situacao,
            'ultimos_cadastros': ultimos_cadastros,
            'colunas': df.columns.tolist(),
            'ultima_atualizacao': datetime.now().isoformat()
        }
        
        print(f"✅ Dashboard gerado com sucesso!")
        print(f"   Total: {resultado['total_cadastros']}")
        print(f"   Situações: {len(situacoes)}")
        print(f"   Responsáveis: {len(responsaveis)}")
        
        return jsonify({
            'sucesso': True,
            'dados': resultado
        })
        
    except Exception as e:
        print(f"❌ Erro detalhado: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@app.route('/api/diagnostico-import')
def api_diagnostico_import():
    """Rota para diagnosticar problemas de import"""
    import os
    import sys
    from pathlib import Path
    
    resultado = {
        'diretorio_atual': str(Path.cwd()),
        'arquivos_py': [],
        'python_path': sys.path,
        'erros': []
    }
    
    # Lista arquivos .py no diretório atual
    for arquivo in Path.cwd().glob('*.py'):
        resultado['arquivos_py'].append(arquivo.name)
    
    # Tenta importar de diferentes formas
    try:
        import google_sheets
        resultado['import_google_sheets'] = 'OK'
    except ImportError as e:
        resultado['import_google_sheets'] = str(e)
        resultado['erros'].append(str(e))
    
    try:
        from processamento import google_sheets
        resultado['import_processamento_google_sheets'] = 'OK'
    except ImportError as e:
        resultado['import_processamento_google_sheets'] = str(e)
    
    try:
        import google_sheets_utils
        resultado['import_google_sheets_utils'] = 'OK'
    except ImportError as e:
        resultado['import_google_sheets_utils'] = str(e)
    
    return jsonify(resultado)

def processar_dataframe_cadastros(df):
    """Função auxiliar para processar o DataFrame"""
    try:
        import pandas as pd
        
        # Remove linhas totalmente vazias
        df = df.dropna(how='all')
        
        # ============================================
        # MAPEAMENTO DAS COLUNAS
        # ============================================
        print("📋 Verificando colunas disponíveis:")
        for col in df.columns:
            print(f"   - '{col}'")
        
        # ============================================
        # 1. DISTRIBUIÇÃO POR SITUAÇÃO
        # ============================================
        situacoes = {}
        if 'SITUAÇÃO' in df.columns:
            situacoes_raw = df['SITUAÇÃO'].dropna()
            situacoes_raw = situacoes_raw[situacoes_raw != '']
            situacoes = situacoes_raw.value_counts().to_dict()
            print(f"📊 Situações: {situacoes}")
        
        # ============================================
        # 2. DISTRIBUIÇÃO POR RESPONSÁVEL
        # ============================================
        responsaveis = {}
        if 'RESPONSÁVEL' in df.columns:
            responsaveis_raw = df['RESPONSÁVEL'].dropna()
            responsaveis_raw = responsaveis_raw[responsaveis_raw != '']
            responsaveis = responsaveis_raw.value_counts().to_dict()
            print(f"👥 Responsáveis: {responsaveis}")
        
        # ============================================
        # 3. PRAZOS
        # ============================================
        prazos = {}
        if 'PRAZO FORNECEDOR' in df.columns:
            com_prazo = df['PRAZO FORNECEDOR'].notna().sum()
            prazos = {
                'com_prazo': int(com_prazo),
                'sem_prazo': int(len(df) - com_prazo)
            }
        
        # ============================================
        # 4. MARCAS POR SITUAÇÃO
        # ============================================
        marcas_por_situacao = {}
        if 'SITUAÇÃO' in df.columns and 'MARCA' in df.columns:
            for situacao in df['SITUAÇÃO'].dropna().unique():
                if situacao and situacao != '':
                    marcas = df[df['SITUAÇÃO'] == situacao]['MARCA'].dropna().tolist()
                    marcas_limpas = [str(m) for m in marcas if m and m != '']
                    if marcas_limpas:
                        marcas_por_situacao[str(situacao)] = marcas_limpas
        
        # ============================================
        # 5. ÚLTIMOS CADASTROS
        # ============================================
        ultimos_cadastros = []
        for idx, row in df.head(10).iterrows():
            item = {
                'marca': str(row.get('MARCA', 'N/A')) if pd.notna(row.get('MARCA')) else 'N/A',
                'situacao': str(row.get('SITUAÇÃO', 'N/A')) if pd.notna(row.get('SITUAÇÃO')) else 'N/A',
                'responsavel': str(row.get('RESPONSÁVEL', 'N/A')) if pd.notna(row.get('RESPONSÁVEL')) else 'N/A',
                'observacao': str(row.get('OBSERVAÇÃO', 'N/A')) if pd.notna(row.get('OBSERVAÇÃO')) else 'N/A',
                'prazo': str(row.get('PRAZO FORNECEDOR', 'N/A')) if pd.notna(row.get('PRAZO FORNECEDOR')) else 'N/A'
            }
            ultimos_cadastros.append(item)
        
        # ============================================
        # RESULTADO
        # ============================================
        resultado = {
            'total_cadastros': len(df),
            'situacoes': situacoes,
            'responsaveis': responsaveis,
            'prazos': prazos,
            'marcas_por_situacao': marcas_por_situacao,
            'ultimos_cadastros': ultimos_cadastros,
            'colunas': df.columns.tolist()
        }
        
        print(f"✅ Dashboard gerado com sucesso!")
        print(f"   Total: {resultado['total_cadastros']}")
        print(f"   Situações: {len(situacoes)}")
        print(f"   Responsáveis: {len(responsaveis)}")
        
        return jsonify({
            'sucesso': True,
            'dados': resultado
        })
        
    except Exception as e:
        print(f"❌ Erro no processamento: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def processar_dataframe(df):
    """Função auxiliar para processar o DataFrame"""
    try:
        if df.empty:
            return jsonify({
                'sucesso': False,
                'erro': 'Planilha vazia'
            })
        
        print(f"✅ {len(df)} linhas encontradas")
        print(f"📋 Colunas: {df.columns.tolist()}")
        
        # Limpa dados
        df = df.dropna(how='all')
        
        # ============================================
        # MÉTRICAS
        # ============================================
        total_cadastros = len(df)
        
        # Distribuição por SITUAÇÃO
        situacoes = {}
        if 'SITUAÇÃO' in df.columns:
            situacoes = df['SITUAÇÃO'].value_counts().to_dict()
        
        # Distribuição por RESPONSÁVEL
        responsaveis = {}
        if 'RESPONSÁVEL' in df.columns:
            responsaveis = df['RESPONSÁVEL'].value_counts().to_dict()
        
        # Prazos
        prazos = {}
        if 'PRAZO FORNECEDOR' in df.columns:
            prazos_preenchidos = df['PRAZO FORNECEDOR'].notna().sum()
            prazos = {
                'com_prazo': int(prazos_preenchidos),
                'sem_prazo': int(len(df) - prazos_preenchidos)
            }
        
        # Marcas por situação
        marcas_por_situacao = {}
        if 'SITUAÇÃO' in df.columns and 'MARCA' in df.columns:
            for situacao in df['SITUAÇÃO'].unique():
                if pd.notna(situacao):
                    marcas = df[df['SITUAÇÃO'] == situacao]['MARCA'].tolist()
                    marcas_por_situacao[str(situacao)] = [m for m in marcas if pd.notna(m)]
        
        # Últimos cadastros
        ultimos_cadastros = []
        for idx, row in df.head(10).iterrows():
            item = {
                'marca': row.get('MARCA', 'N/A'),
                'situacao': row.get('SITUAÇÃO', 'N/A'),
                'responsavel': row.get('RESPONSÁVEL', 'N/A'),
                'observacao': row.get('OBSERVAÇÃO', 'N/A'),
                'prazo': row.get('PRAZO FORNECEDOR', 'N/A')
            }
            ultimos_cadastros.append(item)
        
        # Resultado
        resultado = {
            'total_cadastros': total_cadastros,
            'situacoes': situacoes,
            'responsaveis': responsaveis,
            'prazos': prazos,
            'marcas_por_situacao': marcas_por_situacao,
            'ultimos_cadastros': ultimos_cadastros,
            'colunas': df.columns.tolist()
        }
        
        print(f"✅ Dashboard gerado: {total_cadastros} cadastros")
        
        return jsonify({
            'sucesso': True,
            'dados': resultado
        })
        
    except Exception as e:
        print(f"❌ Erro no processamento: {str(e)}")
        raise
    
    
def obter_dados_aba(sheet_id, aba_nome, limite_linhas=None):
    """Obtém TODOS os dados de uma aba específica para preview"""
    try:
        # Encontra o caminho correto para o credentials.json
        current_dir = Path(__file__).parent
        credentials_path = current_dir / "credentials.json"
        
        if not credentials_path.exists():
            raise FileNotFoundError("Arquivo credentials.json não encontrado")
        
        credentials = service_account.Credentials.from_service_account_file(
            str(credentials_path),
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        
        gc = gspread.authorize(credentials)
        planilha = gc.open_by_key(sheet_id)
        worksheet = planilha.worksheet(aba_nome)
        
        # Obtém TODAS as linhas (sem limite)
        todas_linhas = worksheet.get_all_values()
        
        if not todas_linhas:
            return {
                'colunas': [],
                'dados': [],
                'total_linhas': 0,
                'total_colunas': 0
            }
        
        # A primeira linha são os cabeçalhos
        colunas = todas_linhas[0] if todas_linhas else []
        
        # Converte TODAS as linhas seguintes para dicionários
        dados = []
        for i, linha in enumerate(todas_linhas[1:], 1):
            linha_dict = {}
            for j, valor in enumerate(linha):
                nome_coluna = colunas[j] if j < len(colunas) else f"Coluna_{j+1}"
                linha_dict[nome_coluna] = valor
            dados.append(linha_dict)
        
        return {
            'colunas': colunas,
            'dados': dados,  # TODAS as linhas
            'total_linhas': worksheet.row_count,
            'total_colunas': worksheet.col_count
        }
        
    except Exception as e:
        if current_app:
            current_app.logger.error(f"Erro ao obter dados da aba: {str(e)}")
        else:
            print(f"Erro ao obter dados da aba: {str(e)}")
        raise Exception(f"Erro ao obter dados da aba: {str(e)}")

@app.route("/api/abas-google-sheets")
def api_abas_google_sheets():
    """API para listar abas de uma planilha - AGORA APENAS VISÍVEIS"""
    sheet_id = request.args.get('sheet_id')
    if not sheet_id:
        return jsonify({'error': 'sheet_id é obrigatório'}), 400
    
    try:
        # ALTERAÇÃO: Usa a nova função para abas visíveis
        from google_sheets_utils import listar_abas_visiveis_google_sheets
        abas = listar_abas_visiveis_google_sheets(sheet_id)
        return jsonify({'success': True, 'abas': abas})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
@app.route("/api/abas-google-sheets-visiveis")
def api_abas_google_sheets_visiveis():
    """API para listar apenas abas visíveis de uma planilha"""
    sheet_id = request.args.get('sheet_id')
    if not sheet_id:
        return jsonify({'error': 'sheet_id é obrigatório'}), 400
    
    try:
        from google_sheets_utils import listar_abas_visiveis_google_sheets
        abas = listar_abas_visiveis_google_sheets(sheet_id)
        return jsonify({'success': True, 'abas': abas})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route("/api/preview-aba")
def api_preview_aba():
    """API para preview de uma aba"""
    sheet_id = request.args.get('sheet_id')
    aba_nome = request.args.get('aba_nome')
    
    if not sheet_id or not aba_nome:
        return jsonify({'error': 'sheet_id e aba_nome são obrigatórios'}), 400
    
    try:
        from google_sheets_utils import obter_dados_aba
        preview_data = obter_dados_aba(sheet_id, aba_nome)
        return jsonify({'success': True, 'data': preview_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
@app.route("/comparar-prazos", methods=["POST"])
def comparar_prazos():
    try:
        if ('arquivo_erp' not in request.files) or ('arquivo_marketplace' not in request.files):
            registrar_processo(
                modulo="prazos",
                qtd_itens=0,
                tempo_execucao=0,
                status="erro",
                erro_mensagem="Nenhum arquivo enviado"
            )
            return jsonify({'sucesso': False, 'erro': "Nenhum arquivo enviado"}), 400

        arquivo_erp = request.files['arquivo_erp']
        arquivo_marketplace = request.files['arquivo_marketplace']

        if arquivo_erp.filename == '' or arquivo_marketplace.filename == '':
            registrar_processo(
                modulo="prazos",
                qtd_itens=0,
                tempo_execucao=0,
                status="erro",
                erro_mensagem="Nenhum arquivo selecionado"
            )
            return jsonify({'sucesso': False, 'erro': "Nenhum arquivo selecionado"}), 400

        inicio = datetime.now()
        resultado = processar_comparacao(arquivo_erp, arquivo_marketplace, app.config['UPLOAD_FOLDER'])
        
        if not resultado.get('sucesso', False):
            registrar_processo(
                modulo="prazos",
                qtd_itens=0,
                tempo_execucao=0,
                status="erro",
                erro_mensagem=resultado.get('erro', 'Erro desconhecido')
            )
            raise Exception(resultado.get('erro', 'Erro desconhecido no processamento'))

        # Registrar processo com sucesso
        tempo_segundos = (datetime.now() - inicio).total_seconds()
        registrar_processo(
            modulo="prazos",
            qtd_itens=resultado['total_itens'],
            tempo_execucao=tempo_segundos,
            status="sucesso"
        )

        return jsonify(resultado)

    except Exception as e:
        registrar_processo(
            modulo="prazos",
            qtd_itens=0,
            tempo_execucao=0,
            status="erro",
            erro_mensagem=str(e)
        )
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route("/comparar-prazos", methods=["GET"])
def mostrar_tela_comparacao():
    return render_template(
        "comparar_prazos.html",
        active_page='comparar-prazos',
        active_module='cadastro',
        historico_processos=obter_historico_processos("prazos"),
        processos_hoje=contar_processos_hoje("prazos"),
        stats=get_processing_stats("prazos"),
        page_title='Comparação de Prazos'  # Agora aceita o parâmetro
    )

def contar_processos_hoje(modulo="cadastro"):
    """Função local necessária para o template (não remova!)."""
    from log_utils import contar_processos_hoje as contar_logs  # Importa a função original
    return contar_logs(modulo)  # Delega para a função de log_utils.py

@app.context_processor
def inject_stats():
    return {
        'count_processos_hoje': contar_processos_hoje(),
        'now': datetime.now()
    }

def format_time(seconds):
    """Formata segundos em minutos e segundos"""
    minutes = int(seconds // 60)
    seconds = int(seconds % 60)
    return f"{minutes}m {seconds}s"

@app.errorhandler(500)
def handle_500_error(e):
    return jsonify({
        'status': 'error',
        'message': 'Internal server error',
        'details': str(e) if app.debug else None
    }), 500

@app.route("/configuracoes/google-sheets", methods=["GET", "POST"])
def configurar_google_sheets():
    """Tela de configuração do Google Sheets - Agora mostra apenas abas visíveis"""
    config = carregar_configuracao_google_sheets()
    abas = []
    mensagem = None
    tipo_mensagem = "info"
    erro = None
    
    try:
        if request.method == "POST":
            sheet_id = request.form.get('sheet_id', '').strip()
            acao = request.form.get('acao')
            
            if acao == 'testar':
                sucesso, msg = testar_conexao_google_sheets(sheet_id)
                mensagem = msg
                tipo_mensagem = "success" if sucesso else "danger"
                
            elif acao == 'listar_abas':
                if sheet_id:
                    # ALTERAÇÃO: Usa a nova função para abas visíveis
                    from google_sheets_utils import listar_abas_visiveis_google_sheets
                    abas = listar_abas_visiveis_google_sheets(sheet_id)
                    mensagem = f"{len(abas)} abas visíveis encontradas"
                    tipo_mensagem = "success"
                else:
                    mensagem = "Informe o ID da planilha primeiro"
                    tipo_mensagem = "warning"
                
            elif acao == 'salvar':
                if not sheet_id:
                    mensagem = "ID da planilha é obrigatório"
                    tipo_mensagem = "danger"
                else:
                    # Salva apenas o ID, a aba será selecionada na tela de extração
                    if salvar_configuracao_google_sheets(sheet_id, ''):
                        config = carregar_configuracao_google_sheets()
                        mensagem = "ID da planilha salvo com sucesso! Selecione a aba na tela de extração."
                        tipo_mensagem = "success"
                    else:
                        mensagem = "Erro ao salvar configuração"
                        tipo_mensagem = "danger"
    
    except Exception as e:
        erro = str(e)
        mensagem = f"Erro: {erro}"
        tipo_mensagem = "danger"
    
    return render_template(
        "config_google_sheets.html",
        active_page='configuracao',
        config=config,
        abas=abas,
        mensagem=mensagem,
        tipo_mensagem=tipo_mensagem,
        erro=erro,
        page_title='Extração de Atributos'
    )

@app.route("/validar-xml", methods=["GET", "POST"])
def validar_xml():
    resultado = None
    if request.method == "POST":
        if "arquivo_xml" not in request.files:
            return jsonify({"sucesso": False, "erro": "Nenhum arquivo enviado"}), 400

        arquivo = request.files["arquivo_xml"]
        if arquivo.filename == "":
            return jsonify({"sucesso": False, "erro": "Nenhum arquivo selecionado"}), 400

        resultado = validar_xml_nfe(arquivo) # type: ignore

        return jsonify(resultado)

    return render_template(
        "validar_xml.html",
        historico_processos=obter_historico_processos("xml"),
        processos_hoje=contar_processos_hoje("xml"),
        stats=get_processing_stats("xml")
        
    )

@app.route("/salvar-ordem-fotos", methods=["POST"])
def salvar_ordem_fotos():
    """API para salvar a ordem das fotos - IMPLEMENTAÇÃO INICIAL"""
    try:
        data = request.get_json()
        fotos = data.get('fotos', [])
        
        print(f"📝 Recebida solicitação para salvar ordem de {len(fotos)} fotos")
        
        # ✅ CORREÇÃO: Log para debug
        for i, foto in enumerate(fotos):
            print(f"Foto {i+1}: Produto {foto.get('product_id')}, Foto {foto.get('photo_id')}, Index: {foto.get('new_index')}")
        
        # ⚠️ ATENÇÃO: Esta é uma implementação básica
        # Você precisará implementar a lógica real de atualização na API AnyMarket
        # A API AnyMarket pode não suportar reordenação via REST
        
        return jsonify({
            'sucesso': True,
            'message': f'Ordem de {len(fotos)} fotos processada (implementação em desenvolvimento)',
            'total_fotos': len(fotos)
        })
        
    except Exception as e:
        print(f"❌ Erro ao salvar ordem: {str(e)}")
        return jsonify({
            'sucesso': False,
            'erro': f'Erro ao salvar ordem: {str(e)}'
        }), 500

@app.route('/consultar-produto')
def consultar_produto():
    """Página para consultar produtos por SKU"""
    return render_template(
        'consultar_produto.html',
        active_page='consultar_produto',
        active_module='anymarket',
        page_title='Produtos'
    )

@app.route('/api/anymarket/exportar-excel')
def api_exportar_excel():
    """API para exportar pedidos para Excel - COM TODOS OS FILTROS"""
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            return jsonify({'success': False, 'error': 'Token de autenticação não fornecido'}), 401
        
        token = auth_header.replace('Bearer ', '')
        
        # Coletar todos os filtros
        data_inicio = request.args.get('dataInicio')
        data_fim = request.args.get('dataFim')
        status = request.args.get('status')
        marketplace = request.args.get('marketplace')
        numero_pedido = request.args.get('numeroPedido')  # ✅ NOVO FILTRO
        
        # Buscar TODOS os pedidos com os filtros
        all_orders = []
        page = 1
        limit = 100  # Máximo por página
        
        while True:
            # Construir parâmetros igual à API normal
            params = {
                'offset': (page - 1) * limit,
                'limit': limit,
            }
            
            if status and status.strip():
                params['status'] = status.strip()
                
            if marketplace and marketplace.strip():
                params['marketplace'] = marketplace.strip()
            
            if data_inicio and data_inicio.strip():
                try:
                    datetime.strptime(data_inicio, '%Y-%m-%d')
                    params['createdAfter'] = f"{data_inicio}T00:00:00-03:00"
                except ValueError:
                    pass
            
            if data_fim and data_fim.strip():
                try:
                    datetime.strptime(data_fim, '%Y-%m-%d')
                    params['createdBefore'] = f"{data_fim}T23:59:59-03:00"
                except ValueError:
                    pass
            
            # ✅ FILTRO POR NÚMERO DO PEDIDO
            if numero_pedido and numero_pedido.strip():
                params['marketPlaceNumber'] = numero_pedido.strip()
            
            # Fazer requisição
            headers = {
                'Authorization': f'Bearer {token}',
                'Content-Type': 'application/json',
                'gumgaToken': token
            }
            
            response = requests.get(
                "https://api.anymarket.com.br/v2/orders", 
                params=params, 
                headers=headers, 
                timeout=60
            )
            
            if response.status_code != 200:
                break
                
            data = response.json()
            orders = data.get('content', [])
            
            if not orders:
                break
                
            all_orders.extend(orders)
            
            # Verificar se há mais páginas
            pagination = data.get('page', {})
            if len(orders) < limit:
                break
                
            page += 1
        
        # Criar DataFrame para exportação
        df_data = []
        for order in all_orders:
            df_data.append({
                'ID': order.get('id'),
                'Marketplace': order.get('marketPlace'),
                'Status': order.get('status'),
                'Nº Marketplace': order.get('marketPlaceNumber'),
                'Comprador': order.get('buyer', {}).get('name'),
                'Email': order.get('buyer', {}).get('email'),
                'Telefone': order.get('buyer', {}).get('phone'),
                'Data Criação': order.get('createdAt'),
                'Data Pagamento': order.get('paymentDate'),
                'Quantidade Itens': len(order.get('items', [])),
                'Valor Total': order.get('total'),
                'Frete': order.get('freight'),
                'Desconto': order.get('discount'),
                'Cidade': order.get('buyer', {}).get('city'),
                'Estado': order.get('buyer', {}).get('state')
            })
        
        df = pd.DataFrame(df_data)
        
        # Criar arquivo Excel em memória
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Pedidos', index=False)
            
            # Formatar a planilha
            worksheet = writer.sheets['Pedidos']
            
            # Ajustar largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Retornar arquivo
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'pedidos_anymarket_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )
        
    except Exception as e:
        print(f"❌ Erro na exportação Excel: {str(e)}")
        return jsonify({'success': False, 'error': f'Erro na exportação: {str(e)}'}), 500
    
@app.route('/api/anymarket/produtos/buscar-sku', methods=['POST'])
def api_buscar_produto_sku():
    """API para buscar produto por SKU - ROTA NOVA"""
    try:
        # ✅ USA APENAS A FUNÇÃO NOVA, SEM MEXER NO EXISTENTE
        from processamento.api_anymarket import buscar_produto_por_sku
        
        data = request.get_json()
        sku = data.get('sku', '').strip()
        
        if not sku:
            return jsonify({'sucesso': False, 'erro': 'SKU é obrigatório'}), 400
        
        # Chama a função nova
        resultado = buscar_produto_por_sku(sku)
        
        return jsonify(resultado)
        
    except Exception as e:
        print(f"❌ Erro ao buscar produto por SKU: {str(e)}")
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/anymarket/produtos/<product_id>')
def api_buscar_produto_id(product_id):
    """API para buscar produto por ID"""
    try:
        from processamento.api_anymarket import obter_cliente_anymarket
        
        client = obter_cliente_anymarket()
        resultado = client.buscar_produto_por_id(product_id)
        
        return jsonify(resultado)
        
    except Exception as e:
        print(f"❌ Erro ao buscar produto por ID: {str(e)}")
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/anymarket/produtos')
def api_listar_produtos():
    """API para listar produtos com paginação"""
    try:
        from processamento.api_anymarket import obter_cliente_anymarket
        
        pagina = request.args.get('page', 1, type=int)
        limite = request.args.get('limit', 50, type=int)
        filtro = request.args.get('filtro', '')
        
        client = obter_cliente_anymarket()
        
        params = {
            'limit': limite,
            'offset': (pagina - 1) * limite
        }
        
        if filtro:
            params['title'] = filtro
        
        resultado = client.buscar_produtos_com_filtros(params)
        
        return jsonify(resultado)
        
    except Exception as e:
        print(f"❌ Erro ao listar produtos: {str(e)}")
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
    
@app.route('/perfil-mercado-livre')
def perfil_mercado_livre():
    """Dashboard com dados do perfil do Mercado Livre"""
    esta_autenticado = ml_token_manager.is_authenticated()
    
    dados_perfil = None
    if esta_autenticado:
        dados_perfil = obter_dados_completos_perfil()
    
    return render_template(
        'perfil_mercadolivre.html',
        active_page='perfil_mercado_livre',
        active_module='mercadolivre',
        page_title='Dashboard - Mercado Livre',
        esta_autenticado=esta_autenticado,
        dados_perfil=dados_perfil
    )

@app.route('/api/mercadolivre/exportar-excel', methods=['POST'])
def api_exportar_mlb_excel():
    """API para exportar resultados de MLB para Excel - COM COLUNAS DE VENDAS"""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        data = request.get_json()
        dados_exportacao = data.get('dados', [])
        
        if not dados_exportacao:
            return jsonify({'sucesso': False, 'erro': 'Nenhum dado para exportar'}), 400
        
        # Criar DataFrame
        df = pd.DataFrame(dados_exportacao)
        
        # DEFINIR A ORDEM DAS COLUNAS COM VENDAS
        colunas_ordenadas = [
            'MLB Principal', 'MLB Variação', 'Tipo', 'SKU', 'Título', 
            'Preço', 'Estoque', 
            'Vendidos (API)',          # NOVA COLUNA
            'Vendidos (Real)',          # NOVA COLUNA
            'Vendidos (Variações)',     # NOVA COLUNA
            'Modo Envio', 'Prazo Fabricação', 'Status', 'Frete Grátis',
            'Qtd Variações', 'Catálogo', 'Tipo Anúncio', 'Data Criação', 'Link', 'Erro'
        ]
        
        # Manter apenas as colunas que existem nos dados
        colunas_existentes = []
        for col in colunas_ordenadas:
            if col in df.columns:
                colunas_existentes.append(col)
            else:
                # Se a coluna não existe, criar com valor padrão
                df[col] = '-'
                colunas_existentes.append(col)
        
        # Reordenar o DataFrame
        df = df[colunas_existentes]
        
        # Criar arquivo Excel em memória
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Planilha principal com todos os dados
            df.to_excel(writer, sheet_name='Resultados MLB', index=False)
            
            # Planilha resumida com estatísticas de vendas
            # Calcular totais de vendas
            total_vendas_api = 0
            total_vendas_real = 0
            
            # Tentar extrair dos dados recebidos
            for item in dados_exportacao:
                # Pega vendas (API)
                vendas_api = item.get('Vendidos (API)', 0)
                if isinstance(vendas_api, str) and vendas_api != '-':
                    try:
                        vendas_api = int(vendas_api)
                    except:
                        vendas_api = 0
                elif isinstance(vendas_api, (int, float)):
                    pass
                else:
                    vendas_api = 0
                
                # Pega vendas (Real)
                vendas_real = item.get('Vendidos (Real)', 0)
                if isinstance(vendas_real, str) and vendas_real != '-':
                    try:
                        vendas_real = int(vendas_real)
                    except:
                        vendas_real = 0
                elif isinstance(vendas_real, (int, float)):
                    pass
                else:
                    vendas_real = 0
                
                total_vendas_api += vendas_api
                total_vendas_real += vendas_real
            
            # Estatísticas
            estatisticas_data = {
                'Métrica': [
                    'Total de Itens',
                    'Total de Anúncios Principais',
                    'Total de Variações',
                    'Total Vendas (API Principal)',
                    'Total Vendas (Real - Soma Variações)',
                    'Diferença (API - Real)',
                    'Percentual de Diferença',
                    'Data da Exportação'
                ],
                'Valor': [
                    len(dados_exportacao),
                    data.get('total_principais', 0),
                    data.get('total_variações', 0),
                    f"{total_vendas_api:,.0f}",
                    f"{total_vendas_real:,.0f}",
                    f"{total_vendas_api - total_vendas_real:,.0f}",
                    f"{((total_vendas_api - total_vendas_real) / total_vendas_real * 100):.1f}%" if total_vendas_real > 0 else "0%",
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                ]
            }
            df_estatisticas = pd.DataFrame(estatisticas_data)
            df_estatisticas.to_excel(writer, sheet_name='Estatísticas', index=False)
            
            # Formatar planilha principal
            worksheet = writer.sheets['Resultados MLB']
            
            # Formatação para números
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    if cell.column_letter in ['F', 'G', 'H', 'I', 'J']:  # Colunas de números
                        if cell.value and cell.value != '-':
                            try:
                                # Formatar como número
                                cell.number_format = '#,##0'
                            except:
                                pass
            
            # Destacar variações
            light_blue_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            light_yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            
            # Encontrar a coluna 'Tipo'
            tipo_col_idx = None
            for idx, col in enumerate(df.columns, 1):
                if col == 'Tipo':
                    tipo_col_idx = idx
                    break
            
            if tipo_col_idx:
                for row_idx in range(2, len(dados_exportacao) + 2):
                    tipo_cell = worksheet.cell(row=row_idx, column=tipo_col_idx)
                    if tipo_cell.value and 'Variação' in str(tipo_cell.value):
                        # Cor azul clara para variações
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = light_blue_fill
                    elif tipo_cell.value == 'Principal':
                        # Cor amarela clara para principais
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = light_yellow_fill
            
            # Ajustar largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Formatar cabeçalho
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Adicionar filtros
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Formatar planilha de estatísticas
            worksheet_stats = writer.sheets['Estatísticas']
            for cell in worksheet_stats[1]:
                cell.font = Font(bold=True)
            
            for column in worksheet_stats.columns:
                column_letter = column[0].column_letter
                worksheet_stats.column_dimensions[column_letter].width = 35
        
        output.seek(0)
        
        # Nome do arquivo com timestamp
        nome_arquivo = f"consulta_mlb_vendas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )
        
    except Exception as e:
        print(f"❌ Erro na exportação Excel MLB: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'erro': f'Erro na exportação: {str(e)}'}), 500
    
@app.route('/api/mercadolivre/debug-vendas-site/<mlb>')
def api_debug_vendas_site(mlb):
    """
    Rota de DEBUG para COMPARAR o que a API retorna vs o que aparece no site
    Faz scraping básico da página do produto para ver os dados reais
    """
    try:
        from token_manager_secure import ml_token_manager
        import requests
        from bs4 import BeautifulSoup
        import re
        
        if not ml_token_manager.is_authenticated():
            return jsonify({
                'sucesso': False,
                'erro': 'Não autenticado no Mercado Livre'
            }), 401
        
        token = ml_token_manager.get_valid_token()
        headers_api = {'Authorization': f'Bearer {token}'}
        
        print(f"\n{'='*60}")
        print(f"🔍 DEBUG COMPARATIVO - MLB: {mlb}")
        print(f"{'='*60}")
        
        # =========================================
        # 1. DADOS DA API
        # =========================================
        print("\n📦 1. BUSCANDO DADOS DA API...")
        response_api = requests.get(
            f"https://api.mercadolibre.com/items/{mlb}",
            headers=headers_api,
            timeout=15
        )
        
        if response_api.status_code != 200:
            return jsonify({
                'sucesso': False,
                'erro': f'Erro na API: {response_api.status_code}'
            })
        
        dados_api = response_api.json()
        
        sold_quantity_api = dados_api.get('sold_quantity', 0)
        titulo = dados_api.get('title', '')
        
        print(f"   Título: {titulo[:100]}...")
        print(f"   sold_quantity (API): {sold_quantity_api}")
        
        # =========================================
        # 2. DADOS DO SITE (scraping básico)
        # =========================================
        print("\n🌐 2. BUSCANDO DADOS DO SITE...")
        
        # URL do produto no site
        url_site = f"https://produto.mercadolivre.com.br/{mlb}"
        
        headers_browser = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        response_site = requests.get(url_site, headers=headers_browser, timeout=15)
        
        dados_site = {
            'url': url_site,
            'status_code': response_site.status_code,
            'vendas_texto': None,
            'vendas_numero': None,
            'encontrado': False
        }
        
        if response_site.status_code == 200:
            soup = BeautifulSoup(response_site.text, 'html.parser')
            
            # Procurar por texto de vendas em diferentes lugares
            padroes = [
                r'(\d+)\s*vendidos?',
                r'(\d+[\.,]?\d*)\s*unidades?\s*vendidas?',
                r'Mais\s*de\s*(\d+)\s*vendidos?',
                r'(\d+)\s*compras?'
            ]
            
            # Procurar em todo o texto da página
            texto_pagina = soup.get_text()
            
            for padrao in padroes:
                matches = re.findall(padrao, texto_pagina, re.IGNORECASE)
                if matches:
                    # Pega o primeiro match
                    valor_texto = matches[0]
                    # Limpar e converter para número
                    if isinstance(valor_texto, tuple):
                        valor_texto = valor_texto[0]
                    valor_numero = int(re.sub(r'[^\d]', '', str(valor_texto)))
                    
                    dados_site['vendas_texto'] = valor_texto
                    dados_site['vendas_numero'] = valor_numero
                    dados_site['encontrado'] = True
                    dados_site['padrao_usado'] = padrao
                    break
            
            # Procurar em elementos específicos
            elementos_vendas = soup.find_all(['span', 'div', 'p'], 
                string=re.compile(r'vendidos?|unidades', re.IGNORECASE))
            
            dados_site['elementos_encontrados'] = len(elementos_vendas)
            
            print(f"   Status code: {response_site.status_code}")
            print(f"   Vendas no site: {dados_site['vendas_numero']} (texto: '{dados_site['vendas_texto']}')")
        else:
            print(f"   ❌ Erro ao acessar site: {response_site.status_code}")
        
        # =========================================
        # 3. VERIFICAR HISTÓRICO DE VENDAS (se disponível)
        # =========================================
        print("\n📊 3. VERIFICANDO HISTÓRICO...")
        
        # Tentar buscar vendas em outras fontes
        outras_fontes = {}
        
        # Verificar se tem variações que podem ter vendas
        if 'variations' in dados_api and dados_api['variations']:
            print(f"   ℹ️  Produto tem {len(dados_api['variations'])} variações")
            
            # Buscar detalhes das variações
            variacoes_detalhes = []
            soma_vendas_variacoes = 0
            
            for var in dados_api['variations']:
                var_id = var.get('id')
                if var_id:
                    response_var = requests.get(
                        f"https://api.mercadolibre.com/items/{mlb}/variations/{var_id}",
                        headers=headers_api,
                        timeout=10
                    )
                    if response_var.status_code == 200:
                        var_data = response_var.json()
                        vendas_var = var_data.get('sold_quantity', 0)
                        soma_vendas_variacoes += vendas_var
                        variacoes_detalhes.append({
                            'id': var_id,
                            'sold_quantity': vendas_var,
                            'attributes': var_data.get('attribute_combinations', [])
                        })
            
            outras_fontes['vendas_variacoes'] = soma_vendas_variacoes
            outras_fontes['variacoes_detalhadas'] = variacoes_detalhes
            print(f"   Vendas totais nas variações: {soma_vendas_variacoes}")
        
        # =========================================
        # 4. RESUMO E COMPARAÇÃO
        # =========================================
        print("\n" + "="*60)
        print("📋 RESUMO COMPARATIVO:")
        print("="*60)
        print(f"MLB: {mlb}")
        print(f"Título: {titulo[:80]}...")
        print(f"\n📈 API Mercado Livre: {sold_quantity_api} vendas")
        
        if dados_site['vendas_numero']:
            print(f"🌐 Site Mercado Livre: {dados_site['vendas_numero']} vendas")
            diferenca = sold_quantity_api - dados_site['vendas_numero']
            print(f"\n⚖️ DIFERENÇA: {diferenca} ({'+' if diferenca > 0 else ''}{diferenca})")
            
            if diferenca != 0:
                print(f"\n⚠️  ALERTA: API está retornando {abs(diferenca)} vendas a mais que o site!")
        else:
            print("🌐 Site: Não foi possível extrair vendas")
        
        if 'vendas_variacoes' in outras_fontes:
            print(f"\n📊 Soma das variações: {outras_fontes['vendas_variacoes']}")
        
        print("="*60)
        
        # =========================================
        # 5. RETORNAR TODOS OS DADOS
        # =========================================
        return jsonify({
            'sucesso': True,
            'mlb': mlb,
            'comparacao': {
                'api': sold_quantity_api,
                'site': dados_site['vendas_numero'],
                'diferenca': sold_quantity_api - dados_site['vendas_numero'] if dados_site['vendas_numero'] else None,
                'site_texto': dados_site['vendas_texto'],
                'url_site': url_site
            },
            'dados_api': {
                'id': dados_api.get('id'),
                'title': dados_api.get('title'),
                'sold_quantity': dados_api.get('sold_quantity'),
                'available_quantity': dados_api.get('available_quantity'),
                'status': dados_api.get('status'),
                'has_variations': 'variations' in dados_api and len(dados_api.get('variations', [])) > 0
            },
            'dados_site': dados_site,
            'outras_fontes': outras_fontes,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        print(f"❌ Erro no debug: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'erro': str(e),
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/mercadolivre/contas')
def api_listar_contas():
    """Lista todas as contas"""
    try:
        accounts = ml_token_manager.get_all_accounts()
        return jsonify({
            'sucesso': True,
            'contas': accounts,
            'conta_atual': ml_token_manager.current_account_id
        })
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/contas/adicionar', methods=['POST'])
def api_adicionar_conta_auto():
    """Adiciona nova conta e tenta obter tokens AUTOMATICAMENTE"""
    try:
        data = request.get_json()
        
        account_name = data.get('account_name')
        app_id = data.get('app_id')
        secret_key = data.get('secret_key')
        
        if not account_name or not app_id or not secret_key:
            return jsonify({'sucesso': False, 'erro': 'Nome, App ID e Secret Key são obrigatórios'}), 400
        
        print(f"📥 Tentando adicionar conta: {account_name}")
        
        # Chama o método que tenta obter tokens automaticamente
        account_id, sucesso, mensagem = ml_token_manager.add_account_with_app_credentials(
            account_name, app_id, secret_key
        )
        
        if account_id:
            return jsonify({
                'sucesso': sucesso,
                'mensagem': mensagem,
                'account_id': account_id,
                'autenticada_automaticamente': sucesso
            })
        else:
            return jsonify({'sucesso': False, 'erro': mensagem}), 500
            
    except Exception as e:
        print(f"❌ Erro na API: {str(e)}")
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/contas/<account_id>/adicionar-tokens-manual', methods=['POST'])
def api_adicionar_tokens_manual(account_id):
    """Fallback: Adiciona tokens manualmente"""
    try:
        data = request.get_json()
        access_token = data.get('access_token')
        refresh_token = data.get('refresh_token')
        
        if not access_token or not refresh_token:
            return jsonify({'sucesso': False, 'erro': 'Ambos os tokens são obrigatórios'}), 400
        
        sucesso, mensagem = ml_token_manager.add_tokens_manually(
            account_id, access_token, refresh_token
        )
        
        if sucesso:
            return jsonify({
                'sucesso': True,
                'mensagem': mensagem
            })
        else:
            return jsonify({'sucesso': False, 'erro': mensagem}), 400
            
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/contas/<account_id>/selecionar', methods=['POST'])
def api_selecionar_conta(account_id):
    """Seleciona conta para uso"""
    try:
        if ml_token_manager.set_current_account(account_id):
            return jsonify({
                'sucesso': True,
                'mensagem': 'Conta selecionada com sucesso!'
            })
        else:
            return jsonify({'sucesso': False, 'erro': 'Conta não encontrada'}), 404
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/contas/<account_id>/testar')
def api_testar_conta(account_id):
    """Testa se uma conta está funcionando"""
    try:
        if account_id not in ml_token_manager.accounts:
            return jsonify({'sucesso': False, 'erro': 'Conta não encontrada'}), 404
        
        account = ml_token_manager.accounts[account_id]
        token = account.get('access_token')
        
        if not token:
            return jsonify({
                'sucesso': False,
                'autenticada': False,
                'erro': 'Conta não tem token configurado'
            })
        
        # Testa o token
        headers = {'Authorization': f'Bearer {token}'}
        response = requests.get(
            'https://api.mercadolibre.com/users/me',
            headers=headers,
            timeout=10
        )
        
        if response.status_code == 200:
            user_data = response.json()
            return jsonify({
                'sucesso': True,
                'autenticada': True,
                'nickname': user_data.get('nickname'),
                'user_id': user_data.get('id')
            })
        else:
            return jsonify({
                'sucesso': False,
                'autenticada': False,
                'erro': f'Token inválido (status: {response.status_code})'
            })
            
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@app.route('/api/mercadolivre/contas/<account_id>', methods=['DELETE'])
def api_remover_conta(account_id):
    """Remove conta (não permite remover a atual)"""
    try:
        sucesso, mensagem = ml_token_manager.remove_account(account_id)
        
        if sucesso:
            return jsonify({
                'sucesso': True,
                'mensagem': mensagem
            })
        else:
            return jsonify({'sucesso': False, 'erro': mensagem}), 400
            
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
    
def obter_pedidos_anymarket_30_dias():
    """Obtém todos os pedidos dos últimos 30 dias do AnyMarket"""
    try:
        # Verifica se tem token configurado
        if not verificar_token_anymarket_configurado():
            return {
                'sucesso': False,
                'erro': 'Token do AnyMarket não configurado',
                'token_configurado': False
            }
        
        # Obtém o token
        from processamento.api_anymarket import obter_token_anymarket_seguro
        token = obter_token_anymarket_seguro()
        
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json',
            'gumgaToken': token
        }
        
        # Calcula datas dos últimos 30 dias
        hoje = datetime.now()
        data_30_dias_atras = (hoje - timedelta(days=30)).strftime('%Y-%m-%d')
        data_hoje = hoje.strftime('%Y-%m-%d')
        
        # Configurações da API
        url = "https://api.anymarket.com.br/v2/orders"
        limit = 100  # Máximo por página
        offset = 0
        all_orders = []
        
        print(f"📊 Buscando TODOS os pedidos AnyMarket de {data_30_dias_atras} até {data_hoje}")
        
        # Loop para paginação - busca TODOS os pedidos
        while True:
            params = {
                'limit': limit,
                'offset': offset,
                'createdAfter': f'{data_30_dias_atras}',
                'createdBefore': f'{data_hoje}',
                'sort': 'createdAt',
                'sortDirection': 'DESC'
            }
            
            response = requests.get(url, params=params, headers=headers, timeout=30)
            
            if response.status_code != 200:
                print(f"❌ Erro API AnyMarket: {response.status_code} - {response.text[:200]}")
                break
            
            data = response.json()
            orders = data.get('content', [])
            
            if orders:
                all_orders.extend(orders)
                print(f"✅ Página {offset//limit + 1}: {len(orders)} pedidos (total: {len(all_orders)})")
            
            # Verifica se há mais páginas
            if len(orders) < limit:
                break
            
            offset += limit
            
            # Safety limit - máximo de 500 pedidos
            if offset >= 500:
                print(f"⚠️ Limite de 500 pedidos atingido")
                break
        
        print(f"📦 Total de pedidos coletados: {len(all_orders)}")
        
        # Processa estatísticas detalhadas
        estatisticas = processar_estatisticas_detalhadas_pedidos(all_orders)
        
        return {
            'sucesso': True,
            'token_configurado': True,
            'total_pedidos': len(all_orders),
            'periodo': f'{data_30_dias_atras} até {data_hoje}',
            'estatisticas': estatisticas,
            'timestamp': datetime.now().isoformat()
        }
        
    except Exception as e:
        print(f"❌ Erro ao buscar pedidos AnyMarket: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'sucesso': False,
            'erro': str(e),
            'token_configurado': verificar_token_anymarket_configurado()
        }

@app.route("/api/anymarket/diagnosticar-imagens", methods=["POST"])
def api_diagnosticar_imagens():
    """API para diagnóstico de múltiplos produtos (AGORA ATÉ 1000 PRODUTOS)"""
    try:
        data = request.get_json()
        product_ids = data.get('product_ids', [])
        
        if not product_ids:
            return jsonify({'sucesso': False, 'erro': 'Nenhum ID fornecido'}), 400
        
        # ✅✅✅ ALTERAÇÃO PRINCIPAL: Aumentar para 1000 produtos
        LIMITE_MAXIMO = 1000
        if len(product_ids) > LIMITE_MAXIMO:
            product_ids = product_ids[:LIMITE_MAXIMO]
        
        resultados = []
        total_imagens_analisadas = 0
        total_erros = 0
        
        from processamento.api_anymarket import consultar_api_anymarket, obter_token_anymarket_seguro
        token = obter_token_anymarket_seguro()
        
        for product_id in product_ids:
            try:
                # Consultar imagens do produto
                resultado = consultar_api_anymarket(product_id, token)
                
                if resultado.get('sucesso') and resultado.get('dados'):
                    imagens = resultado['dados']
                    total_imagens_analisadas += len(imagens)
                    
                    # Analisar cada imagem
                    imagens_com_erro = []
                    for img in imagens:
                        erro = verificar_erro_imagem(img)
                        if erro:
                            imagens_com_erro.append({
                                'id': img.get('id'),
                                'index': img.get('index'),
                                'url': img.get('url'),
                                'erro': erro,
                                'campos_faltantes': identificar_campos_faltantes(img),
                                'status': img.get('status_api'),
                                'status_message': img.get('statusMessage')
                            })
                            total_erros += 1
                    
                    if imagens_com_erro:
                        resultados.append({
                            'product_id': product_id,
                            'total_imagens': len(imagens),
                            'imagens_com_erro': imagens_com_erro,
                            'resumo_erros': len(imagens_com_erro)
                        })
                else:
                    resultados.append({
                        'product_id': product_id,
                        'erro_consulta': resultado.get('erro'),
                        'imagens_com_erro': []
                    })
                    
            except Exception as e:
                resultados.append({
                    'product_id': product_id,
                    'erro': str(e),
                    'imagens_com_erro': []
                })
        
        # Gerar relatório em formato para Excel
        relatorio_excel = []
        for resultado in resultados:
            for img_erro in resultado.get('imagens_com_erro', []):
                relatorio_excel.append({
                    'ID_PRODUTO': resultado['product_id'],
                    'ID_IMG': img_erro.get('id'),
                    'POSICAO': img_erro.get('index'),
                    'TIPO_ERRO': img_erro.get('erro'),
                    'URL': img_erro.get('url'),
                    'CAMPOS_FALTANTES': ', '.join(img_erro.get('campos_faltantes', [])),
                    'STATUS_API': img_erro.get('status'),
                    'STATUS_MESSAGE': img_erro.get('status_message')
                })
        
        return jsonify({
            'sucesso': True,
            'total_produtos': len(product_ids),
            'total_imagens_analisadas': total_imagens_analisadas,
            'total_imagens_com_erro': total_erros,
            'produtos_com_erro': sum(1 for r in resultados if r.get('imagens_com_erro')),
            'resultados': resultados,
            'relatorio_excel': relatorio_excel,
            'limite_utilizado': LIMITE_MAXIMO,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

def verificar_erro_imagem(imagem):
    """Verifica se uma imagem tem erros baseado nos critérios"""
    erro = None
    
    # Critério 1: Estrutura incompleta
    campos_esperados = ['thumbnailUrl', 'lowResolutionUrl', 'standardUrl', 'originalImage']
    campos_faltantes = [campo for campo in campos_esperados if not imagem.get(campo)]
    
    if len(campos_faltantes) >= 2:  # Se faltam 2 ou mais campos importantes
        erro = f'Estrutura incompleta - Faltam {len(campos_faltantes)} campos'
    
    # Critério 2: Status de erro
    elif imagem.get('status_api') and imagem.get('status_api').lower() not in ['active', 'ok', 'success']:
        erro = f'Status problemático: {imagem.get("status_api")}'
    
    # Critério 3: URL inválida ou não disponível
    elif not imagem.get('url') or imagem.get('status') == 'indisponivel':
        erro = 'URL não disponível'
    
    return erro

def identificar_campos_faltantes(imagem):
    """Identifica quais campos estão faltando na imagem"""
    campos_esperados = ['thumbnailUrl', 'lowResolutionUrl', 'standardUrl', 'originalImage']
    return [campo for campo in campos_esperados if not imagem.get(campo)]


def processar_estatisticas_detalhadas_pedidos(orders):
    """Processa estatísticas detalhadas dos pedidos - VERSÃO CORRIGIDA"""
    if not orders:
        return {
            'total_pedidos': 0,
            'valor_total': 0,
            'total_itens': 0,
            'pedidos_por_dia': [],
            'top_produtos_quantidade': [],
            'top_produtos_valor': [],
            'status_distribuicao': {},
            'marketplace_distribuicao': {},
            'resumo': {
                'valor_total': 0,
                'total_pedidos': 0,
                'total_itens': 0,
                'ticket_medio_pedido': 0,
                'ticket_medio_item': 0,
                'media_itens_por_pedido': 0,
                'pedidos_concluidos': 0,
                'pedidos_pendentes': 0,
                'pedidos_cancelados': 0,
                'pedidos_faturados': 0,
                'marketplace_principal': 'N/A'
            }
        }
    
    # Dicionários para agrupamentos
    pedidos_por_dia = defaultdict(int)
    status_distribuicao = defaultdict(int)
    marketplace_distribuicao = defaultdict(int)
    
    # Dicionários para produtos - VERSÃO CORRIGIDA
    produtos_por_quantidade = defaultdict(int)
    produtos_por_valor = defaultdict(float)
    produtos_info = {}
    
    # Variáveis para estatísticas básicas
    valor_total = 0
    total_itens = 0
    
    for order in orders:
        # Valor total do pedido - CORRIGIDO
        try:
            valor_pedido = float(order.get('total', 0) or 0)
            valor_total += valor_pedido
        except:
            valor_pedido = 0
        
        # Data do pedido
        created_at = order.get('createdAt', '')
        if created_at:
            try:
                # Extrai apenas a data
                data_dia = created_at.split('T')[0] if 'T' in created_at else created_at[:10]
                pedidos_por_dia[data_dia] += 1
            except:
                pass
        
        # Status
        status = order.get('status', 'DESCONHECIDO')
        status_distribuicao[status] += 1
        
        # Marketplace
        marketplace = order.get('marketPlace', 'DESCONHECIDO')
        marketplace_distribuicao[marketplace] += 1
        
        # Itens do pedido - VERSÃO CORRIGIDA
        items = order.get('items', [])
        if not items:
            # Tenta obter itens de outra forma
            items = order.get('orderItems', [])
        
        for item in items:
            try:
                # Quantidade - CORRIGIDO
                quantidade = int(item.get('amount', 1) or 1)
                total_itens += quantidade
                
                # Preço - CORRIGIDO
                preco_unitario = 0
                try:
                    preco_unitario = float(item.get('price', 0) or 0)
                except:
                    # Tenta outros campos de preço
                    preco_unitario = float(item.get('value', 0) or 0)
                
                valor_item = preco_unitario * quantidade
                
                # Informações do produto - CORRIGIDO
                nome_produto = 'Produto Desconhecido'
                sku = 'SKU Desconhecido'
                
                # Tenta obter nome do produto de várias formas
                if item.get('product'):
                    nome_produto = item['product'].get('title', 'Produto Desconhecido')
                    sku = item['product'].get('partnerId', 'SKU Desconhecido')
                elif item.get('sku'):
                    nome_produto = item['sku'].get('title', 'Produto Desconhecido')
                    sku = item['sku'].get('partnerId', 'SKU Desconhecido')
                elif item.get('description'):
                    nome_produto = item['description']
                
                # Tenta obter SKU de várias formas
                if sku == 'SKU Desconhecido':
                    sku = item.get('partnerId', 'SKU Desconhecido')
                
                # Chave única para o produto
                chave_produto = f"{sku}|{nome_produto}"
                
                # DEBUG: Mostra informações do produto
                print(f"🔍 Produto encontrado: {nome_produto[:30]}... - Quantidade: {quantidade} - Preço: {preco_unitario}")
                
                # Atualiza contadores
                produtos_por_quantidade[chave_produto] += quantidade
                produtos_por_valor[chave_produto] += valor_item
                
                # Armazena informações detalhadas
                if chave_produto not in produtos_info:
                    produtos_info[chave_produto] = {
                        'nome': nome_produto,
                        'sku': sku,
                        'preco_medio': preco_unitario
                    }
                    
            except Exception as e:
                print(f"⚠️ Erro ao processar item: {str(e)}")
                continue
    
    print(f"📊 Total de itens processados: {total_itens}")
    print(f"📊 Total de produtos únicos: {len(produtos_por_quantidade)}")
    
    # Prepara top produtos por quantidade
    top_produtos_quantidade = []
    for chave, quantidade in sorted(produtos_por_quantidade.items(), key=lambda x: x[1], reverse=True)[:15]:
        info = produtos_info.get(chave, {})
        preco_medio = info.get('preco_medio', 0)
        valor_total_produto = produtos_por_valor.get(chave, 0)
        
        top_produtos_quantidade.append({
            'nome': info.get('nome', 'Desconhecido'),
            'sku': info.get('sku', 'N/A'),
            'quantidade': quantidade,
            'valor_total': valor_total_produto,
            'preco_medio': valor_total_produto / quantidade if quantidade > 0 else preco_medio
        })
    
    # Prepara top produtos por valor
    top_produtos_valor = []
    for chave, valor_total_produto in sorted(produtos_por_valor.items(), key=lambda x: x[1], reverse=True)[:15]:
        info = produtos_info.get(chave, {})
        quantidade = produtos_por_quantidade.get(chave, 0)
        
        top_produtos_valor.append({
            'nome': info.get('nome', 'Desconhecido'),
            'sku': info.get('sku', 'N/A'),
            'valor_total': valor_total_produto,
            'quantidade': quantidade,
            'preco_medio': valor_total_produto / quantidade if quantidade > 0 else info.get('preco_medio', 0)
        })
    
    # Prepara dados para gráfico de pedidos por dia (últimos 30 dias)
    pedidos_por_dia_lista = []
    for i in range(30):
        data = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        quantidade = pedidos_por_dia.get(data, 0)
        pedidos_por_dia_lista.append({
            'data': data,
            'quantidade': quantidade,
            'dia_semana': ['Dom', 'Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb'][
                (datetime.now() - timedelta(days=i)).weekday()
            ]
        })
    
    # Inverte para ordem cronológica
    pedidos_por_dia_lista.reverse()
    
    # Resumo geral
    resumo = {
        'valor_total': valor_total,
        'total_pedidos': len(orders),
        'total_itens': total_itens,
        'ticket_medio_pedido': valor_total / len(orders) if orders else 0,
        'ticket_medio_item': valor_total / total_itens if total_itens else 0,
        'media_itens_por_pedido': total_itens / len(orders) if orders else 0,
        'pedidos_concluidos': sum(1 for o in orders if str(o.get('status', '')).upper() == 'CONCLUDED'),
        'pedidos_pendentes': sum(1 for o in orders if str(o.get('status', '')).upper() == 'PENDING'),
        'pedidos_cancelados': sum(1 for o in orders if str(o.get('status', '')).upper() == 'CANCELED'),
        'pedidos_faturados': sum(1 for o in orders if str(o.get('status', '')).upper() == 'INVOICED'),
        'marketplace_principal': max(marketplace_distribuicao.items(), key=lambda x: x[1])[0] if marketplace_distribuicao else 'N/A'
    }
    
    print(f"🎯 RESUMO:")
    print(f"   Valor Total: R$ {resumo['valor_total']:.2f}")
    print(f"   Total Pedidos: {resumo['total_pedidos']}")
    print(f"   Total Itens: {resumo['total_itens']}")
    print(f"   Ticket Médio: R$ {resumo['ticket_medio_pedido']:.2f}")
    print(f"   Itens/Pedido: {resumo['media_itens_por_pedido']:.1f}")
    
    return {
        'total_pedidos': len(orders),
        'valor_total': valor_total,
        'total_itens': total_itens,
        'pedidos_por_dia': pedidos_por_dia_lista,
        'top_produtos_quantidade': top_produtos_quantidade,
        'top_produtos_valor': top_produtos_valor,
        'status_distribuicao': dict(status_distribuicao),
        'marketplace_distribuicao': dict(marketplace_distribuicao),
        'resumo': resumo
    }
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        usuario = Usuario.query.filter_by(username=username).first()
        
        if usuario and usuario.check_password(password):
            if not usuario.is_active:
                flash('Usuário desativado.', 'danger')
                return redirect(url_for('login'))
            
            login_user(usuario)
            usuario.last_login = datetime.now()
            db.session.commit()
            
            flash(f'✅ Bem-vindo, {usuario.username}! Perfil: {usuario.perfil}', 'success')
            
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)
            
            if usuario.is_master():
                return redirect(url_for('dashboard_master'))
            elif usuario.perfil == 'SAC':
                return redirect(url_for('dashboard_sac'))
            elif usuario.perfil == 'Cadastro':
                return redirect(url_for('dashboard_cadastro'))
            else:
                return redirect(url_for('dashboard'))
        else:
            flash('Usuário ou senha inválidos', 'danger')
    
    return render_template('login.html', page_title='Login')

@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard padrão (redireciona para o específico)"""
    if current_user.is_master():
        return redirect(url_for('dashboard_master'))
    elif current_user.perfil == 'SAC':
        return redirect(url_for('dashboard_sac'))
    elif current_user.perfil == 'Cadastro':
        return redirect(url_for('dashboard_cadastro'))
    return redirect(url_for('home'))

@app.route('/dashboard/master')
@login_required
@master_required
def dashboard_master():
    """Dashboard completo para Master"""
    from models import ItemProcessado  # ← Import dentro da função
    
    # Estatísticas gerais do sistema
    stats = {
        'total_usuarios': Usuario.query.count(),
        'total_processos': Processo.query.count(),
        'processos_hoje': Processo.query.filter(db.func.date(Processo.data) == datetime.now().date()).count(),
        'total_itens_processados': 0  # ← Valor padrão
    }
    
   #  Tenta buscar total de itens processados (se a tabela existir)
    try:
        total = db.session.query(db.func.sum(ItemProcessado.id)).scalar()
        stats['total_itens_processados'] = total or 0
    except Exception as e:
        print(f"Erro ao buscar itens processados: {e}")
    
    return render_template('dashboard_master.html', stats=stats, page_title='Dashboard Master')

@app.route('/dashboard/sac')
@login_required
@permissao_modulo('pedidos')
def dashboard_sac():
    """Dashboard para SAC - foco em pedidos"""
    return render_template('dashboard_sac.html', page_title='Dashboard SAC')

@app.route('/dashboard/cadastro')
@login_required
@permissao_modulo('produtos')
def dashboard_cadastro():
    """Dashboard para Cadastro - foco em produtos"""
    return render_template('dashboard_cadastro.html', page_title='Dashboard Cadastro')

@app.route('/logout')
@login_required
def logout():
    """Faz logout do usuário"""
    logout_user()
    flash('✅ Logout realizado com sucesso', 'info')
    return redirect(url_for('login'))

@app.route('/perfil')
@login_required
def perfil():
    """Página de perfil do usuário"""
    return render_template('perfil.html', page_title='Meu Perfil')

@app.route('/alterar-senha', methods=['POST'])
@login_required
def alterar_senha():
    """Altera a senha do usuário"""
    data = request.get_json()
    senha_atual = data.get('senha_atual')
    nova_senha = data.get('nova_senha')
    
    if not current_user.check_password(senha_atual):
        return jsonify({'sucesso': False, 'erro': 'Senha atual incorreta'}), 401
    
    current_user.set_password(nova_senha)
    db.session.commit()
    
    return jsonify({'sucesso': True, 'mensagem': 'Senha alterada com sucesso!'})

@app.route('/api/anymarket/diagnosticar-produto', methods=['POST'])
def api_diagnosticar_produto():
    """API para diagnóstico individual de produto - VERSÃO CORRIGIDA"""
    try:
        data = request.get_json()
        product_id = data.get('product_id')
        
        if not product_id:
            return jsonify({'sucesso': False, 'erro': 'ID do produto é obrigatório'}), 400
        
        # Consultar imagens do produto
        from processamento.api_anymarket import consultar_api_anymarket, obter_token_anymarket_seguro
        token = obter_token_anymarket_seguro()
        
        resultado = consultar_api_anymarket(product_id, token)
        
        if not resultado.get('sucesso'):
            return jsonify({
                'sucesso': False,
                'erro': resultado.get('erro', 'Erro na consulta'),
                'product_id': product_id
            })
        
        if not resultado.get('dados'):
            return jsonify({
                'sucesso': True,
                'product_id': product_id,
                'mensagem': 'Nenhuma imagem encontrada',
                'imagens_com_erro': [],
                'imagens_ok': []
            })
        
        imagens = resultado['dados']
        imagens_com_erro = []
        imagens_ok = []
        
        # Analisar cada imagem
        for img in imagens:
            erro = analisar_imagem_erro_corrigido(img)
            if erro:
                imagens_com_erro.append({
                    'id': img.get('id'),
                    'index': img.get('index'),
                    'url': img.get('url'),
                    'erro': erro['descricao'],
                    'tipo_erro': erro['tipo'],
                    'status_api': img.get('status_api'),
                    'status_message': img.get('statusMessage'),
                    'dimensoes': f"{img.get('originalWidth', 0)}x{img.get('originalHeight', 0)}"
                })
            else:
                imagens_ok.append({
                    'id': img.get('id'),
                    'index': img.get('index'),
                    'url': img.get('url'),
                    'status': 'OK',
                    'dimensoes': f"{img.get('originalWidth', 0)}x{img.get('originalHeight', 0)}"
                })
        
        # DEBUG: Log para verificar análise
        print(f"🔍 Análise produto {product_id}:")
        print(f"   Total imagens: {len(imagens)}")
        print(f"   Com erro: {len(imagens_com_erro)}")
        print(f"   OK: {len(imagens_ok)}")
        
        if imagens_com_erro:
            for erro in imagens_com_erro[:3]:  # Mostra até 3 erros
                print(f"   Erro posição {erro['index']}: {erro['tipo_erro']}")
        
        return jsonify({
            'sucesso': True,
            'product_id': product_id,
            'total_imagens': len(imagens),
            'imagens_com_erro': imagens_com_erro,
            'imagens_ok': imagens_ok,
            'resumo': {
                'com_erro': len(imagens_com_erro),
                'ok': len(imagens_ok),
                'percentual_erro': (len(imagens_com_erro) / len(imagens) * 100) if imagens else 0
            }
        })
        
    except Exception as e:
        print(f"❌ Erro no diagnóstico: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

def analisar_imagem_erro_corrigido(imagem):
    """Analisa imagem para identificar erros - VERSÃO SUPER PRECISA"""
    
    # DEBUG: Verificar estrutura da imagem
    print(f"📸 Analisando imagem {imagem.get('id')} - Posição {imagem.get('index')}")
    print(f"   Status: {imagem.get('status_api')}")
    print(f"   StatusMessage: {imagem.get('statusMessage', 'N/A')}")
    print(f"   URL: {imagem.get('url', 'N/A')[:50]}...")
    
    # CRITÉRIO 1: Status explicitamente ERROR
    status = imagem.get('status_api', '').upper()
    if status == 'ERROR':
        print(f"   ⚠️  DETECTADO: Status ERROR")
        return {
            'tipo': 'STATUS_ERROR',
            'descricao': f'ERROR: {imagem.get("statusMessage", "Imagem não processada")}'
        }
    
    # CRITÉRIO 2: StatusMessage contém erro (mesmo se status não for ERROR)
    status_message = imagem.get('statusMessage', '').lower()
    if any(erro in status_message for erro in ['problema', 'erro', 'error', 'não encontrada', 'not found', 'fail']):
        print(f"   ⚠️  DETECTADO: StatusMessage com erro")
        return {
            'tipo': 'STATUS_MESSAGE_ERROR',
            'descricao': f'StatusMessage indica erro: {imagem.get("statusMessage", "N/A")}'
        }
    
    # CRITÉRIO 3: Dimensões zeradas (0x0) - indica falha no processamento
    if (imagem.get('originalWidth') == 0 or imagem.get('originalHeight') == 0) and status != 'PROCESSED':
        print(f"   ⚠️  DETECTADO: Dimensões zeradas")
        return {
            'tipo': 'DIMENSOES_ZERADAS',
            'descricao': f'Dimensões inválidas: {imagem.get("originalWidth", 0)}x{imagem.get("originalHeight", 0)}'
        }
    
    # CRITÉRIO 4: Faltam campos essenciais de processamento
    campos_processados = ['thumbnailUrl', 'lowResolutionUrl', 'standardUrl', 'originalImage']
    campos_faltantes = [campo for campo in campos_processados if not imagem.get(campo)]
    
    if len(campos_faltantes) >= 3:  # Se faltam 3 ou mais campos de processamento
        print(f"   ⚠️  DETECTADO: Faltam {len(campos_faltantes)} campos de processamento")
        return {
            'tipo': 'CAMPOS_PROCESSAMENTO_FALTANTES',
            'descricao': f'Faltam campos de processamento: {", ".join(campos_faltantes)}'
        }
    
    # Se não encontrou nenhum erro
    print(f"   ✅ IMAGEM OK")
    return None

def obter_estatisticas_anymarket_7_dias():
    """Obtém estatísticas dos últimos 7 dias do AnyMarket"""
    try:
        # Verifica se tem token configurado
        if not verificar_token_anymarket_configurado():
            return {
                'sucesso': False,
                'erro': 'Token do AnyMarket não configurado',
                'token_configurado': False
            }
        
        # Obtém o token
        from processamento.api_anymarket import obter_token_anymarket_seguro
        token = obter_token_anymarket_seguro()
        
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json',
            'gumgaToken': token
        }
        
        # Calcula datas dos últimos 7 dias
        hoje = datetime.now()
        data_7_dias_atras = (hoje - timedelta(days=7)).strftime('%Y-%m-%d')
        data_hoje = hoje.strftime('%Y-%m-%d')
        
        # Faz a requisição para pedidos dos últimos 7 dias
        url = "https://api.anymarket.com.br/v2/orders"
        params = {
            'limit': 100,  # Aumentei o limite para pegar mais pedidos
            'offset': 0,
            'createdAfter': f'{data_7_dias_atras}T00:00:00-03:00',
            'createdBefore': f'{data_hoje}T23:59:59-03:00',
            'sort': 'createdAt',
            'sortDirection': 'DESC'
        }
        
        print(f"📊 Buscando pedidos AnyMarket de {data_7_dias_atras} até {data_hoje}")
        
        response = requests.get(url, params=params, headers=headers, timeout=30)
        
        if response.status_code != 200:
            print(f"❌ Erro API AnyMarket: {response.status_code} - {response.text[:200]}")
            return {
                'sucesso': False,
                'erro': f'Erro na API AnyMarket: {response.status_code}',
                'status_code': response.status_code,
                'token_configurado': True
            }
        
        data = response.json()
        orders = data.get('content', [])
        total_elements = data.get('page', {}).get('totalElements', 0)
        
        print(f"✅ {len(orders)} pedidos encontrados (total: {total_elements})")
        
        # Se não encontrou pedidos, pode ser problema de timezone
        if not orders:
            # Tenta sem timezone específico
            params_sem_tz = {
                'limit': 100,
                'offset': 0,
                'createdAfter': f'{data_7_dias_atras}',
                'createdBefore': f'{data_hoje}',
                'sort': 'createdAt',
                'sortDirection': 'DESC'
            }
            
            print("🔄 Tentando sem timezone específico...")
            response_sem_tz = requests.get(url, params=params_sem_tz, headers=headers, timeout=30)
            
            if response_sem_tz.status_code == 200:
                data_sem_tz = response_sem_tz.json()
                orders = data_sem_tz.get('content', [])
                total_elements = data_sem_tz.get('page', {}).get('totalElements', 0)
                print(f"✅ {len(orders)} pedidos encontrados (sem timezone)")
        
        # Processa as estatísticas
        estatisticas = processar_estatisticas_pedidos(orders)
        
        return {
            'sucesso': True,
            'token_configurado': True,
            'total_pedidos': total_elements,
            'pedidos_ultimos_7_dias': len(orders),
            'estatisticas': estatisticas,
            'timestamp': datetime.now().isoformat()
        }
        
    except Exception as e:
        print(f"❌ Erro ao buscar estatísticas AnyMarket: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'sucesso': False,
            'erro': str(e),
            'token_configurado': verificar_token_anymarket_configurado() if 'verificar_token_anymarket_configurado' in locals() else False
        }

def processar_estatisticas_pedidos(orders):
    """Processa estatísticas dos pedidos"""
    if not orders:
        return {
            'total_pedidos': 0,
            'valor_total': 0,
            'pedidos_por_dia': [],
            'top_produtos': [],
            'status_distribuicao': {},
            'marketplace_distribuicao': {}
        }
    
    # Cálculo de estatísticas
    valor_total = sum(float(order.get('total', 0)) for order in orders)
    
    # Agrupa por data
    pedidos_por_dia = {}
    for order in orders:
        created_at = order.get('createdAt', '')
        if created_at:
            # Extrai apenas a data (YYYY-MM-DD)
            data = created_at.split('T')[0] if 'T' in created_at else created_at[:10]
            pedidos_por_dia[data] = pedidos_por_dia.get(data, 0) + 1
    
    # Distribuição por status
    status_distribuicao = {}
    for order in orders:
        status = order.get('status', 'DESCONHECIDO')
        status_distribuicao[status] = status_distribuicao.get(status, 0) + 1
    
    # Distribuição por marketplace
    marketplace_distribuicao = {}
    for order in orders:
        marketplace = order.get('marketPlace', 'DESCONHECIDO')
        marketplace_distribuicao[marketplace] = marketplace_distribuicao.get(marketplace, 0) + 1
    
    # Top produtos (baseado em itens dos pedidos)
    top_produtos = {}
    for order in orders:
        items = order.get('items', [])
        for item in items:
            product_id = item.get('sku', {}).get('partnerId') or item.get('product', {}).get('title', 'Produto Desconhecido')
            quantidade = item.get('amount', 1)
            top_produtos[product_id] = top_produtos.get(product_id, 0) + quantidade
    
    # Ordena top produtos
    top_produtos_lista = sorted(top_produtos.items(), key=lambda x: x[1], reverse=True)[:10]
    
    # Formata pedidos por dia para gráfico
    pedidos_por_dia_lista = []
    for i in range(7):
        data = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        quantidade = pedidos_por_dia.get(data, 0)
        pedidos_por_dia_lista.append({
            'data': data,
            'quantidade': quantidade,
            'dia_semana': ['Dom', 'Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb'][
                (datetime.now() - timedelta(days=i)).weekday()
            ]
        })
    
    # Inverte para ordem cronológica
    pedidos_por_dia_lista.reverse()
    
    return {
        'total_pedidos': len(orders),
        'valor_total': valor_total,
        'pedidos_por_dia': pedidos_por_dia_lista,
        'top_produtos': top_produtos_lista,
        'status_distribuicao': status_distribuicao,
        'marketplace_distribuicao': marketplace_distribuicao,
        'ticket_medio': valor_total / len(orders) if orders else 0,
        'pedidos_pendentes': sum(1 for o in orders if o.get('status') == 'PENDING'),
        'pedidos_concluidos': sum(1 for o in orders if o.get('status') == 'CONCLUDED'),
        'pedidos_cancelados': sum(1 for o in orders if o.get('status') == 'CANCELED')
    }

# ROTA 1: Produto completo (atributos + qualidade)
@app.route('/api/mercadolivre/produto-completo/<mlb>')
def api_produto_completo(mlb):
    """Retorna produto com todos atributos + qualidade"""
    try:
        # Usa a instância global já configurada
        resultado = ml_api_secure.buscar_todos_atributos_produto(mlb)
        return jsonify(resultado)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500

# ROTA 2: Atributos da categoria
@app.route('/api/mercadolivre/atributos-categoria/<category_id>')
def api_atributos_categoria(category_id):
    """Retorna todos os atributos possíveis de uma categoria"""
    try:
        headers = ml_api_secure._get_headers()
        url = f"https://api.mercadolibre.com/categories/{category_id}/attributes"
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            atributos = response.json()
            atributos_filtrados = []
            for attr in atributos:
                atributos_filtrados.append({
                    'id': attr.get('id'),
                    'name': attr.get('name'),
                    'type': attr.get('type'),
                    'required': attr.get('required', False),
                    'values': [
                        {'id': v.get('id'), 'name': v.get('name')}
                        for v in attr.get('values', [])
                    ]
                })
            return jsonify({
                'sucesso': True,
                'atributos': atributos_filtrados
            })
        return jsonify({'sucesso': False, 'erro': 'Erro ao buscar categoria'})
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)})

# ROTA 3: Otimizar qualidade
@app.route('/api/mercadolivre/otimizar-qualidade', methods=['POST'])
def api_otimizar_qualidade():
    """Otimiza a qualidade do anúncio automaticamente"""
    try:
        data = request.json
        mlb = data.get('mlb')
        auto_corrigir = data.get('auto_corrigir', True)
        
        resultado = ml_api_secure.otimizar_qualidade_produto(mlb, auto_corrigir)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500

# ROTA 4: Alterar múltiplos atributos
@app.route('/api/mercadolivre/alterar-multiplos-atributos', methods=['POST'])
def api_alterar_multiplos_atributos():
    """Altera múltiplos atributos de uma vez"""
    try:
        data = request.json
        mlb = data.get('mlb')
        atributos = data.get('atributos', {})
        
        if not mlb:
            return jsonify({'sucesso': False, 'erro': 'MLB não informado'})
        
        if not atributos:
            return jsonify({'sucesso': False, 'erro': 'Nenhum atributo para alterar'})
        
        resultado = ml_api_secure.alterar_multiplos_atributos(mlb, atributos)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500

# ROTA 5: Dashboard do produto (HTML)
@app.route('/produto/dashboard')
def produto_dashboard():
    """Tela de dashboard do produto"""
    return render_template('mercadolivre/produto_dashboard.html')

def obter_dados_completos_perfil():
    """Obtém dados básicos do perfil - VERSÃO CORRIGIDA"""
    try:
        token = ml_token_manager.get_valid_token()
        if not token:
            return {'erro': 'Token não disponível'}
        
        headers = {'Authorization': f'Bearer {token}'}
        base_url = "https://api.mercadolibre.com"
        
        # Dados básicos do usuário
        response_user = requests.get(f"{base_url}/users/me", headers=headers, timeout=10)
        if response_user.status_code != 200:
            return {'erro': 'Erro ao buscar dados do usuário'}
        
        user_data = response_user.json()
        user_id = user_data['id']
        print(f"👤 Usuário: {user_data.get('nickname')} (ID: {user_id})")
        
        # Buscar seller_reputation para nível da conta - MAIS CONFIÁVEL
        nivel_conta = "Não identificado"
        try:
            response_reputation = requests.get(
                f"{base_url}/users/{user_id}/seller_reputation",
                headers=headers,
                timeout=10
            )
            if response_reputation.status_code == 200:
                reputation_data = response_reputation.json()
                print(f"📊 Seller reputation: {reputation_data}")
                
                # Extrair nível de diferentes campos possíveis
                if 'level_id' in reputation_data:
                    nivel_conta = reputation_data['level_id']
                elif 'seller_level' in reputation_data:
                    nivel_conta = reputation_data['seller_level']
                elif 'power_seller_status' in reputation_data:
                    nivel_conta = reputation_data['power_seller_status']
                    
                print(f"✅ Nível da conta: {nivel_conta}")
            else:
                print(f"❌ Erro seller_reputation: {response_reputation.status_code}")
        except Exception as e:
            print(f"❌ Erro nível: {e}")
        
        # Buscar vendas com PAGINAÇÃO para pegar mais de 50
        data_30_dias_atras = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%dT%H:%M:%S.000-03:00')
        
        todas_vendas = []
        offset = 0
        limit = 50
        
        while True:
            params_vendas = {
                'seller': user_id,
                'order.date_created.from': data_30_dias_atras,
                'sort': 'date_desc',
                'limit': limit,
                'offset': offset
            }
            
            response_vendas = requests.get(
                f"{base_url}/orders/search", 
                headers=headers, 
                params=params_vendas, 
                timeout=20
            )
            
            if response_vendas.status_code != 200:
                print(f"❌ Erro vendas (offset {offset}): {response_vendas.status_code}")
                break
            
            vendas_data = response_vendas.json()
            vendas_lote = vendas_data.get('results', [])
            
            if not vendas_lote:
                break
                
            todas_vendas.extend(vendas_lote)
            print(f"✅ Lote {offset//limit + 1}: {len(vendas_lote)} vendas")
            
            # Se veio menos que o limite, é a última página
            if len(vendas_lote) < limit:
                break
                
            offset += limit
            
            # Limitar a 200 vendas máximo para não sobrecarregar
            if offset >= 200:
                break
        
        print(f"📈 Total de vendas encontradas: {len(todas_vendas)}")
        
        # Buscar anúncios ativos
        anuncios_ativos = 0
        try:
            response_anuncios = requests.get(
                f"{base_url}/users/{user_id}/items/search?status=active&limit=50",
                headers=headers,
                timeout=10
            )
            if response_anuncios.status_code == 200:
                anuncios_data = response_anuncios.json()
                anuncios_ativos = len(anuncios_data.get('results', []))
                print(f"✅ Anúncios ativos: {anuncios_ativos}")
        except Exception as e:
            print(f"❌ Erro anúncios: {e}")
        
        # Calcular métricas básicas
        total_vendas = len(todas_vendas)
        valor_total_vendas = sum(float(order.get('total_amount', order.get('total', 0))) for order in todas_vendas)
        
        # Processar últimas vendas
        ultimas_vendas_detalhadas = []
        for venda in todas_vendas[:10]:  # Apenas 10 últimas para mostrar
            total_venda = venda.get('total_amount') or venda.get('total') or 0
            
            ultimas_vendas_detalhadas.append({
                'id': venda.get('id', 'N/A'),
                'data': venda.get('date_created', 'N/A'),
                'status': venda.get('status', 'N/A'),
                'total': float(total_venda),
                'itens': len(venda.get('order_items', [])),
                'comprador': venda.get('buyer', {}).get('nickname', 'N/A')
            })
        
        return {
            'dados_usuario': user_data,
            'nivel_conta': nivel_conta,
            'metricas': {
                'total_vendas_30_dias': total_vendas,
                'valor_total_vendas': valor_total_vendas,
                'ticket_medio': valor_total_vendas / total_vendas if total_vendas > 0 else 0,
                'anuncios_ativos': anuncios_ativos,
                'vendas_por_dia': round(total_vendas / 30, 2) if total_vendas > 0 else 0,
                'total_vendas_encontradas': total_vendas  # Para mostrar no template
            },
            'ultimas_vendas': ultimas_vendas_detalhadas,
            'timestamp': datetime.now().isoformat()
        }
        
    except Exception as e:
        print(f"❌ Erro ao buscar dados: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'erro': f'Erro interno: {str(e)}'}

# ──────────────────────────────────────────────────────────────
# WEBHOOK — recebe notificações do Mercado Livre
# ──────────────────────────────────────────────────────────────
@app.route('/mercadolivre/oauth/webhook', methods=['POST'])
def webhook_mercadolivre():
    """
    Endpoint registrado no painel do seu aplicativo ML.
    Salva cada evento no banco e responde 200 imediatamente,
    evitando que o ML fique reenviando indefinidamente.
    """
    from models import MLWebhookEvent
    import json as _json

    data = request.get_json(silent=True) or {}

    topic          = data.get('topic') or data.get('type', 'unknown')
    resource       = data.get('resource', '')
    user_id        = str(data.get('user_id', ''))
    attempts       = int(data.get('attempts', 1))
    application_id = str(data.get('application_id', ''))

    try:
        evento = MLWebhookEvent(
            topic          = topic,
            resource       = resource,
            user_id        = user_id,
            attempts       = attempts,
            application_id = application_id,
            payload        = _json.dumps(data),
            processed      = False,
        )
        db.session.add(evento)
        db.session.commit()
        app.logger.info(f"[ML-Webhook] topic={topic} resource={resource} id={evento.id}")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"[ML-Webhook] Erro ao salvar evento: {e}")

    # Sempre responde 200 — obrigatório para o ML parar de reenviar
    return "", 200


# ──────────────────────────────────────────────────────────────
# API — dados para o dashboard de eventos
# ──────────────────────────────────────────────────────────────
@app.route('/api/ml/webhook-events')
@login_required
def api_webhook_events():
    """Retorna últimos 200 eventos + contadores por tópico (últimos 7 dias)."""
    from models import MLWebhookEvent
    from sqlalchemy import func

    eventos = (
        MLWebhookEvent.query
        .order_by(MLWebhookEvent.received_at.desc())
        .limit(200)
        .all()
    )

    sete_dias = datetime.utcnow() - timedelta(days=7)
    contadores = (
        db.session.query(MLWebhookEvent.topic, func.count(MLWebhookEvent.id))
        .filter(MLWebhookEvent.received_at >= sete_dias)
        .group_by(MLWebhookEvent.topic)
        .all()
    )

    return jsonify({
        'eventos':    [e.to_dict() for e in eventos],
        'contadores': {t: c for t, c in contadores},
        'total':      len(eventos),
    })


# ──────────────────────────────────────────────────────────────
# DASHBOARD — página de eventos do Mercado Livre
# ──────────────────────────────────────────────────────────────
@app.route('/dashboard/ml-eventos')
@login_required
def dashboard_ml_eventos():
    """Renderiza o dashboard de eventos do Mercado Livre."""
    return render_template('mercadolivre/dashboard_eventos.html')



# ============================================
# API PARA DASHBOARD MASTER - DADOS DOS WEBHOOKS
# ============================================

# ============================================
# API PARA DASHBOARD MASTER - DADOS DOS WEBHOOKS
# ============================================
 
@app.route('/api/ml/master/resumo')
@login_required
@master_required
@cache.cached(timeout=60)
def api_master_resumo():
    from models import MLWebhookEvent
    from sqlalchemy import func, and_, or_
    from datetime import datetime, timedelta
 
    try:
        data_30d = datetime.utcnow() - timedelta(days=30)
        data_7d  = datetime.utcnow() - timedelta(days=7)
 
        eventos_30d = MLWebhookEvent.query.filter(
            MLWebhookEvent.received_at >= data_30d
        ).all()
 
        catalogos_novos    = 0
        sugestoes_catalogo = 0
        anuncios_pausados  = 0
        anuncios_excluidos = 0
        anuncios_reativados = 0
 
        for evento in eventos_30d:
            resource = (evento.resource or '').lower()
            topic    = (evento.topic    or '').lower()
            payload  = evento.get_data()
 
            if 'items' in resource or 'item' in topic:
                status = str(payload.get('status', '')).lower()
                if 'paused' in status or 'pause' in resource:
                    anuncios_pausados += 1
                elif 'deleted' in status or 'closed' in status or 'delete' in resource:
                    anuncios_excluidos += 1
                elif 'active' in status:
                    anuncios_reativados += 1
            elif 'catalog' in resource or 'catalog' in topic:
                catalogos_novos += 1
            elif 'suggestion' in resource or 'benchmark' in resource:
                sugestoes_catalogo += 1
 
        total_7d = MLWebhookEvent.query.filter(
            MLWebhookEvent.received_at >= data_7d
        ).count()
 
        # Volume diário (últimos 7 dias)
        volume_diario = []
        for i in range(6, -1, -1):
            dia       = (datetime.utcnow() - timedelta(days=i)).date()
            dia_inicio = datetime(dia.year, dia.month, dia.day)
            dia_fim    = dia_inicio + timedelta(days=1)
            qtd = MLWebhookEvent.query.filter(
                and_(MLWebhookEvent.received_at >= dia_inicio,
                     MLWebhookEvent.received_at <  dia_fim)
            ).count()
            volume_diario.append({'data': dia.strftime('%d/%m'), 'total': qtd})
 
        # Top sellers
        top_sellers = db.session.query(
            MLWebhookEvent.user_id,
            func.count(MLWebhookEvent.id).label('total')
        ).filter(
            MLWebhookEvent.received_at >= data_30d,
            MLWebhookEvent.user_id.isnot(None),
            MLWebhookEvent.user_id != '',
            MLWebhookEvent.user_id != 'null'
        ).group_by(MLWebhookEvent.user_id).order_by(
            func.count(MLWebhookEvent.id).desc()
        ).limit(10).all()
 
        # Alertas de anúncios
        anuncios_alertas = []
        eventos_items = MLWebhookEvent.query.filter(
            and_(MLWebhookEvent.received_at >= data_30d,
                 or_(MLWebhookEvent.resource.like('%items%'),
                     MLWebhookEvent.topic.like('%item%')))
        ).order_by(MLWebhookEvent.received_at.desc()).limit(20).all()
 
        for a in eventos_items:
            payload  = a.get_data()
            resource = (a.resource or '').lower()
            if 'paused' in resource or 'pause' in resource:
                tipo = 'pausado'
            elif 'deleted' in resource or 'closed' in resource:
                tipo = 'excluido'
            else:
                tipo = 'alerta'
            anuncios_alertas.append({
                'id': a.id, 'tipo': tipo,
                'resource': a.resource, 'user_id': a.user_id,
                'status': payload.get('status', 'N/A'),
                'received_at': a.received_at.isoformat()
            })
 
        # Catálogos
        catalogos = MLWebhookEvent.query.filter(
            or_(MLWebhookEvent.resource.like('%catalog%'),
                MLWebhookEvent.topic.like('%catalog%'))
        ).order_by(MLWebhookEvent.received_at.desc()).limit(50).all()
 
        catalogos_list = [{
            'id': c.id, 'topic': c.topic,
            'resource': c.resource, 'user_id': c.user_id,
            'status': c.get_data().get('status', 'N/A'),
            'received_at': c.received_at.isoformat()
        } for c in catalogos]
 
        # Sugestões
        sugestoes = MLWebhookEvent.query.filter(
            or_(MLWebhookEvent.resource.like('%benchmark%'),
                MLWebhookEvent.topic.like('%suggestion%'))
        ).order_by(MLWebhookEvent.received_at.desc()).limit(50).all()
 
        sugestoes_list = [{
            'id': s.id, 'resource': s.resource,
            'user_id': s.user_id, 'status': 'sugestão',
            'received_at': s.received_at.isoformat()
        } for s in sugestoes]
 
        # Timeline
        timeline = MLWebhookEvent.query.order_by(
            MLWebhookEvent.received_at.desc()
        ).limit(100).all()
 
        timeline_list = [{
            'id': t.id, 'topic': t.topic,
            'resource': t.resource, 'user_id': t.user_id,
            'attempts': t.attempts, 'processed': t.processed,
            'received_at': t.received_at.isoformat()
        } for t in timeline]
 
        return jsonify({
            'kpis': {
                'catalogos_novos':      catalogos_novos,
                'sugestoes_catalogo':   sugestoes_catalogo,
                'anuncios_pausados':    anuncios_pausados,
                'anuncios_excluidos':   anuncios_excluidos,
                'anuncios_reativados':  anuncios_reativados,
                'total_7d':             total_7d,
                'total_catalog_eventos': len(catalogos_list)
            },
            'volume_diario':  volume_diario,
            'top_sellers':    [{'user_id': s.user_id, 'total': s.total} for s in top_sellers],
            'anuncios_alertas': anuncios_alertas,
            'catalogos':      catalogos_list,
            'sugestoes':      sugestoes_list,
            'timeline':       timeline_list
        })
 
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================
# API PARA DASHBOARD SAC
# ============================================

# ============================================
# API PARA DASHBOARD SAC
# ============================================
 
# Tópicos reconhecidos pelo SAC
_SAC_QUESTION_TOPICS  = ('questions', 'messages')
_SAC_ORDER_TOPICS     = ('orders', 'shipments')
_SAC_CLAIM_TOPICS     = ('claims', 'mediations', 'complaints', 'post_sale')
_SAC_PAYMENT_TOPICS   = ('payments',)
 
 
def _sac_extract_payload(evento):
    """Extrai campos úteis do payload do webhook."""
    try:
        import json
        data = json.loads(evento.payload or '{}')
    except Exception:
        data = {}
    return data
 
 
def _sac_classify_topic(topic):
    """Retorna categoria do tópico: question | order | claim | payment | other."""
    t = (topic or '').lower()
    if any(k in t for k in _SAC_QUESTION_TOPICS):
        return 'question'
    if any(k in t for k in _SAC_ORDER_TOPICS):
        return 'order'
    if any(k in t for k in _SAC_CLAIM_TOPICS):
        return 'claim'
    if any(k in t for k in _SAC_PAYMENT_TOPICS):
        return 'payment'
    return 'other'
 
 
def _extract_question_text(payload, resource):
    """Tenta extrair o texto da pergunta do payload ou retorna o resource."""
    text = (payload.get('text')
            or payload.get('question', {}).get('text')
            or payload.get('body')
            or '')
    return text[:300] if text else resource
 
 
def _extract_claim_status(payload):
    """Tenta extrair o status da reclamação."""
    status = (payload.get('status')
              or payload.get('resolution', {}).get('reason')
              or payload.get('stage')
              or 'aberta')
    return str(status).lower()
 
 
def _extract_buyer_info(payload):
    """Tenta extrair nome/id do comprador."""
    buyer = (payload.get('buyer', {}) or {})
    return buyer.get('nickname') or buyer.get('id') or ''
 
 
@app.route('/api/ml/sac/resumo')
@login_required
@permissao_modulo('pedidos')
@cache.cached(timeout=60)
def api_sac_resumo():
    from models import MLWebhookEvent
    from sqlalchemy import func, and_, or_
    from datetime import datetime, timedelta
 
    try:
        data_30d     = datetime.utcnow() - timedelta(days=30)
        data_7d      = datetime.utcnow() - timedelta(days=7)
        hoje_inicio  = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
 
        # ── Filtros base ───────────────────────────────────────────────
        filtro_questions = or_(
            *[MLWebhookEvent.topic.ilike(f'%{t}%') for t in _SAC_QUESTION_TOPICS]
        )
        filtro_orders = or_(
            *[MLWebhookEvent.topic.ilike(f'%{t}%') for t in _SAC_ORDER_TOPICS],
            MLWebhookEvent.resource.ilike('%order%'),
            MLWebhookEvent.resource.ilike('%shipment%')
        )
        filtro_claims = or_(
            *[MLWebhookEvent.topic.ilike(f'%{t}%') for t in _SAC_CLAIM_TOPICS]
        )
 
        # ── KPIs perguntas ─────────────────────────────────────────────
        total_perguntas      = MLWebhookEvent.query.filter(filtro_questions).count()
        perguntas_hoje       = MLWebhookEvent.query.filter(
            and_(MLWebhookEvent.received_at >= hoje_inicio, filtro_questions)
        ).count()
        perguntas_7d         = MLWebhookEvent.query.filter(
            and_(MLWebhookEvent.received_at >= data_7d, filtro_questions)
        ).count()
 
        # ── KPIs reclamações ───────────────────────────────────────────
        total_reclamacoes    = MLWebhookEvent.query.filter(filtro_claims).count()
        reclamacoes_7d       = MLWebhookEvent.query.filter(
            and_(MLWebhookEvent.received_at >= data_7d, filtro_claims)
        ).count()
        reclamacoes_hoje     = MLWebhookEvent.query.filter(
            and_(MLWebhookEvent.received_at >= hoje_inicio, filtro_claims)
        ).count()
 
        # ── KPIs pedidos / pagamentos ──────────────────────────────────
        pedidos_7d   = MLWebhookEvent.query.filter(
            and_(MLWebhookEvent.received_at >= data_7d, filtro_orders)
        ).count()
        pagamentos_7d = MLWebhookEvent.query.filter(
            and_(MLWebhookEvent.received_at >= data_7d,
                 MLWebhookEvent.topic.ilike('%payment%'))
        ).count()
 
        # ── Lista perguntas (últimas 80, 30 dias) ──────────────────────
        perguntas_eventos = MLWebhookEvent.query.filter(
            and_(MLWebhookEvent.received_at >= data_30d, filtro_questions)
        ).order_by(MLWebhookEvent.received_at.desc()).limit(80).all()
 
        perguntas_pendentes  = 0
        perguntas_respondidas = 0
        perguntas_list = []
 
        for e in perguntas_eventos:
            payload = _sac_extract_payload(e)
            # status da pergunta vem do payload, não do campo processed
            status_raw = (payload.get('status') or payload.get('answer', {}).get('status') or '').lower()
            if status_raw in ('answered', 'respondida', 'closed_by_seller'):
                status_label = 'respondida'
                perguntas_respondidas += 1
            else:
                status_label = 'pendente'
                perguntas_pendentes += 1
 
            perguntas_list.append({
                'id':          e.id,
                'resource':    e.resource,
                'user_id':     e.user_id,
                'status':      status_label,
                'texto':       _extract_question_text(payload, e.resource),
                'comprador':   _extract_buyer_info(payload),
                'attempts':    e.attempts,
                'received_at': e.received_at.isoformat() if e.received_at else None,
            })
 
        # ── Lista reclamações (últimas 60, 30 dias) ────────────────────
        reclamacoes_eventos = MLWebhookEvent.query.filter(
            and_(MLWebhookEvent.received_at >= data_30d, filtro_claims)
        ).order_by(MLWebhookEvent.received_at.desc()).limit(60).all()
 
        reclamacoes_abertas    = 0
        reclamacoes_resolvidas = 0
        reclamacoes_list = []
 
        for e in reclamacoes_eventos:
            payload    = _sac_extract_payload(e)
            status_raw = _extract_claim_status(payload)
 
            if any(k in status_raw for k in ('closed', 'resolved', 'encerrada', 'resolvida')):
                status_label = 'resolvida'
                reclamacoes_resolvidas += 1
            else:
                status_label = 'aberta'
                reclamacoes_abertas += 1
 
            reclamacoes_list.append({
                'id':          e.id,
                'topic':       e.topic,
                'resource':    e.resource,
                'user_id':     e.user_id,
                'status':      status_label,
                'comprador':   _extract_buyer_info(payload),
                'motivo':      payload.get('reason') or payload.get('resolution_reason') or '',
                'attempts':    e.attempts,
                'received_at': e.received_at.isoformat() if e.received_at else None,
            })
 
        # ── Lista pedidos ──────────────────────────────────────────────
        pedidos_eventos = MLWebhookEvent.query.filter(
            and_(MLWebhookEvent.received_at >= data_7d, filtro_orders)
        ).order_by(MLWebhookEvent.received_at.desc()).limit(30).all()
 
        pedidos_list = [{
            'id':          p.id,
            'resource':    p.resource,
            'user_id':     p.user_id,
            'status':      p.get_data().get('status', 'recebido'),
            'attempts':    p.attempts,
            'received_at': p.received_at.isoformat() if p.received_at else None,
        } for p in pedidos_eventos]
 
        # ── Volume diário de perguntas (7 dias) ────────────────────────
        volume_perguntas = []
        for i in range(6, -1, -1):
            dia        = (datetime.utcnow() - timedelta(days=i)).date()
            dia_inicio = datetime(dia.year, dia.month, dia.day)
            dia_fim    = dia_inicio + timedelta(days=1)
            qtd = MLWebhookEvent.query.filter(
                and_(MLWebhookEvent.received_at >= dia_inicio,
                     MLWebhookEvent.received_at <  dia_fim,
                     filtro_questions)
            ).count()
            volume_perguntas.append({'data': dia.strftime('%d/%m'), 'total': qtd})
 
        # ── Volume diário de reclamações (7 dias) ──────────────────────
        volume_reclamacoes = []
        for i in range(6, -1, -1):
            dia        = (datetime.utcnow() - timedelta(days=i)).date()
            dia_inicio = datetime(dia.year, dia.month, dia.day)
            dia_fim    = dia_inicio + timedelta(days=1)
            qtd = MLWebhookEvent.query.filter(
                and_(MLWebhookEvent.received_at >= dia_inicio,
                     MLWebhookEvent.received_at <  dia_fim,
                     filtro_claims)
            ).count()
            volume_reclamacoes.append({'data': dia.strftime('%d/%m'), 'total': qtd})
 
        return jsonify({
            'kpis': {
                'perguntas_total':        total_perguntas,
                'perguntas_pendentes':    perguntas_pendentes,
                'perguntas_respondidas':  perguntas_respondidas,
                'perguntas_hoje':         perguntas_hoje,
                'perguntas_7d':           perguntas_7d,
                'reclamacoes_total':      total_reclamacoes,
                'reclamacoes_abertas':    reclamacoes_abertas,
                'reclamacoes_resolvidas': reclamacoes_resolvidas,
                'reclamacoes_hoje':       reclamacoes_hoje,
                'reclamacoes_7d':         reclamacoes_7d,
                'pedidos_7d':             pedidos_7d,
                'pagamentos_7d':          pagamentos_7d,
            },
            'perguntas':           perguntas_list,
            'reclamacoes':         reclamacoes_list,
            'pedidos':             pedidos_list,
            'volume_perguntas':    volume_perguntas,
            'volume_reclamacoes':  volume_reclamacoes,
        })
 
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 5000


# Import necessário para usar o 'or_' nas queries
from sqlalchemy import or_

@app.route('/api/ml/debug-banco')
@login_required  
def debug_banco():
    from sqlalchemy import text
    
    # Todas as tabelas que existem no banco
    tabelas = db.session.execute(text("""
        SELECT table_name 
        FROM information_schema.tables 
        WHERE table_schema = 'public'
        ORDER BY table_name
    """)).fetchall()

    # Contar registros em qualquer tabela que tenha "webhook" ou "ml" no nome
    contagens = []
    for (tabela,) in tabelas:
        if any(k in tabela.lower() for k in ['webhook', 'ml_', 'mercado']):
            try:
                total = db.session.execute(
                    text(f'SELECT COUNT(*) FROM "{tabela}"')
                ).scalar()
                contagens.append({'tabela': tabela, 'total': total})
            except Exception as e:
                contagens.append({'tabela': tabela, 'erro': str(e)})

    return jsonify({
        'todas_tabelas': [t[0] for t in tabelas],
        'tabelas_ml': contagens
    })

@app.route('/api/ml/testar-dados')
@login_required
def testar_dados_webhook():
    from models import MLWebhookEvent
    
    total = MLWebhookEvent.query.count()
    ultimos = MLWebhookEvent.query.order_by(MLWebhookEvent.received_at.desc()).limit(5).all()
    
    return jsonify({
        'total_eventos': total,
        'ultimos_eventos': [
            {
                'id': e.id,
                'topic': e.topic,
                'resource': e.resource,
                'received_at': e.received_at.isoformat()
            } for e in ultimos
        ]
    })

@app.route('/api/mercadolivre/qualidade-detalhada/<mlb>')
def api_qualidade_detalhada(mlb):
    """
    Retorna a qualidade detalhada do anúncio com buckets e variáveis de melhoria.
    Endpoint oficial: /item/{mlb}/performance
    """
    try:
        from mercadolivre_api_secure import ml_api_secure
        
        headers = ml_api_secure._get_headers()
        url = f"https://api.mercadolibre.com/item/{mlb}/performance"
        
        response = requests.get(url, headers=headers, timeout=15)
        
        if response.status_code != 200:
            # Fallback: cálculo local
            resultado_produto = ml_api_secure.buscar_todos_atributos_produto(mlb)
            if resultado_produto.get('sucesso'):
                qualidade = resultado_produto.get('qualidade', {})
                return jsonify({
                    'sucesso': True,
                    'score': qualidade.get('pontuacao', 0),
                    'nivel': qualidade.get('nivel', 'Básica'),
                    'buckets': [],
                    'dicas': qualidade.get('dicas', []),
                    'origem': 'calculado'
                })
            return jsonify({'sucesso': False, 'erro': 'Não foi possível buscar qualidade'})
        
        dados = response.json()
        
        # Processa os buckets em um formato amigável para o frontend
        buckets_processados = []
        todas_dicas = []
        
        for bucket in dados.get('buckets', []):
            bucket_info = {
                'key': bucket.get('key'),
                'title': bucket.get('title'),
                'status': bucket.get('status'),
                'score': bucket.get('score', 0),
                'variables': []
            }
            
            for var in bucket.get('variables', []):
                # Cada variável contém regras com links
                for rule in var.get('rules', []):
                    if rule.get('status') == 'PENDING':
                        wordings = rule.get('wordings', {})
                        acao = {
                            'key': var.get('key'),
                            'title': var.get('title'),
                            'label': wordings.get('label', 'Melhorar'),
                            'link': wordings.get('link', ''),
                            'progress': rule.get('progress', 0),
                            'score': var.get('score', 0)
                        }
                        bucket_info['variables'].append(acao)
                        todas_dicas.append(acao)
            
            buckets_processados.append(bucket_info)
        
        # Mapeia nível para português
        nivel_map = {
            'bad': 'Básica',
            'medium': 'Satisfatória',
            'good': 'Profissional',
            'professional': 'Profissional'
        }
        nivel = dados.get('level_wording') or nivel_map.get(dados.get('level', ''), 'Básica')
        
        return jsonify({
            'sucesso': True,
            'score': dados.get('score', 0),
            'nivel': nivel,
            'buckets': buckets_processados,
            'dicas': todas_dicas[:20],
            'entity_id': dados.get('entity_id'),
            'calculated_at': dados.get('calculated_at')
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
    
if __name__ == "__main__":
    app.run(debug=False)