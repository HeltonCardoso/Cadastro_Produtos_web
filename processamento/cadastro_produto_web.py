import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
import warnings
import logging
import re
import unicodedata
import gspread
from google.oauth2 import service_account

# Tentativa de importar a função original (fallback)
try:
    from processamento.google_sheets import ler_planilha_google
except ImportError:
    ler_planilha_google = None

PLANILHA_MODELO_FIXA = "Template_Produtos_Mpozenato_Cadastro_.xlsx"

warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")

logger = logging.getLogger('cadastro')


# ============================================
# FUNÇÕES PARA CREDENCIAIS DO GOOGLE (CORRIGIDAS)
# ============================================

def _get_credentials_path():
    """
    Encontra o arquivo credentials.json em múltiplos locais.
    Prioriza o Secret File do Render.
    """
    # 1. Secret File do Render (prioridade máxima)
    secret_path = '/etc/secrets/credentials.json'
    if os.path.exists(secret_path):
        print(f"✅ Credentials encontrado em Secret File: {secret_path}")
        return secret_path
    
    # 2. Pasta processamento
    local_path = Path(__file__).parent / "credentials.json"
    if local_path.exists():
        print(f"✅ Credentials encontrado em: {local_path}")
        return str(local_path)
    
    # 3. Raiz do projeto
    root_path = Path('credentials.json')
    if root_path.exists():
        print(f"✅ Credentials encontrado em: {root_path}")
        return str(root_path)
    
    # 4. Pasta config
    config_path = Path('config/credentials.json')
    if config_path.exists():
        print(f"✅ Credentials encontrado em: {config_path}")
        return str(config_path)
    
    # 5. Pasta modelos
    modelos_path = Path('modelos/credentials.json')
    if modelos_path.exists():
        print(f"✅ Credentials encontrado em: {modelos_path}")
        return str(modelos_path)
    
    raise FileNotFoundError(
        "credentials.json não encontrado. Verifique se o Secret File está configurado no Render "
        "ou se o arquivo existe no projeto."
    )


def obter_cliente_google():
    """Retorna cliente autenticado do Google Sheets"""
    creds_path = _get_credentials_path()
    credentials = service_account.Credentials.from_service_account_file(
        creds_path,
        scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    return gspread.authorize(credentials)


def ler_planilha_google_com_credencial(sheet_id, aba_nome):
    """
    Lê uma planilha do Google Sheets usando as credenciais corretas.
    Esta é a VERSÃO CORRIGIDA que funciona com o Secret File do Render.
    """
    try:
        print(f"📊 Lendo planilha Google: ID={sheet_id}, Aba={aba_nome}")
        
        client = obter_cliente_google()
        planilha = client.open_by_key(sheet_id)
        worksheet = planilha.worksheet(aba_nome)
        
        # Obtém todos os dados
        dados = worksheet.get_all_values()
        
        if not dados or len(dados) == 0:
            return pd.DataFrame()
        
        # Primeira linha como cabeçalho
        cabecalhos = dados[0]
        # Restante dos dados
        dados_linhas = dados[1:]
        
        # Converte para DataFrame
        df = pd.DataFrame(dados_linhas, columns=cabecalhos)
        
        print(f"✅ Planilha carregada: {len(df)} linhas, {len(df.columns)} colunas")
        return df
        
    except Exception as e:
        raise Exception(f"Erro ao acessar Google Sheets: {str(e)}")


def ler_planilha_google_fallback(sheet_id, aba_nome):
    """
    Função de fallback que tenta usar o módulo original se disponível,
    ou a versão corrigida.
    """
    # Tenta usar a versão corrigida primeiro
    try:
        return ler_planilha_google_com_credencial(sheet_id, aba_nome)
    except Exception as e:
        print(f"⚠️ Erro na versão corrigida: {e}")
        
        # Fallback para a função original (se existir)
        if ler_planilha_google:
            try:
                print("🔄 Tentando usar função original ler_planilha_google...")
                return ler_planilha_google(sheet_id, aba_nome)
            except Exception as e2:
                print(f"⚠️ Função original também falhou: {e2}")
        
        raise Exception(f"Não foi possível ler a planilha: {str(e)}")


# ============================================
# FUNÇÕES AUXILIARES EXISTENTES
# ============================================

def sanitize_filename(texto):
    """Remove acentos, caracteres especiais e sanitiza para nome de arquivo"""
    texto = unicodedata.normalize('NFKD', str(texto))
    texto = texto.encode('ASCII', 'ignore').decode('ASCII')
    return re.sub(r'[^\w\-_]', '', texto).strip().replace(' ', '_')


def copiar_validacoes(worksheet):
    return list(worksheet.data_validations.dataValidation)


def reaplicar_validacoes(worksheet, validacoes):
    for dv in validacoes:
        worksheet.add_data_validation(dv)


def limpar_moeda(valor):
    if pd.isna(valor):
        return None

    valor = str(valor)

    return (
        valor
        .replace("R$", "")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", ".")
    )


def ajustar_decimal(valor):
    if pd.isna(valor):
        return None

    valor = str(valor).strip()

    # Corrige decimal brasileiro
    if "," in valor and "." not in valor:
        valor = valor.replace(",", ".")

    return valor


# ============================================
# FUNÇÃO PRINCIPAL DE PROCESSAMENTO
# ============================================

def executar_processamento(planilha_origem, planilha_destino=None):
    """
    Executa o processamento de cadastro de produtos.
    
    Args:
        planilha_origem: Pode ser:
            - dict com 'sheet_id' e 'aba' (Google Sheets)
            - string com caminho de arquivo Excel local
        planilha_destino: Caminho da planilha destino (opcional)
    
    Returns:
        tuple: (caminho_saida, qtd_produtos, tempo_segundos, produtos_processados)
    """
    inicio = datetime.now()
    produtos_processados = []

    # ============================================
    # DEFINE PLANILHA DESTINO (MODELO FIXO)
    # ============================================
    if planilha_destino is None:
        # Procura o modelo em possíveis locais
        possiveis_modelos = [
            Path(__file__).parent.parent / "modelos" / PLANILHA_MODELO_FIXA,
            Path(PLANILHA_MODELO_FIXA),
            Path("modelos") / PLANILHA_MODELO_FIXA,
        ]
        
        modelo_path = None
        for path in possiveis_modelos:
            if path.exists():
                modelo_path = path
                break
        
        if not modelo_path:
            raise Exception(f"Modelo fixo não encontrado. Procurei em: {possiveis_modelos}")
        
        planilha_destino = str(modelo_path)
        print(f"✅ Usando modelo fixo: {planilha_destino}")

    # ============================================
    # LÊ PLANILHA ORIGEM
    # ============================================
    print("📥 Lendo planilha de origem...")
    
    if isinstance(planilha_origem, dict):
        # Origem: Google Sheets (USANDO VERSÃO CORRIGIDA)
        sheet_id = planilha_origem['sheet_id']
        aba_nome = planilha_origem['aba']
        print(f"📊 Lendo do Google Sheets: {sheet_id} / {aba_nome}")
        df = ler_planilha_google_fallback(sheet_id, aba_nome)
    else:
        # Origem: Arquivo local
        print(f"📁 Lendo arquivo local: {planilha_origem}")
        df = pd.read_excel(planilha_origem)

    if df.empty:
        raise Exception("A planilha de origem está vazia")

    print(f"✅ Planilha carregada: {len(df)} linhas")

    # ============================================
    # LIMPEZA E VALIDAÇÃO DOS DADOS
    # ============================================
    
    # Remove linhas completamente vazias
    df = df.dropna(how="all")

    # Remove repetições de cabeçalhos no meio da planilha
    df = df[~df.apply(
        lambda row: all(str(row[c]).strip().upper() == c.upper() for c in df.columns if c in row),
        axis=1
    )]

    # Colunas esperadas
    colunas_esperadas = [
        "EAN", "NOMEONCLICK", "NOMEE-COMMERCE", "TIPODEPRODUTO",
        "EMBALTURA", "EMBLARGURA", "EMBCOMPRIMENTO", "VOLUMES",
        "EANCOMPONENTES", "MARCA", "CUSTO", "DE", "POR", "FORNECEDOR",
        "OUTROS", "IPI", "FRETE", "NCM", "CODFORN", "CATEGORIA", "GRUPO",
        "COMPLEMENTO", "DISPONIBILIDADEWEB", "DESCRICAOHTML", "PESOBRUTO",
        "PESOLIQUIDO", "VOLPESOBRUTO", "VOLPESOLIQ", "VOLLARGURA",
        "VOLALTURA", "VOLCOMPRIMENTO", "CATEGORIAPRINCIPALTRAY",
        "CATEGORIAPRINCIPALJET", "NIVELADICIONAL1JET", "CUSTOTOTAL"
    ]

    colunas_faltando = [col for col in colunas_esperadas if col not in df.columns]
    if colunas_faltando:
        raise Exception(f"Planilha Online faltando as seguintes colunas: {', '.join(colunas_faltando)}")

    # Remove linhas onde colunas têm o nome da coluna (ex: linha com "EAN" na coluna EAN)
    df = df[~df.apply(lambda row: any(
        str(row[c]).strip().upper() == c.upper() for c in df.columns if pd.notna(row[c])
    ), axis=1)]

    # ============================================
    # PROCESSAMENTO DOS DADOS
    # ============================================
    
    logs = []
    dados_sheets = {
        "PRODUTO": [],
        "PRECO": [],
        "LOJA WEB": [],
        "KIT": [],
        "VOLUME": []
    }

    produto_dict = {
        str(row["EAN"]).strip(): row["NOMEONCLICK"] if pd.notna(row["NOMEONCLICK"]) else "Nome Desconhecido"
        for _, row in df.iterrows()
    }

    marcas_cadastradas = set()
    data_atual = datetime.now()
    data_formatada = data_atual.strftime("%d/%m/%Y")
    data_mais_20_anos = data_atual.replace(year=data_atual.year + 30)
    data_formatada_mais_20_anos = data_mais_20_anos.strftime("%d/%m/%Y")

    for idx, row in df.iterrows():
        ean = str(row["EAN"]).strip()
        nome_onclick = row["NOMEONCLICK"]
        nome_ecommerce = row["NOMEE-COMMERCE"]
        logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] Anuncio Criado: {ean} - {nome_ecommerce}")
        
        tipo_produto = str(row["TIPODEPRODUTO"]).strip().upper() if pd.notna(row["TIPODEPRODUTO"]) else ""
        altura = row["EMBALTURA"]
        largura = row["EMBLARGURA"]
        comprimento = row["EMBCOMPRIMENTO"]

        # Corrige o erro de conversão de "VOLUMES"
        try:
            valor_volumes = ajustar_decimal(row["VOLUMES"])
            volumes = int(float(valor_volumes)) if valor_volumes else 1
        except Exception:
            volumes = 1

        componentes = row["EANCOMPONENTES"]
        marca = row["MARCA"]
        custo = limpar_moeda(row["CUSTO"])
        preco_venda = limpar_moeda(row["DE"])
        preco_promo = limpar_moeda(row["POR"])
        fornecedor = row["FORNECEDOR"]
        outros = row["OUTROS"]
        ipi = limpar_moeda(row["IPI"])
        frete = limpar_moeda(row["FRETE"])
        ncm = row["NCM"]
        cod_forn = row["CODFORN"]
        categoria = row["CATEGORIA"]
        grupo = row["GRUPO"]

        marca_web = nome_ecommerce.split("-")[-1].strip() if isinstance(nome_ecommerce, str) and "-" in nome_ecommerce else ""
        complemento = row["COMPLEMENTO"]
        disponibilidade_web = row["DISPONIBILIDADEWEB"]
        descricao_html = row["DESCRICAOHTML"]
        peso_bruto = row["PESOBRUTO"]
        peso_liquido = row["PESOLIQUIDO"]
        vol_peso_bruto = row["VOLPESOBRUTO"]
        vol_peso_liquido = row["VOLPESOLIQ"]
        vol_largura = row["VOLLARGURA"]
        vol_altura = row["VOLALTURA"]
        vol_comprimento = row["VOLCOMPRIMENTO"]
        tipo_produto_valor = 0 if tipo_produto == "ACABADO" else 2
        nome_reduzido = nome_onclick[:25] if isinstance(nome_onclick, str) else ""

        marcas_cadastradas.add(marca)

        dados_sheets["PRODUTO"].append([
            ean, cod_forn, tipo_produto_valor, nome_onclick, nome_reduzido, nome_onclick, nome_onclick, None,
            marca, categoria, grupo, None, None, complemento, None, None, "F", "F", "F", None, volumes,
            peso_bruto, peso_liquido, largura, altura, comprimento, None, 90, 1000,
            disponibilidade_web, "F", "F", ncm, None, "0", "T", "F", "F", "NAO", nome_ecommerce, marca_web,
            "90 dias após o recebimento do produto", disponibilidade_web, descricao_html, "F", "F"
        ])

        # Processamento de KIT
        if tipo_produto == "KIT" and pd.notna(componentes):
            componentes_list = str(componentes).split("/")
            componentes_contados = {}
            for comp in componentes_list:
                comp_ean = comp.strip()
                componentes_contados[comp_ean] = componentes_contados.get(comp_ean, 0) + 1
            for comp_ean, quantidade in componentes_contados.items():
                nome_componente = produto_dict.get(comp_ean, "Desconhecido")
                dados_sheets["KIT"].append([ean, comp_ean, nome_componente, str(quantidade), "", "0"])

        # Processamento de VOLUMES
        for i in range(volumes):
            if volumes == 1:
                dados_sheets["VOLUME"].append([
                    ean, nome_onclick, peso_bruto, peso_liquido, largura, altura, "",
                    comprimento, "", "BOX", "T", i + 1
                ])
            else:
                dados_sheets["VOLUME"].append([
                    ean, nome_onclick, vol_peso_bruto, vol_peso_liquido, vol_largura, vol_altura, "",
                    vol_comprimento, "", "BOX", "T", i + 1
                ])

        dados_sheets["PRECO"].append([
            ean, fornecedor, custo, ipi, "", frete, limpar_moeda(row["CUSTOTOTAL"]), preco_venda, preco_promo, preco_promo,
            data_formatada, data_formatada_mais_20_anos, "", "F"
        ])

        dados_sheets["LOJA WEB"].append([
            ean, "", "", "", row["CATEGORIAPRINCIPALTRAY"], "", "", "", "T", "T", "", "", "",
            row["CATEGORIAPRINCIPALJET"], row["NIVELADICIONAL1JET"], "", "", "T", "T"
        ])

        produtos_processados.append({
            'ean': ean,
            'nome': nome_ecommerce,
            'status': 'sucesso',
            'data_processamento': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })

    # ============================================
    # SALVA PLANILHA DE SAÍDA
    # ============================================
    
    print("💾 Salvando planilha de saída...")
    
    wb = load_workbook(planilha_destino)
    estilo_invisivel = NamedStyle(name="aspas_invisiveis")
    estilo_invisivel.number_format = '@'
    wb.add_named_style(estilo_invisivel)

    ws_tipo_importacao = wb["Tipo Importacao"]
    validacoes_tipo_importacao = copiar_validacoes(ws_tipo_importacao)

    for sheet_name, data in dados_sheets.items():
        ws = wb[sheet_name]
        start_row = 3 if sheet_name == "PRODUTO" else 2

        # Limpa linhas existentes
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # Preenche novos dados
        for i, row_data in enumerate(data, start=start_row):
            for j, value in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=value)

    # Aplica estilo para aspas invisíveis
    for sheet_name in dados_sheets.keys():
        ws = wb[sheet_name]
        start_row = 3 if sheet_name == "PRODUTO" else 2
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                if cell.value == "'":
                    cell.style = "aspas_invisiveis"
                    cell.value = "'"
                    cell.fill = None

    # Define nome do arquivo de saída
    marcas_validas = [m for m in marcas_cadastradas if isinstance(m, str) and m.strip() and m.strip().upper() != "MARCA"]
    marca_unica = next(iter(marcas_validas)) if marcas_validas else "saida"
    marca_sanitizada = sanitize_filename(marca_unica)
    novo_nome = f"Template_Produtos_Mpozenato_Cadastro_{marca_sanitizada}.xlsx"
    
    # Garante que a pasta uploads existe
    os.makedirs("uploads", exist_ok=True)
    caminho_saida = os.path.join("uploads", novo_nome)

    wb.save(caminho_saida)

    # Reabre para reaplicar validações da aba "Tipo Importacao"
    wb = load_workbook(caminho_saida)
    ws_tipo_importacao = wb["Tipo Importacao"]
    reaplicar_validacoes(ws_tipo_importacao, validacoes_tipo_importacao)
    wb.save(caminho_saida)

    # Salva log
    with open('uploads/logs_processamento.txt', 'w', encoding='utf-8') as f:
        for linha in logs:
            f.write(linha + '\n')

    fim = datetime.now()
    duracao = (fim - inicio).total_seconds()
    qtd_produtos = len(df)

    print(f"✅ Processamento concluído! {qtd_produtos} produtos processados em {duracao:.2f} segundos")
    print(f"📁 Arquivo gerado: {caminho_saida}")

    return caminho_saida, qtd_produtos, duracao, produtos_processados