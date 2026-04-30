"""
faltando_simples.py
===================
Mostra APENAS as características faltando para cada MLB
Baseado no endpoint de catálogo que funcionou.

Uso:
    python faltando_simples.py [arquivo_entrada.xlsx]
"""

import sys
import json
import requests
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from app import app
from token_manager_secure import ml_token_manager

BASE = "https://api.mercadolibre.com"


def get_headers():
    token = ml_token_manager.get_valid_token()
    return {"Authorization": f"Bearer {token}"}


def consultar_catalog_quality(mlb):
    """Endpoint que realmente funciona para características faltando"""
    url = f"{BASE}/catalog_quality/status?item_id={mlb}&v=3"
    try:
        response = requests.get(url, headers=get_headers(), timeout=15)
        if response.status_code == 200:
            return response.json()
        return None
    except:
        return None


def consultar_health_actions(mlb):
    """Tenta pegar ações (campanhas, etc)"""
    url = f"{BASE}/items/{mlb}/health/actions"
    try:
        response = requests.get(url, headers=get_headers(), timeout=15)
        if response.status_code == 200:
            return response.json()
        return None
    except:
        return None


def extrair_caracteristicas_faltando(catalog_data):
    """Extrai APENAS as características que faltam"""
    if not catalog_data:
        return []
    
    faltando = []
    
    adoption = catalog_data.get('adoption_status', {})
    
    # Pega atributos faltando de diferentes seções
    for section in ['all', 'required', 'pi', 'ft']:
        section_data = adoption.get(section, {})
        if section_data:
            missing = section_data.get('missing_attributes', [])
            if missing and isinstance(missing, list):
                faltando.extend(missing)
    
    # Remove duplicatas
    faltando = list(set(faltando))
    
    # Mapeia códigos para nomes legíveis em português
    mapeamento = {
        'GTIN': 'Código de barras (GTIN/EAN/UPC)',
        'BRAND': 'Marca do produto',
        'MODEL': 'Modelo do produto',
        'COLOR': 'Cor do produto',
        'SIZE': 'Tamanho do produto',
        'GENDER': 'Gênero',
        'MATERIAL': 'Material',
        'WEIGHT': 'Peso',
        'WIDTH': 'Largura',
        'HEIGHT': 'Altura',
        'DEPTH': 'Profundidade',
        'AGE_GROUP': 'Faixa etária',
        'RELEASE_YEAR': 'Ano de lançamento',
        'RELEASE_SEASON': 'Estação',
        'CHARACTER': 'Personagem',
        'FOOTWEAR_TYPE': 'Tipo de calçado',
        'FOOTWEAR_MATERIAL': 'Material do calçado',
        'IS_SUITABLE_FOR_FIRST_STEPS': 'Primeiros passos',
        'VOLUME': 'Volume',
        'CAPACITY': 'Capacidade',
        'POWER': 'Potência',
        'CONDITION': 'Condição do produto',
        'WARRANTY': 'Garantia',
        'ORIGIN': 'Origem',
        'CERTIFICATION': 'Certificação',
    }
    
    # Converte para nomes legíveis
    resultado = []
    for attr in faltando:
        nome = mapeamento.get(attr, attr)
        resultado.append(nome)
    
    return resultado


def extrair_acoes(actions_data):
    """Extrai ações recomendadas (campanhas, clips, etc)"""
    if not actions_data:
        return []
    
    acoes = []
    for action in actions_data.get('actions', []):
        name = action.get('name', '')
        if name:
            acoes.append(name)
    
    return acoes


def processar_mlb(mlb):
    """Processa um MLB e retorna o que falta"""
    print(f"  🔍 Analisando {mlb}...", end=" ")
    
    resultado = {
        'mlb': mlb,
        'caracteristicas_faltando': [],
        'acoes_necessarias': [],
        'erro': None
    }
    
    # Consulta características do catálogo (essa funciona)
    catalog_data = consultar_catalog_quality(mlb)
    if catalog_data:
        caracteristicas = extrair_caracteristicas_faltando(catalog_data)
        resultado['caracteristicas_faltando'] = caracteristicas
        print(f"✅ {len(caracteristicas)} características faltando")
    else:
        print(f"⚠️ Não foi possível consultar")
        resultado['erro'] = "Catalogo não disponível"
        return resultado
    
    # Tenta consultar ações (opcional, pode não funcionar)
    actions_data = consultar_health_actions(mlb)
    if actions_data:
        acoes = extrair_acoes(actions_data)
        resultado['acoes_necessarias'] = acoes
    
    return resultado


def ler_mlbs_do_excel(arquivo):
    """Lê MLBs do Excel (primeira coluna)"""
    try:
        wb = load_workbook(arquivo)
        ws = wb.active
        
        mlbs = []
        for row in range(2, ws.max_row + 1):
            mlb_cell = ws.cell(row, 1).value
            if mlb_cell:
                mlb = str(mlb_cell).strip().upper()
                # Remove MLB se já tiver
                mlb = mlb.replace('MLB', '').strip()
                mlb = f'MLB{mlb}'
                mlbs.append(mlb)
        
        return mlbs
    except Exception as e:
        print(f"❌ Erro ao ler Excel: {e}")
        return []


def criar_excel_resultado(resultados, arquivo_saida):
    """Cria Excel apenas com características faltando"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Caracteristicas_Faltando"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Cabeçalhos
    cabecalhos = ['MLB', 'Total Faltando', 'Características que Faltam', 'Ações Recomendadas']
    
    for col, header in enumerate(cabecalhos, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Preencher dados
    for row_idx, r in enumerate(resultados, 2):
        caracteristicas = r.get('caracteristicas_faltando', [])
        acoes = r.get('acoes_necessarias', [])
        
        ws.cell(row_idx, 1, r['mlb'])
        ws.cell(row_idx, 2, len(caracteristicas))
        
        # Características faltando (uma por linha)
        cell_carac = ws.cell(row_idx, 3, "\n".join(caracteristicas) if caracteristicas else "Nenhuma característica faltando")
        cell_carac.alignment = left_alignment
        
        # Ações
        cell_acoes = ws.cell(row_idx, 4, "\n".join(acoes) if acoes else "Nenhuma ação específica")
        cell_acoes.alignment = left_alignment
    
    # Ajustar largura
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    
    # Ajustar altura das linhas
    for row in range(2, len(resultados) + 2):
        ws.row_dimensions[row].height = max(40, len(str(ws.cell(row, 3).value or '')) / 4)
    
    wb.save(arquivo_saida)
    return arquivo_saida


def main():
    arquivo_entrada = sys.argv[1] if len(sys.argv) > 1 else "mlbs_entrada.xlsx"
    
    print(f"\n{'='*70}")
    print(f"  VERIFICADOR DE CARACTERÍSTICAS FALTANDO")
    print(f"{'='*70}")
    print(f"📁 Arquivo: {arquivo_entrada}\n")
    
    if not Path(arquivo_entrada).exists():
        print(f"❌ Arquivo '{arquivo_entrada}' não encontrado!")
        print(f"\nCrie um Excel com uma coluna 'MLB' na primeira coluna")
        sys.exit(1)
    
    # Testar token
    print("🔍 Verificando autenticação...")
    try:
        token = ml_token_manager.get_valid_token()
        headers = get_headers()
        test_response = requests.get(f"{BASE}/users/me", headers=headers, timeout=10)
        if test_response.status_code == 200:
            user = test_response.json()
            print(f"✅ Autenticado: {user.get('nickname', 'Usuário')}\n")
        else:
            print(f"❌ Token inválido!")
            sys.exit(1)
    except Exception as e:
        print(f"❌ Erro de autenticação: {e}")
        sys.exit(1)
    
    # Ler MLBs
    mlbs = ler_mlbs_do_excel(arquivo_entrada)
    
    if not mlbs:
        print(f"❌ Nenhum MLB encontrado!")
        sys.exit(1)
    
    print(f"📋 Encontrados {len(mlbs)} MLBs\n")
    
    # Processar
    resultados = []
    
    with app.app_context():
        for i, mlb in enumerate(mlbs, 1):
            print(f"[{i}/{len(mlbs)}]", end=" ")
            resultado = processar_mlb(mlb)
            resultados.append(resultado)
            
            # Mostrar características se tiver
            if resultado['caracteristicas_faltando']:
                print(f"     📝 Faltam: {', '.join(resultado['caracteristicas_faltando'][:3])}")
                if len(resultado['caracteristicas_faltando']) > 3:
                    print(f"        ... e mais {len(resultado['caracteristicas_faltando']) - 3}")
    
    # Gerar Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo_saida = f"caracteristicas_faltando_{timestamp}.xlsx"
    
    print(f"\n📊 Gerando relatório...")
    criar_excel_resultado(resultados, arquivo_saida)
    
    # Resumo
    print(f"\n{'='*70}")
    print("  RESUMO")
    print(f"{'='*70}")
    
    total = len(resultados)
    com_caracteristicas = sum(1 for r in resultados if r['caracteristicas_faltando'])
    total_caracteristicas = sum(len(r['caracteristicas_faltando']) for r in resultados)
    
    print(f"\n✅ Relatório: {arquivo_saida}")
    print(f"\n📊 MLBs analisados: {total}")
    print(f"⚠️ MLBs com características faltando: {com_caracteristicas}")
    print(f"📝 Total de características faltando: {total_caracteristicas}")
    
    # Listar os piores
    print(f"\n📋 MLBs que precisam de mais atenção:")
    piores = sorted(resultados, key=lambda x: len(x['caracteristicas_faltando']), reverse=True)[:5]
    for r in piores:
        if r['caracteristicas_faltando']:
            print(f"   • {r['mlb']}: {len(r['caracteristicas_faltando'])} características faltando")
    
    print(f"\n{'='*70}\n")


if __name__ == "__main__":
    main()