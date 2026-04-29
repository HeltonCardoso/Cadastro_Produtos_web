"""
excel_quality_checker_completo.py
==================================
Consulta MÚLTIPLOS endpoints do Mercado Livre para ter TODAS as informações de qualidade:
- Performance (score básico)
- Diagnóstico completo do anúncio
- Qualidade do anúncio (campos, características)
- Recomendações de catálogo
- Necessidade de campanhas

Uso:
    python excel_quality_checker_completo.py [arquivo_entrada.xlsx]
"""

import sys
import json
import requests
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from app import app
from token_manager_secure import ml_token_manager

BASE = "https://api.mercadolibre.com"


def get_headers():
    token = ml_token_manager.get_valid_token()
    return {"Authorization": f"Bearer {token}"}


def consultar_item(mlb):
    """Consulta dados básicos do item"""
    url = f"{BASE}/items/{mlb}"
    try:
        response = requests.get(url, headers=get_headers(), timeout=15)
        if response.status_code == 200:
            return response.json()
        return None
    except:
        return None


def consultar_performance(mlb):
    """Consulta performance/qualidade do anúncio"""
    url = f"{BASE}/item/{mlb}/performance"
    try:
        response = requests.get(url, headers=get_headers(), timeout=15)
        if response.status_code == 200:
            return response.json()
        return None
    except:
        return None


def consultar_qualidade_anuncio(mlb):
    """Endpoint de qualidade do anúncio (mais detalhado)"""
    url = f"{BASE}/items/{mlb}/quality"
    try:
        response = requests.get(url, headers=get_headers(), timeout=15)
        if response.status_code == 200:
            return response.json()
        return None
    except:
        return None


def consultar_recomendacoes(mlb):
    """Recomendações para melhorar o anúncio"""
    url = f"{BASE}/items/{mlb}/recommendations"
    try:
        response = requests.get(url, headers=get_headers(), timeout=15)
        if response.status_code == 200:
            return response.json()
        return None
    except:
        return None


def consultar_catalog_requirements(mlb):
    """Verifica se precisa criar catálogo"""
    url = f"{BASE}/items/{mlb}/catalog_requirements"
    try:
        response = requests.get(url, headers=get_headers(), timeout=15)
        if response.status_code == 200:
            return response.json()
        return None
    except:
        return None


def check_campaign_needs(item_data):
    """Analisa se precisa de campanha baseado nos dados do item"""
    if not item_data:
        return []
    
    needs = []
    
    # Verificar se tem anúncios grátis disponíveis
    if item_data.get('listing_type_id') == 'free':
        needs.append("✅ Já está usando anúncio grátis")
    else:
        needs.append("💰 Considere usar anúncio grátis para economizar")
    
    # Verificar visualizações baixas (indica necessidade de campanha)
    if item_data.get('sold_quantity', 0) == 0:
        needs.append("📢 Poucas vendas - Considere campanha de impulsionamento")
    
    # Verificar classificação do anúncio
    if item_data.get('condition') == 'not_specified':
        needs.append("⚠️ Precisa especificar condição do produto (novo/usado)")
    
    return needs


def analisar_catalogo(catalog_data, item_data):
    """Analisa se precisa criar/integrar com catálogo"""
    if catalog_data:
        return {
            'precisa_catalogo': True,
            'detalhes': "Anúncio pode ser integrado ao catálogo do Mercado Livre",
            'acoes': ["Cadastrar produto no catálogo oficial"]
        }
    
    # Verificar se tem dados de catálogo no item
    if item_data:
        if not item_data.get('catalog_product_id'):
            return {
                'precisa_catalogo': True,
                'detalhes': "Produto não está vinculado ao catálogo do Mercado Livre",
                'acoes': ["Buscar produto no catálogo", "Criar ficha de catálogo se não existir"]
            }
    
    return {
        'precisa_catalogo': False,
        'detalhes': "Produto já está no catálogo",
        'acoes': []
    }


def analisar_caracteristicas(item_data):
    """Analisa características que faltam no anúncio"""
    if not item_data:
        return {'faltam': [], 'completas': 0, 'total': 0}
    
    attributes = item_data.get('attributes', [])
    total_attributes = len(attributes)
    filled_attributes = [a for a in attributes if a.get('value_name')]
    
    # Identificar atributos obrigatórios faltando
    required_missing = []
    for attr in attributes:
        if attr.get('tags') and 'required' in attr.get('tags', []):
            if not attr.get('value_name'):
                required_missing.append({
                    'nome': attr.get('name', ''),
                    'id': attr.get('id', ''),
                    'exemplo': attr.get('value_type', '')
                })
    
    return {
        'faltam': required_missing,
        'completas': len(filled_attributes),
        'total': total_attributes,
        'percentual': (len(filled_attributes) / total_attributes * 100) if total_attributes > 0 else 0
    }


def analisar_tudo(mlb):
    """Consulta TODOS os endpoints e compila o que falta"""
    
    resultado = {
        'mlb': mlb,
        'score': None,
        'nivel': None,
        'faltam_pontos': None,
        'precisa_campanha': [],
        'precisa_catalogo': {},
        'caracteristicas_faltando': [],
        'pendencias_gerais': [],
        'acoes_prioritarias': [],
        'recomendacoes': [],
        'erros': []
    }
    
    # 1. Consultar performance (score básico)
    print(f"  📊 Consultando performance...", end=" ")
    performance = consultar_performance(mlb)
    if performance:
        resultado['score'] = performance.get('score', 0)
        resultado['nivel'] = performance.get('level', 'unknown')
        resultado['faltam_pontos'] = max(0, 80 - resultado['score'])
        
        # Extrair pendências do performance
        for bucket in performance.get('buckets', []):
            for var in bucket.get('variables', []):
                for rule in var.get('rules', []):
                    if rule.get('status') in ['bad', 'medium', 'critical']:
                        wording = rule.get('wordings', {})
                        resultado['pendencias_gerais'].append({
                            'categoria': bucket.get('title', ''),
                            'item': var.get('title', ''),
                            'problema': wording.get('label', ''),
                            'solucao': wording.get('action', '')
                        })
        print("✅" if performance else "❌")
    else:
        print("❌")
        resultado['erros'].append("Não foi possível consultar performance")
    
    # 2. Consultar dados do item
    print(f"  📝 Consultando dados do item...", end=" ")
    item_data = consultar_item(mlb)
    if item_data:
        # Analisar características
        caracteristicas = analisar_caracteristicas(item_data)
        resultado['caracteristicas_faltando'] = caracteristicas['faltam']
        
        # Verificar necessidade de campanha
        resultado['precisa_campanha'] = check_campaign_needs(item_data)
        
        print("✅")
    else:
        print("❌")
        resultado['erros'].append("Não foi possível consultar dados do item")
    
    # 3. Consultar qualidade do anúncio
    print(f"  ⭐ Consultando qualidade detalhada...", end=" ")
    qualidade = consultar_qualidade_anuncio(mlb)
    if qualidade:
        # Extrair recomendações específicas
        for rec in qualidade.get('recommendations', []):
            resultado['recomendacoes'].append({
                'tipo': rec.get('type', ''),
                'descricao': rec.get('description', ''),
                'prioridade': rec.get('priority', '')
            })
        print("✅")
    else:
        print("⚠️")
    
    # 4. Verificar necessidade de catálogo
    print(f"  📦 Verificando catálogo...", end=" ")
    catalog_data = consultar_catalog_requirements(mlb)
    resultado['precisa_catalogo'] = analisar_catalogo(catalog_data, item_data)
    print("✅" if resultado['precisa_catalogo'] else "❌")
    
    # 5. Consultar recomendações
    print(f"  💡 Buscando recomendações...", end=" ")
    recomendacoes_api = consultar_recomendacoes(mlb)
    if recomendacoes_api:
        for rec in recomendacoes_api.get('recommendations', []):
            resultado['acoes_prioritarias'].append({
                'acao': rec.get('action', ''),
                'motivo': rec.get('reason', ''),
                'impacto': rec.get('impact', '')
            })
        print("✅")
    else:
        print("⚠️")
    
    return resultado


def criar_excel_completo(arquivo_entrada, resultados):
    """Cria Excel com todas as informações detalhadas"""
    
    wb = Workbook()
    
    # Planilha 1: Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo_Geral"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Cabeçalhos do resumo
    cabecalhos_resumo = [
        'MLB', 'Score', 'Nível', 'Faltam Pontos',
        'Precisa Catálogo?', 'Precisa Campanha?',
        'Total Características', 'Características Completas',
        'Total Pendências', 'Status'
    ]
    
    for col, header in enumerate(cabecalhos_resumo, 1):
        cell = ws_resumo.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Preencher resumo
    for row_idx, r in enumerate(resultados, 2):
        analise = r['analise']
        
        precisa_catalogo = "✅ Sim" if analise.get('precisa_catalogo', {}).get('precisa_catalogo') else "❌ Não"
        precisa_campanha = "📢 Sim" if analise.get('precisa_campanha') else "✅ Não"
        
        total_carac = analise.get('caracteristicas_faltando', [])
        total_carac_count = len(total_carac) if total_carac else 0
        
        # Status
        if analise.get('erros'):
            status = "❌ ERRO"
            status_fill = critical_fill
        elif analise.get('faltam_pontos', 100) == 0:
            status = "✅ OTIMIZADO"
            status_fill = success_fill
        elif analise.get('faltam_pontos', 100) < 20:
            status = "⚠️ QUASE LÁ"
            status_fill = warning_fill
        else:
            status = "🔴 PRECISA MELHORAR"
            status_fill = critical_fill
        
        valores = [
            r['mlb'],
            analise.get('score', 'N/A'),
            analise.get('nivel', 'N/A'),
            analise.get('faltam_pontos', 'N/A'),
            precisa_catalogo,
            precisa_campanha,
            analise.get('caracteristicas_faltando', []) if isinstance(analise.get('caracteristicas_faltando'), list) else 'N/A',
            'N/A',  # Placeholder
            len(analise.get('pendencias_gerais', [])),
            status
        ]
        
        for col, valor in enumerate(valores, 1):
            cell = ws_resumo.cell(row_idx, col, valor)
            cell.alignment = left_alignment
            if col == 10:  # Coluna Status
                cell.fill = status_fill
                cell.font = Font(bold=True)
    
    # Ajustar colunas do resumo
    for col in range(1, len(cabecalhos_resumo) + 1):
        ws_resumo.column_dimensions[chr(64 + col)].width = 20
    
    # Planilha 2: Pendências detalhadas
    ws_pendencias = wb.create_sheet("Pendencias_Detalhadas")
    
    cab_pendencias = ['MLB', 'Categoria', 'Item', 'Problema', 'Solução']
    for col, header in enumerate(cab_pendencias, 1):
        cell = ws_pendencias.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
    
    row_idx = 2
    for r in resultados:
        for pendencia in r['analise'].get('pendencias_gerais', []):
            ws_pendencias.cell(row_idx, 1, r['mlb'])
            ws_pendencias.cell(row_idx, 2, pendencia.get('categoria', ''))
            ws_pendencias.cell(row_idx, 3, pendencia.get('item', ''))
            ws_pendencias.cell(row_idx, 4, pendencia.get('problema', ''))
            ws_pendencias.cell(row_idx, 5, pendencia.get('solucao', ''))
            row_idx += 1
    
    # Planilha 3: Características faltando
    ws_carac = wb.create_sheet("Caracteristicas_Faltando")
    
    cab_carac = ['MLB', 'Característica', 'Exemplo', 'Ação Necessária']
    for col, header in enumerate(cab_carac, 1):
        cell = ws_carac.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
    
    row_idx = 2
    for r in resultados:
        for carac in r['analise'].get('caracteristicas_faltando', []):
            ws_carac.cell(row_idx, 1, r['mlb'])
            ws_carac.cell(row_idx, 2, carac.get('nome', ''))
            ws_carac.cell(row_idx, 3, carac.get('exemplo', ''))
            ws_carac.cell(row_idx, 4, f"Preencher característica {carac.get('nome', '')}")
            row_idx += 1
    
    # Planilha 4: Ações prioritárias
    ws_acoes = wb.create_sheet("Acoes_Prioritarias")
    
    cab_acoes = ['MLB', 'Ação', 'Motivo', 'Impacto Esperado']
    for col, header in enumerate(cab_acoes, 1):
        cell = ws_acoes.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
    
    row_idx = 2
    for r in resultados:
        # Ações de campanha
        for campanha in r['analise'].get('precisa_campanha', []):
            ws_acoes.cell(row_idx, 1, r['mlb'])
            ws_acoes.cell(row_idx, 2, campanha)
            ws_acoes.cell(row_idx, 3, "Melhorar visibilidade")
            ws_acoes.cell(row_idx, 4, "Aumentar vendas")
            row_idx += 1
        
        # Ações de catálogo
        if r['analise'].get('precisa_catalogo', {}).get('precisa_catalogo'):
            for acao in r['analise']['precisa_catalogo'].get('acoes', []):
                ws_acoes.cell(row_idx, 1, r['mlb'])
                ws_acoes.cell(row_idx, 2, acao)
                ws_acoes.cell(row_idx, 3, r['analise']['precisa_catalogo'].get('detalhes', ''))
                ws_acoes.cell(row_idx, 4, "Melhorar qualidade e visibilidade")
                row_idx += 1
    
    return wb


def main():
    arquivo_entrada = sys.argv[1] if len(sys.argv) > 1 else "mlbs_entrada.xlsx"
    
    print(f"\n{'='*80}")
    print(f"  DIAGNÓSTICO COMPLETO DE QUALIDADE - MLBs")
    print(f"{'='*80}")
    print(f"📁 Arquivo: {arquivo_entrada}\n")
    
    # Verificar arquivo
    if not Path(arquivo_entrada).exists():
        print(f"❌ Arquivo '{arquivo_entrada}' não encontrado!")
        sys.exit(1)
    
    # Ler MLBs do Excel
    wb_in = load_workbook(arquivo_entrada)
    ws_in = wb_in.active
    
    print("📖 Lendo MLBs...")
    mlbs = []
    for row in range(2, ws_in.max_row + 1):
        mlb_cell = ws_in.cell(row, 1).value
        if mlb_cell:
            mlb = str(mlb_cell).strip().upper()
            if not mlb.startswith('MLB'):
                mlb = f'MLB{mlb}'
            mlbs.append(mlb)
    
    print(f"✅ Encontrados {len(mlbs)} MLBs\n")
    
    # Processar cada MLB
    resultados = []
    
    with app.app_context():
        for i, mlb in enumerate(mlbs, 1):
            print(f"\n[{i}/{len(mlbs)}] Analisando {mlb}:")
            analise = analisar_tudo(mlb)
            resultados.append({
                'mlb': mlb,
                'analise': analise
            })
            print(f"  {'='*50}")
    
    # Criar Excel com resultados
    print(f"\n📊 Gerando relatório completo...")
    wb_resultado = criar_excel_completo(arquivo_entrada, resultados)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo_saida = f"diagnostico_completo_{timestamp}.xlsx"
    wb_resultado.save(arquivo_saida)
    
    # Resumo final
    print(f"\n{'='*80}")
    print("  RESUMO DO DIAGNÓSTICO")
    print(f"{'='*80}")
    
    precisa_catalogo = sum(1 for r in resultados if r['analise'].get('precisa_catalogo', {}).get('precisa_catalogo'))
    precisa_campanha = sum(1 for r in resultados if r['analise'].get('precisa_campanha'))
    tem_erro = sum(1 for r in resultados if r['analise'].get('erros'))
    
    print(f"\n📊 Total analisado: {len(resultados)}")
    print(f"📦 Precisam criar/integrar catálogo: {precisa_catalogo}")
    print(f"📢 Precisam de campanha de impulsionamento: {precisa_campanha}")
    print(f"❌ Com erros de consulta: {tem_erro}")
    
    print(f"\n✅ Relatório salvo em: {arquivo_saida}")
    print(f"\n📋 O Excel contém 4 abas:")
    print(f"   1. Resumo_Geral - Visão geral de todos MLBs")
    print(f"   2. Pendencias_Detalhadas - Lista completa do que falta")
    print(f"   3. Caracteristicas_Faltando - Atributos não preenchidos")
    print(f"   4. Acoes_Prioritarias - O que fazer primeiro (campanha, catálogo, etc)")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    main()